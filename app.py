# -*- coding: utf-8 -*-
"""
AHVT - نظام إدارة الأداء والأثر التطوعي
نسخة Flask + SQLite (تعمل على سيرفر ويمكن للفريق كله الوصول إليها)
"""
import os
import io
import json
import sqlite3
import uuid
import shutil
import qrcode
from PIL import Image

# Compact reporting/export stack
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT, TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from openpyxl import Workbook
from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None

from flask import (
    Flask, request, redirect, url_for, render_template, session,
    g, flash, jsonify, send_file, send_from_directory
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Persistent storage: on Render/production set HIKMA_DATA_DIR to the mounted
# persistent disk (for example /data). Locally it defaults to the project folder.
DATA_DIR = os.environ.get("HIKMA_DATA_DIR", BASE_DIR)
os.makedirs(DATA_DIR, exist_ok=True)
DB_PATH = os.path.join(DATA_DIR, "hikma.db")

app = Flask(__name__)
app.secret_key = os.environ.get("HIKMA_SECRET_KEY", "hikma-impact-dev-secret-change-me")
app.jinja_env.filters["fromjson"] = lambda value: json.loads(value or "{}")
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

@app.route("/uploads/<path:filename>")
def uploaded_asset(filename):
    """Serve persisted user uploads from the configured data directory."""
    return send_from_directory(UPLOAD_DIR, filename, max_age=3600)

def save_image_as_jpg(file_storage, prefix="img"):
    if not file_storage or not file_storage.filename:
        return ""
    try:
        img = Image.open(file_storage.stream).convert("RGB")
        name = f"{prefix}-{uuid.uuid4().hex[:12]}.jpg"
        img.save(os.path.join(UPLOAD_DIR, name), "JPEG", quality=92, optimize=True)
        return f"uploads/{name}"
    except Exception:
        return ""

def save_uploaded_video(file_storage, prefix="video"):
    if not file_storage or not file_storage.filename:
        return ""
    ext = os.path.splitext(secure_filename(file_storage.filename))[1].lower()
    if ext not in {".mp4", ".webm", ".mov", ".m4v", ".avi"}:
        return ""
    name = f"{prefix}-{uuid.uuid4().hex[:12]}{ext}"
    file_storage.save(os.path.join(UPLOAD_DIR, name))
    return f"uploads/{name}"

def save_uploaded_audio(file_storage, prefix="audio"):
    if not file_storage or not file_storage.filename:
        return ""
    ext = os.path.splitext(secure_filename(file_storage.filename))[1].lower()
    if ext not in {".mp3", ".wav", ".m4a", ".aac", ".ogg"}:
        return ""
    name = f"{prefix}-{uuid.uuid4().hex[:12]}{ext}"
    file_storage.save(os.path.join(UPLOAD_DIR, name))
    return f"uploads/{name}"

def save_multiple_attachments(files, parent_type, parent_id, db, title_prefix=""):
    saved=[]
    for f in files or []:
        if not f or not f.filename:
            continue
        ext=os.path.splitext(secure_filename(f.filename))[1].lower()
        if ext in {".jpg",".jpeg",".png",".webp",".gif"}:
            path=save_image_as_jpg(f, f"{parent_type}-asset"); media_type="image"
        elif ext in {".mp4",".webm",".mov",".m4v",".avi"}:
            path=save_uploaded_video(f, f"{parent_type}-asset"); media_type="video"
        else:
            continue
        if path:
            db.execute("INSERT INTO attachments(id,parent_type,parent_id,file_path,media_type,title,sort_order,created_at) VALUES(?,?,?,?,?,?,?,?)", (uid("att"),parent_type,parent_id,path,media_type,title_prefix or secure_filename(f.filename),len(saved),now_iso()))
            saved.append(path)
    return saved

def attachments_for(db, parent_type, parent_id):
    return db.execute("SELECT * FROM attachments WHERE parent_type=? AND parent_id=? ORDER BY sort_order,created_at", (parent_type,parent_id)).fetchall()

CRITERIA_KEYS = ["attendance", "taskCompletion", "initiativeParticipation",
                  "commitment", "teamwork", "creativity"]
CRITERIA_LABELS = {
    "attendance": "الحضور", "taskCompletion": "تنفيذ المهام",
    "initiativeParticipation": "المشاركة بالمبادرات", "commitment": "الالتزام",
    "teamwork": "العمل الجماعي", "creativity": "الإبداع"
}
DEFAULT_WEIGHTS = {"attendance": 20, "taskCompletion": 25, "initiativeParticipation": 20,
                    "commitment": 15, "teamwork": 10, "creativity": 10}
DEFAULT_POINTS = {"attendance": 10, "task": 10, "leader": 25, "participation": 15, "excellent": 20}


# ============================================================ DB ============================================================
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH, timeout=30)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
        g.db.execute("PRAGMA busy_timeout = 30000")
        # WAL prevents normal reads from blocking writes and is much safer when
        # Gunicorn has more than one worker touching the SQLite database.
        g.db.execute("PRAGMA journal_mode = WAL")
        g.db.execute("PRAGMA synchronous = NORMAL")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DB_PATH, timeout=30)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys = ON")
    db.execute("PRAGMA busy_timeout = 30000")
    db.execute("PRAGMA journal_mode = WAL")
    db.execute("PRAGMA synchronous = NORMAL")
    db.executescript("""
    CREATE TABLE IF NOT EXISTS users(
        id TEXT PRIMARY KEY, name TEXT, email TEXT UNIQUE, username TEXT, password_hash TEXT,
        role TEXT, active INTEGER DEFAULT 1, must_set_password INTEGER DEFAULT 0, last_login TEXT, created_at TEXT);
    CREATE TABLE IF NOT EXISTS members(
        id TEXT PRIMARY KEY, name TEXT, email TEXT, phone TEXT, committee TEXT,
        position TEXT, join_date TEXT, status TEXT, notes TEXT, created_at TEXT, photo TEXT);
    CREATE TABLE IF NOT EXISTS news(
        id TEXT PRIMARY KEY, title TEXT, slug TEXT UNIQUE, excerpt TEXT, content TEXT,
        cover_image TEXT, category TEXT, author TEXT, status TEXT, published_at TEXT, created_at TEXT);
    CREATE TABLE IF NOT EXISTS pages(
        id TEXT PRIMARY KEY, title TEXT, slug TEXT UNIQUE, content TEXT, status TEXT,
        show_in_nav INTEGER DEFAULT 0, sort_order INTEGER DEFAULT 0, created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS administrators(
        id TEXT PRIMARY KEY, name TEXT, position TEXT, committee TEXT,
        date TEXT, responsibilities TEXT, photo TEXT);
    CREATE TABLE IF NOT EXISTS committees(
        id TEXT PRIMARY KEY, name TEXT, head TEXT, description TEXT);
    CREATE TABLE IF NOT EXISTS initiatives(
        id TEXT PRIMARY KEY, name TEXT, date TEXT, location TEXT, manager TEXT,
        committee TEXT, hours REAL, status TEXT, description TEXT, goals TEXT);
    CREATE TABLE IF NOT EXISTS initiative_participants(
        initiative_id TEXT, member_id TEXT, start_time TEXT, end_time TEXT, hours REAL DEFAULT 0,
        PRIMARY KEY(initiative_id, member_id));
    CREATE TABLE IF NOT EXISTS tasks(
        id TEXT PRIMARY KEY, title TEXT, assignee TEXT, deadline TEXT,
        priority TEXT, status TEXT, description TEXT, created_at TEXT);
    CREATE TABLE IF NOT EXISTS attendance(
        id TEXT PRIMARY KEY, member_id TEXT, date TEXT, status TEXT, initiative_id TEXT,
        start_time TEXT, end_time TEXT, hours REAL DEFAULT 0);
    CREATE TABLE IF NOT EXISTS evaluations(
        id TEXT PRIMARY KEY, evaluated_user_id TEXT, evaluator_id TEXT, evaluator_name TEXT,
        date TEXT, type TEXT, notes TEXT,
        c_attendance INTEGER, c_taskCompletion INTEGER, c_initiativeParticipation INTEGER,
        c_commitment INTEGER, c_teamwork INTEGER, c_creativity INTEGER);
    CREATE TABLE IF NOT EXISTS points(
        id TEXT PRIMARY KEY, member_id TEXT, value REAL, source TEXT, date TEXT);
    CREATE TABLE IF NOT EXISTS audit_logs(
        id TEXT PRIMARY KEY, action TEXT, target TEXT, by TEXT, date TEXT);
    CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY, value TEXT);
    CREATE TABLE IF NOT EXISTS notifications(
        id TEXT PRIMARY KEY, user_id TEXT, target_role TEXT, type TEXT, title TEXT, body TEXT,
        url TEXT, read_at TEXT, created_at TEXT);
    CREATE TABLE IF NOT EXISTS applications(
        id TEXT PRIMARY KEY, applicant_name TEXT, email TEXT, phone TEXT, university TEXT,
        department TEXT, stage TEXT, skills TEXT, interests TEXT, committee TEXT, motivation TEXT,
        status TEXT DEFAULT 'new', created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS events(
        id TEXT PRIMARY KEY, title TEXT, date TEXT, time TEXT, location TEXT, description TEXT,
        status TEXT DEFAULT 'upcoming', cover_image TEXT, created_at TEXT);
    CREATE TABLE IF NOT EXISTS partners(
        id TEXT PRIMARY KEY, name TEXT, description TEXT, logo TEXT, url TEXT, sort_order INTEGER DEFAULT 0, active INTEGER DEFAULT 1);
    CREATE TABLE IF NOT EXISTS media(
        id TEXT PRIMARY KEY, title TEXT, url TEXT, category TEXT, initiative_id TEXT, public INTEGER DEFAULT 1, created_at TEXT);
    CREATE TABLE IF NOT EXISTS goals(
        id TEXT PRIMARY KEY, title TEXT, target REAL DEFAULT 0, current REAL DEFAULT 0, period TEXT, status TEXT DEFAULT 'active', created_at TEXT);
    CREATE TABLE IF NOT EXISTS approvals(
        id TEXT PRIMARY KEY, kind TEXT, target_id TEXT, title TEXT, status TEXT DEFAULT 'pending',
        requested_by TEXT, reviewed_by TEXT, notes TEXT, created_at TEXT, reviewed_at TEXT);
    CREATE TABLE IF NOT EXISTS security_sessions(
        id TEXT PRIMARY KEY, user_id TEXT, ip TEXT, user_agent TEXT, created_at TEXT, last_seen TEXT);
    CREATE TABLE IF NOT EXISTS site_sections(
        id TEXT PRIMARY KEY, section_key TEXT UNIQUE, title TEXT, visible INTEGER DEFAULT 1, sort_order INTEGER DEFAULT 0,
        background TEXT DEFAULT '', accent TEXT DEFAULT '', updated_at TEXT);
    CREATE TABLE IF NOT EXISTS nav_items(
        id TEXT PRIMARY KEY, label TEXT, endpoint TEXT, url TEXT, icon TEXT DEFAULT '', visible INTEGER DEFAULT 1,
        sort_order INTEGER DEFAULT 0, is_system INTEGER DEFAULT 1, created_at TEXT);
    CREATE TABLE IF NOT EXISTS role_permissions(
        id TEXT PRIMARY KEY, role TEXT, permission TEXT, allowed INTEGER DEFAULT 1, UNIQUE(role,permission));
    CREATE TABLE IF NOT EXISTS live_sessions(
        id TEXT PRIMARY KEY, title TEXT, description TEXT, initiative_id TEXT, mode TEXT DEFAULT 'internal',
        external_url TEXT, external_platform TEXT, status TEXT DEFAULT 'live', created_by TEXT,
        started_at TEXT, ended_at TEXT, created_at TEXT);
    CREATE TABLE IF NOT EXISTS live_peers(
        id TEXT PRIMARY KEY, session_id TEXT, peer_id TEXT, role TEXT, created_at TEXT, last_seen TEXT,
        UNIQUE(session_id,peer_id));
    CREATE TABLE IF NOT EXISTS live_signals(
        id TEXT PRIMARY KEY, session_id TEXT, sender_id TEXT, target_id TEXT, kind TEXT, payload TEXT,
        created_at TEXT, delivered_at TEXT);
    CREATE TABLE IF NOT EXISTS certificates(
        id TEXT PRIMARY KEY, certificate_no TEXT UNIQUE, recipient_name TEXT, certificate_type TEXT,
        initiative_id TEXT, initiative_name TEXT, issue_date TEXT, hours REAL DEFAULT 0,
        note TEXT, issued_by TEXT, created_at TEXT, status TEXT DEFAULT 'valid',
        verify_token TEXT UNIQUE, qr_path TEXT, template TEXT DEFAULT 'classic', recipient_member_id TEXT,
        issuer_name TEXT, custom_title TEXT, custom_intro TEXT, custom_body TEXT, custom_footer TEXT,
        logo1 TEXT, logo2 TEXT, logo3 TEXT);
    CREATE TABLE IF NOT EXISTS honor_list(
        id TEXT PRIMARY KEY, name TEXT NOT NULL, photo TEXT, honor_type TEXT, reason TEXT,
        achievement TEXT, honor_date TEXT, occasion TEXT, department TEXT, certificate_id TEXT,
        badge TEXT, description TEXT, public INTEGER DEFAULT 1, created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS attachments(
        id TEXT PRIMARY KEY, parent_type TEXT, parent_id TEXT, file_path TEXT, media_type TEXT,
        title TEXT, sort_order INTEGER DEFAULT 0, created_at TEXT);
    CREATE TABLE IF NOT EXISTS card_templates(
        id TEXT PRIMARY KEY, name TEXT, kind TEXT, width_mm REAL DEFAULT 85.6, height_mm REAL DEFAULT 54,
        bg TEXT DEFAULT '#071522', accent TEXT DEFAULT '#20B486', text_color TEXT DEFAULT '#FFFFFF',
        font TEXT DEFAULT 'Tajawal', logo TEXT DEFAULT '', public_default INTEGER DEFAULT 0,
        fields_json TEXT DEFAULT '{}', created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS podcast_shows(
        id TEXT PRIMARY KEY, title TEXT, subtitle TEXT, description TEXT, cover_image TEXT, host TEXT,
        status TEXT DEFAULT 'active', created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS podcast_guests(
        id TEXT PRIMARY KEY, name TEXT, bio TEXT, photo TEXT, specialty TEXT, created_at TEXT);
    CREATE TABLE IF NOT EXISTS podcast_episodes(
        id TEXT PRIMARY KEY, show_id TEXT, season INTEGER DEFAULT 1, episode_no INTEGER DEFAULT 1,
        title TEXT, description TEXT, guest_id TEXT, host TEXT, director TEXT, producer TEXT,
        recorded_at TEXT, published_at TEXT, duration TEXT, cover_image TEXT, audio_file TEXT, video_file TEXT,
        transcript TEXT, status TEXT DEFAULT 'draft', created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS podcast_clips(
        id TEXT PRIMARY KEY, episode_id TEXT, title TEXT, file_path TEXT, cover_image TEXT, created_at TEXT);
    CREATE TABLE IF NOT EXISTS podcast_equipment(
        id TEXT PRIMARY KEY, category TEXT, name TEXT, quantity INTEGER DEFAULT 1, status TEXT DEFAULT 'available', notes TEXT);
    CREATE TABLE IF NOT EXISTS temporary_volunteers(
        id TEXT PRIMARY KEY, name TEXT, phone TEXT, department TEXT, stage TEXT, specialty TEXT,
        task TEXT, committee TEXT, start_date TEXT, end_date TEXT, hours REAL DEFAULT 0,
        benefits TEXT, status TEXT DEFAULT 'active', photo TEXT, card_public INTEGER DEFAULT 1,
        created_at TEXT);
    CREATE TABLE IF NOT EXISTS academy_courses(
        id TEXT PRIMARY KEY, title TEXT, description TEXT, instructor TEXT, level TEXT, duration TEXT,
        cover_image TEXT, public INTEGER DEFAULT 1, status TEXT DEFAULT 'published', created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS academy_payments(
        id TEXT PRIMARY KEY, course_id TEXT, enrollment_id TEXT, learner_key TEXT, learner_name TEXT, learner_email TEXT,
        amount REAL DEFAULT 0, currency TEXT DEFAULT 'IQD', proof_path TEXT, learner_note TEXT, admin_note TEXT,
        status TEXT DEFAULT 'pending', created_at TEXT, reviewed_at TEXT, reviewed_by TEXT);
    CREATE TABLE IF NOT EXISTS academy_lessons(
        id TEXT PRIMARY KEY, course_id TEXT, title TEXT, content TEXT, video TEXT, duration TEXT, sort_order INTEGER DEFAULT 0, created_at TEXT);
    CREATE TABLE IF NOT EXISTS academy_enrollments(
        id TEXT PRIMARY KEY, course_id TEXT, learner_key TEXT, learner_name TEXT, learner_email TEXT,
        progress REAL DEFAULT 0, enrolled_at TEXT, last_seen TEXT, completed_at TEXT);
    CREATE TABLE IF NOT EXISTS academy_progress(
        id TEXT PRIMARY KEY, enrollment_id TEXT, lesson_id TEXT, completed INTEGER DEFAULT 0,
        score REAL DEFAULT 0, completed_at TEXT, UNIQUE(enrollment_id,lesson_id));
    CREATE TABLE IF NOT EXISTS academy_quizzes(
        id TEXT PRIMARY KEY, lesson_id TEXT, question TEXT, options_json TEXT, answer_index INTEGER DEFAULT 0,
        points REAL DEFAULT 1, sort_order INTEGER DEFAULT 0);
    CREATE TABLE IF NOT EXISTS academy_resources(
        id TEXT PRIMARY KEY, course_id TEXT, lesson_id TEXT, title TEXT, file_path TEXT, url TEXT, resource_type TEXT, created_at TEXT);
    CREATE TABLE IF NOT EXISTS finance_entries(
        id TEXT PRIMARY KEY, entry_type TEXT, title TEXT, amount REAL DEFAULT 0, category TEXT, entry_date TEXT,
        description TEXT, public INTEGER DEFAULT 0, created_at TEXT);
    CREATE TABLE IF NOT EXISTS equipment_loans(
        id TEXT PRIMARY KEY, equipment_id TEXT, member_id TEXT, issued_at TEXT, returned_at TEXT, status TEXT DEFAULT 'issued', notes TEXT);
    CREATE TABLE IF NOT EXISTS task_completions(
        id TEXT PRIMARY KEY, task_id TEXT, member_id TEXT, completed_at TEXT, notes TEXT,
        UNIQUE(task_id, member_id));
    CREATE TABLE IF NOT EXISTS chat_conversations(
        id TEXT PRIMARY KEY, user_a TEXT NOT NULL, user_b TEXT NOT NULL, created_at TEXT, updated_at TEXT,
        UNIQUE(user_a,user_b));
    CREATE TABLE IF NOT EXISTS chat_messages(
        id TEXT PRIMARY KEY, conversation_id TEXT NOT NULL, sender_id TEXT NOT NULL, body TEXT,
        media_path TEXT, media_type TEXT, created_at TEXT, edited_at TEXT, deleted_at TEXT);
    CREATE TABLE IF NOT EXISTS chat_reactions(
        id TEXT PRIMARY KEY, message_id TEXT NOT NULL, user_id TEXT NOT NULL, reaction TEXT NOT NULL, created_at TEXT,
        UNIQUE(message_id,user_id));
    CREATE TABLE IF NOT EXISTS site_notifications(
        id TEXT PRIMARY KEY, title TEXT, body TEXT, active INTEGER DEFAULT 1, created_at TEXT, expires_at TEXT);
    """)
    db.commit()
    if db.execute("SELECT COUNT(*) c FROM settings").fetchone()["c"] == 0:
        db.execute("INSERT INTO settings(key,value) VALUES('teamName',?)", ("فريق الحكمة التطوعي",))
        db.execute("INSERT INTO settings(key,value) VALUES('subtitle',?)",
                   ("فريق الحكمة التطوعي — AHVT",))
        db.execute("INSERT INTO settings(key,value) VALUES('weights',?)", (json.dumps(DEFAULT_WEIGHTS),))
        db.execute("INSERT INTO settings(key,value) VALUES('points',?)", (json.dumps(DEFAULT_POINTS),))
        # Lightweight migrations for existing deployments.
    def ensure_column(table, column, definition):
        cols = {r["name"] for r in db.execute(f"PRAGMA table_info({table})").fetchall()}
        if column not in cols:
            db.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")
    ensure_column("members", "photo", "TEXT")
    ensure_column("members", "department", "TEXT")
    ensure_column("members", "stage", "TEXT")
    ensure_column("members", "membership_no", "TEXT")
    ensure_column("members", "card_issue_date", "TEXT")
    ensure_column("members", "card_expiry", "TEXT")
    ensure_column("temporary_volunteers", "benefits", "TEXT")
    ensure_column("temporary_volunteers", "card_public", "INTEGER DEFAULT 1")
    ensure_column("members", "nfc_uid", "TEXT")
    ensure_column("certificates", "status", "TEXT DEFAULT 'valid'")
    ensure_column("certificates", "verify_token", "TEXT")
    ensure_column("certificates", "qr_path", "TEXT")
    ensure_column("certificates", "template", "TEXT DEFAULT 'classic'")
    ensure_column("certificates", "recipient_member_id", "TEXT")
    ensure_column("certificates", "issuer_name", "TEXT")
    ensure_column("certificates", "custom_title", "TEXT")
    ensure_column("certificates", "custom_intro", "TEXT")
    ensure_column("certificates", "custom_body", "TEXT")
    ensure_column("certificates", "custom_footer", "TEXT")
    ensure_column("certificates", "logo1", "TEXT")
    ensure_column("certificates", "logo2", "TEXT")
    ensure_column("certificates", "logo3", "TEXT")
    ensure_column("administrators", "photo", "TEXT")
    ensure_column("administrators", "department", "TEXT")
    ensure_column("administrators", "card_issue_date", "TEXT")
    ensure_column("administrators", "card_expiry", "TEXT")
    ensure_column("administrators", "membership_no", "TEXT")
    ensure_column("administrators", "nfc_uid", "TEXT")
    ensure_column("users", "username", "TEXT")
    ensure_column("users", "active", "INTEGER DEFAULT 1")
    ensure_column("users", "must_set_password", "INTEGER DEFAULT 0")
    ensure_column("users", "last_login", "TEXT")
    ensure_column("users", "password_plain", "TEXT")
    ensure_column("live_sessions", "channel_name", "TEXT")
    ensure_column("live_sessions", "broadcast_type", "TEXT DEFAULT 'لجنة'")
    ensure_column("live_sessions", "event_location", "TEXT")
    ensure_column("live_sessions", "scheduled_date", "TEXT")
    ensure_column("live_sessions", "scheduled_time", "TEXT")
    ensure_column("live_sessions", "external_links", "TEXT DEFAULT '[]'")
    ensure_column("live_sessions", "viewer_peak", "INTEGER DEFAULT 0")
    ensure_column("members", "public_profile", "INTEGER DEFAULT 1")
    ensure_column("administrators", "public_profile", "INTEGER DEFAULT 1")
    ensure_column("events", "qr_path", "TEXT")
    ensure_column("tasks", "qr_path", "TEXT")
    db.execute("""CREATE TABLE IF NOT EXISTS live_source_requests(
        id TEXT PRIMARY KEY, session_id TEXT, requester_name TEXT, requester_user_id TEXT,
        device_label TEXT, status TEXT DEFAULT 'pending', created_at TEXT, reviewed_at TEXT,
        reviewed_by TEXT
    )""")
    ensure_column("live_sessions", "recording_path", "TEXT")
    ensure_column("academy_courses", "is_paid", "INTEGER DEFAULT 0")
    ensure_column("academy_courses", "price_iqd", "REAL DEFAULT 0")
    ensure_column("academy_courses", "payment_required", "INTEGER DEFAULT 0")
    ensure_column("academy_lessons", "resources_json", "TEXT DEFAULT '[]'")
    ensure_column("academy_lessons", "quiz_enabled", "INTEGER DEFAULT 0")
    ensure_column("members", "card_public", "INTEGER DEFAULT 0")
    ensure_column("administrators", "card_public", "INTEGER DEFAULT 0")
    ensure_column("tasks", "latitude", "REAL")
    ensure_column("tasks", "longitude", "REAL")
    ensure_column("tasks", "map_visible", "INTEGER DEFAULT 0")
    ensure_column("podcast_equipment", "qr_path", "TEXT")
    ensure_column("live_peers", "label", "TEXT")
    ensure_column("live_peers", "source_index", "INTEGER DEFAULT 0")
    ensure_column("certificates", "writer_name", "TEXT")
    ensure_column("certificates", "paper_size", "TEXT DEFAULT 'A4 landscape'")
    ensure_column("certificates", "custom_eyebrow", "TEXT")
    ensure_column("certificates", "custom_date_label", "TEXT")
    ensure_column("certificates", "custom_hours_label", "TEXT")
    ensure_column("certificates", "writer_label", "TEXT")
    ensure_column("certificates", "design_bg", "TEXT")
    ensure_column("certificates", "design_accent", "TEXT")
    ensure_column("certificates", "design_text", "TEXT")
    ensure_column("certificates", "design_font", "TEXT")
    ensure_column("media", "media_type", "TEXT DEFAULT 'link'")
    ensure_column("media", "thumbnail", "TEXT")
    ensure_column("media", "sort_order", "INTEGER DEFAULT 0")
    ensure_column("news", "featured", "INTEGER DEFAULT 0")
    ensure_column("news", "scheduled_at", "TEXT")
    ensure_column("initiatives", "latitude", "REAL")
    ensure_column("initiatives", "longitude", "REAL")
    ensure_column("initiatives", "map_status", "TEXT DEFAULT 'completed'")
    ensure_column("initiatives", "map_note", "TEXT")
    ensure_column("initiatives", "map_date", "TEXT")
    ensure_column("initiatives", "map_visible", "INTEGER DEFAULT 1")
    ensure_column("initiative_participants", "start_time", "TEXT")
    ensure_column("initiative_participants", "end_time", "TEXT")
    ensure_column("initiative_participants", "hours", "REAL DEFAULT 0")
    ensure_column("attendance", "start_time", "TEXT")
    ensure_column("attendance", "end_time", "TEXT")
    ensure_column("attendance", "hours", "REAL DEFAULT 0")
    # Backfill newer appearance/public settings without destroying existing configuration.
    defaults = {
        "permissionsEnabled": "1",
        "accentColor": "#20B486", "navyColor": "#071A2F", "backgroundColor": "#050B13",
        "fontFamily": "Tajawal", "heroTitle": "أثرٌ يُقاس، وقيادةٌ تُصنع",
        "heroText": "منصة AHVT لإدارة الأثر، المبادرات، والأداء التطوعي.",
        "announcement": "", "customCss": "", "siteMode": "public",
        "teamLogo": "", "universityLogo": "", "favicon": "",
        "telegramUrl": "https://t.me/Hikmaht_bot", "instagramUrl": "https://www.instagram.com/hikma.ahvt?igsi=MXhodTN3cndpenI0NQ==", "siteDescription": "منصة AHVT لإدارة الأثر، المبادرات، والأداء التطوعي.",
        "showPublicAdmins": "1", "showPublicNews": "1", "maintenanceMode": "0",
        "joinButtonVisible": "1", "joinButtonText": "انضم ↗", "joinButtonIcon": "↗", "joinButtonPlacement": "hero", "joinButtonMode": "telegram", "adminLinkVisible": "1",
        "heroBackground": "", "heroVideo": "", "publicBackground": "", "cinematicMode": "1",
        "seoTitle": "AHVT | فريق الحكمة التطوعي", "seoDescription": "منصة AHVT لإدارة الأثر، المبادرات، والأداء التطوعي.", "maintenanceMode": "0",
        "publicNavJson": "", "publicSectionsJson": "", "siteMode": "public",
        "liveEnabled": "0", "liveTitle": "البث المباشر من الميدان", "liveDescription": "تابع مبادرات فريق الحكمة التطوعي مباشرةً.", "liveUrl": "", "livePlatform": "YouTube", "liveMode": "internal", "liveSessionId": "", "liveStunServers": "stun:stun.l.google.com:19302",
        "certificateIssuer": "فريق الحكمة التطوعي",
        "darkMode": "1",
    }
    for k,v in defaults.items():
        db.execute("INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)", (k,v))

    # Normalize the official team name on existing deployments without touching user data.
    db.execute("UPDATE settings SET value=REPLACE(value, ?, ?) WHERE value LIKE ?",
               ("فريق الحكمة الطلابي", "فريق الحكمة التطوعي", "%فريق الحكمة الطلابي%"))
    db.execute("UPDATE settings SET value=? WHERE key='teamName' AND (value IS NULL OR value='' OR value='AHVT')", ("فريق الحكمة التطوعي",))
    db.execute("UPDATE settings SET value=REPLACE(value, ?, ?) WHERE value LIKE ?", ("HIKMA IMPACT", "AHVT", "%HIKMA IMPACT%"))

    # ======================================================== INITIAL ADMIN ACCOUNTS ========================================================
    # These directory accounts are created without a usable password.
    # The Creator assigns each account its own password from the Admin Accounts page.
    seed_admins = [
        ("بان حسين", "ban.hussein"),
        ("محمد صادق جاسم", "mohammad.sadiq"),
        ("رامي راسم", "rami.rasim"),
        ("علي احمد", "ali.ahmad"),
        ("منتظر حيدر", "muntather.haider"),
    ]
    for admin_name, username in seed_admins:
        existing = db.execute("SELECT id FROM users WHERE lower(username)=?", (username.lower(),)).fetchone()
        if not existing:
            db.execute(
                "INSERT INTO users(id,name,email,username,password_hash,password_plain,role,active,must_set_password,last_login,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                (uid("usr"), admin_name, None, username, generate_password_hash(uuid.uuid4().hex), uuid.uuid4().hex, "ADMIN", 1, 1, None, now_iso())
            )
        # Also keep the public administrative directory populated.
        dir_exists = db.execute("SELECT id FROM administrators WHERE name=?", (admin_name,)).fetchone()
        if not dir_exists:
            db.execute(
                "INSERT INTO administrators(id,name,position,committee,date,responsibilities) VALUES(?,?,?,?,?,?)",
                (uid("adm"), admin_name, "إداري", "", date.today().isoformat(), "إدارة ومتابعة أعمال الفريق")
            )

    # ======================================================== OWNER ACCOUNT ========================================================
    # Owner credentials are kept stable across restarts.  The plain value is
    # intentionally retained because the owner requested simple credential
    # storage; the login also accepts the legacy Werkzeug hash.
    OWNER_EMAIL = "Abdulrahman.a.alani1@gmail.com".strip().lower()
    # init_db() runs before a Flask application context exists.
    # Read the existing owner password directly from this initialization DB
    # connection instead of calling get_setting(), which uses flask.g.
    owner_password_row = db.execute(
        "SELECT value FROM settings WHERE key=?",
        ("ownerPassword",)
    ).fetchone()
    OWNER_PASSWORD = os.environ.get("HIKMA_OWNER_PASSWORD", "ABAMAL0027")
    owner = db.execute("SELECT id FROM users WHERE lower(email)=?", (OWNER_EMAIL,)).fetchone()
    if owner:
        # Do NOT reset the owner password on every Render restart.
        db.execute(
            "UPDATE users SET name=?, username=?, role=?, active=1, must_set_password=0, password_hash=?, password_plain=? WHERE id=?",
            ("Abdulrahman Alani", "abdulrahman", "CREATOR", generate_password_hash(OWNER_PASSWORD), OWNER_PASSWORD, owner["id"])
        )
    else:
        db.execute(
            "INSERT INTO users(id,name,email,username,password_hash,password_plain,role,active,must_set_password,last_login,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
            (
                "owner_" + uuid.uuid4().hex[:12],
                "Abdulrahman Alani",
                OWNER_EMAIL,
                "abdulrahman",
                generate_password_hash(OWNER_PASSWORD),
                OWNER_PASSWORD,
                "CREATOR",
                1,
                0,
                None,
                now_iso(),
            )
        )
    db.execute("INSERT OR IGNORE INTO settings(key,value) VALUES('ownerPassword',?)", (OWNER_PASSWORD,))
    default_nav = [
        ("home","الرئيسية","public_home","/", "⌂",1,10,1),
        ("about","عن الفريق","about_public","/about", "◈",1,20,1),
        ("admins","الإداريون","admins_list","/administrators", "♜",1,30,1),
        ("news","الأخبار","public_news","/news", "▣",1,40,1),
        ("initiatives","المبادرات","initiatives_list","/initiatives", "✦",1,50,1),
        ("committees","اللجان","committees_list","/committees", "◫",1,60,1),
        ("events","الفعاليات","events_public","/events", "◷",1,70,1),
        ("impact","الأثر","impact_public","/impact", "◉",1,80,1),
        ("impact_map","خريطة الأثر","impact_map_public","/impact-map", "⌖",1,85,1),
        ("media","المعرض","media_public","/media", "▧",1,90,1),
        ("partners","الشركاء","partners_public","/partners", "◇",1,100,1),
        ("honor","قائمة الشرف","honor_public","/honor", "🏅",1,105,1),
        ("search","بحث","search_public","/search", "⌕",1,110,1),
        ("cards","بطاقات الفريق","cards_admin","/admin/cards", "▣",1,115,1),
        ("podcast","البودكاست","podcast_public","/podcast", "◉",1,120,1),
        ("podcast_admin","استوديو البودكاست","podcast_admin","/admin/podcast", "🎙",1,125,1),
        ("academy","أكاديمية AHVT","academy_public","/academy", "🎓",1,130,1),
        ("calendar","التقويم","calendar_public","/calendar", "🗓",1,140,1),
        ("media_center","مركز الوسائط","media_center_public","/media-center", "🎞",1,150,1),
        ("equipment","المعدات","equipment_public","/equipment", "🧰",1,160,1),
        ("finance","الشفافية المالية","finance_public","/finance", "💠",1,170,1),
        ("assistant","مساعد AHVT","assistant_public","/assistant", "🧠",1,180,1),
    ]
    for key,label,endpoint,url,icon,visible,sort_order,is_system in default_nav:
        db.execute("INSERT OR IGNORE INTO nav_items(id,label,endpoint,url,icon,visible,sort_order,is_system,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
                   ("nav_"+key,label,endpoint,url,icon,visible,sort_order,is_system,now_iso()))
    default_sections=[
        ("hero","البداية السينمائية",1,10,"hero",""),("news","آخر الأخبار",1,20,"light",""),
        ("impact","الأثر بالأرقام",1,30,"light",""),("impact_map","خريطة الأثر",1,35,"tint",""),("live","البث المباشر",1,38,"dark",""),("initiatives","المبادرات",1,40,"tint",""),
        ("media","الصور والفيديو",1,50,"light",""),("administrators","الإداريون",1,60,"light",""),
        ("committees","اللجان",1,70,"light",""),("events","الفعاليات",1,80,"dark",""),
        ("partners","الشركاء",1,90,"light",""),("join","انضم إلى الفريق",0,100,"join",""),
    ]
    for key,title,visible,sort_order,bg,accent in default_sections:
        db.execute("INSERT OR IGNORE INTO site_sections(id,section_key,title,visible,sort_order,background,accent,updated_at) VALUES(?,?,?,?,?,?,?,?)",
                   ("sec_"+key,key,title,visible,sort_order,bg,accent,now_iso()))
    # Keep the public join CTA only in the hero; disable legacy standalone join section.
    db.execute("UPDATE site_sections SET visible=0 WHERE section_key='join'")
    permission_names=[
        "dashboard","members","administrators","committees","initiatives","tasks","attendance","evaluations",
        "achievements","news","events","media","partners","pages","applications","reports","analytics","goals",
        "academy","academy_manage","calendar","media_center","equipment","finance","assistant","site_notifications",
        "approvals","notifications","audit","security","insights","risk","decisions","appearance","navigation",
        "sections","permissions","backup","system","delete","publish","upload_media","manage_admins","live","certificates","honor","cards","podcast","podcast_admin"
    ]
    for role in ("ADMIN","SUPER_ADMIN","CREATOR"):
        for perm in permission_names:
            allowed = 1 if role in ("SUPER_ADMIN","CREATOR") or perm not in ("permissions","backup","system","security","delete","manage_admins") else 0
            db.execute("INSERT OR IGNORE INTO role_permissions(id,role,permission,allowed) VALUES(?,?,?,?)",(uid("perm"),role,perm,allowed))
    for kind, name in (("member","HIKMA Member 2026"),("admin","HIKMA Admin 2026")):
        if not db.execute("SELECT 1 FROM card_templates WHERE kind=? LIMIT 1",(kind,)).fetchone():
            db.execute("INSERT INTO card_templates(id,name,kind,width_mm,height_mm,bg,accent,text_color,font,public_default,fields_json,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",(uid("ct"),name,kind,85.6,54,"#071522","#20B486","#FFFFFF","Tajawal",0,json.dumps({"show_department":True,"show_stage":True,"show_committee":True,"show_position":True,"show_dates":True,"show_nfc":True},ensure_ascii=False),now_iso(),now_iso()))
    db.commit()
    db.close()


def uid(prefix="id"):
    return f"{prefix}_{uuid.uuid4().hex[:12]}"


def now_iso():
    return datetime.utcnow().isoformat()


# ============================================================ SETTINGS HELPERS ============================================================
def get_setting(key, default=None):
    row = get_db().execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    return row["value"] if row else default


def get_weights():
    return json.loads(get_setting("weights", json.dumps(DEFAULT_WEIGHTS)))


def get_points_config():
    return json.loads(get_setting("points", json.dumps(DEFAULT_POINTS)))


def set_setting(key, value):
    db = get_db()
    db.execute("INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
               (key, value))
    db.commit()


# Payment display settings (safe: never store card CVV/full card data)
set_setting("payment_name", get_setting("payment_name", "")) if False else None

# ============================================================ AUTH ============================================================
def current_user():
    uid_ = session.get("user_id")
    if not uid_:
        return None
    return get_db().execute("SELECT * FROM users WHERE id=?", (uid_,)).fetchone()

def is_admin():
    u = current_user()
    return bool(u and u["role"] in ("CREATOR", "SUPER_ADMIN", "ADMIN"))

def is_creator():
    u = current_user()
    return bool(u and u["role"] in ("CREATOR", "SUPER_ADMIN"))

PERMISSION_ENDPOINTS = {
    "admin_dashboard":"dashboard","members_list":"members","member_view":"members","admins_list":"administrators","admin_new":"administrators","admin_edit":"administrators","admin_delete":"delete",
    "committees_list":"committees","committee_new":"committees","committee_edit":"committees","committee_delete":"delete",
    "initiatives_list":"initiatives","initiative_new":"initiatives","initiative_edit":"initiatives","initiative_delete":"delete","initiative_participants":"initiatives",
    "tasks_list":"tasks","task_new":"tasks","task_edit":"tasks","task_delete":"delete","attendance_list":"attendance","attendance_new":"attendance",
    "evaluations_list":"evaluations","evaluation_new":"evaluations","achievements":"achievements","news_admin":"news","news_new":"news","news_edit":"news","news_delete":"delete",
    "events_admin":"events","event_delete":"delete","partners_admin":"partners","media_admin":"media","media_upload":"upload_media",
    "page_admin":"pages","page_new":"pages","page_edit":"pages","page_delete":"delete","reports_page":"reports","analytics_page":"analytics","goals_page":"goals",
    "approvals_page":"approvals","approval_review":"approvals","notifications_page":"notifications","audit_log":"audit","security_page":"security","insights_page":"insights","risk_page":"risk","decision_center":"decisions",
    "settings_page":"appearance","certificates_admin":"certificates","honor_admin":"honor","honor_new":"honor","honor_edit":"honor","honor_delete":"delete","certificate_legacy":"certificates","navigation_page":"navigation","sections_page":"sections","permissions_page":"permissions","live_admin":"live","backup_export":"backup","admin_users":"manage_admins","admin_user_new":"manage_admins","admin_user_role":"manage_admins",
    "academy_admin":"academy_manage","equipment_admin":"equipment","finance_admin":"finance","assistant_admin":"assistant","site_notifications_admin":"site_notifications",
}

def has_permission(permission):
    u=current_user()
    if not u: return False
    if u["role"]=="CREATOR": return True
    # Granular permissions can be enabled/disabled by the owner from Settings.
    if get_setting("permissionsEnabled", "1") != "1":
        return u["role"] in ("ADMIN", "SUPER_ADMIN")
    row=get_db().execute("SELECT allowed FROM role_permissions WHERE role=? AND permission=?",(u["role"],permission)).fetchone()
    return bool(row and row["allowed"])

ADMIN_GET_ENDPOINTS = {
    "settings_page", "certificates_admin", "certificate_edit", "certificate_legacy", "honor_admin", "honor_new", "honor_edit", "honor_delete", "cards_admin", "card_print", "audit_log", "backup_export", "admin_new", "admin_edit", "admin_delete",
    "admin_users", "admin_user_role", "admin_login", "admin_dashboard", "news_admin", "news_new",
    "news_edit", "news_delete", "page_admin", "page_new", "page_edit", "page_delete",
    "notifications_page", "analytics_page", "goals_page", "approvals_page", "events_admin",
    "partners_admin", "media_admin", "security_page", "certificate", "certificate_view", "insights_page", "risk_page", "decision_center", "navigation_page", "navigation_delete", "sections_page", "permissions_page", "media_upload", "control_center", "upload_public_asset", "live_admin", "academy_payments_admin",
}
PUBLIC_ENDPOINTS = {"public_home", "public_news", "public_news_detail", "public_page",
                    "committees_list", "initiatives_list", "initiative_view", "members_list", "member_view", "achievements", "honor_public",
                    "admins_list", "events_public", "media_public", "partners_public", "impact_public",
                    "about_public", "search_public", "volunteer_redirect", "api_impact", "impact_map_public", "live_public",
                    "login", "admin_login", "logout", "static", "signup", "academy_learn", "academy_lesson_complete", "academy_quiz_submit", "academy_media", "academy_payment", "academy_my_payments", "live_room", "live_join", "live_peers_api", "live_signal", "live_leave", "certificate_verify", "podcast_public", "podcast_episode", "card_public", "chat_page", "chat_open", "chat_send", "chat_edit", "chat_delete", "chat_react", "scan_member_action", "scan_task_action", "scan_equipment_action"}

@app.before_request
def access_control():
    endpoint = request.endpoint or ""
    # Public GET pages are view-only. Any write request requires an admin.
    if endpoint in PUBLIC_ENDPOINTS:
        return
    if request.method in ("POST", "PUT", "PATCH", "DELETE"):
        if not is_admin():
            if request.is_json:
                return jsonify({"ok": False, "error": "Admin access required"}), 403
            flash("هذه العملية متاحة للإدارة فقط", "error")
            return redirect(url_for("admin_login"))
    if endpoint in ADMIN_GET_ENDPOINTS or any(x in endpoint for x in ("_new", "_edit", "_delete", "backup_")):
        if not is_admin():
            flash("هذه الصفحة متاحة للإدارة فقط", "error")
            return redirect(url_for("admin_login"))
        perm=PERMISSION_ENDPOINTS.get(endpoint)
        if perm and not has_permission(perm):
            return render_template("403.html", permission=perm), 403

def get_live_embed_url(url):
    url=(url or "").strip()
    if not url: return ""
    from urllib.parse import urlparse, parse_qs
    try:
        u=urlparse(url)
        host=u.netloc.lower()
        if "youtube.com" in host:
            if u.path.startswith("/watch"):
                vid=parse_qs(u.query).get("v", [""])[0]
                return f"https://www.youtube.com/embed/{vid}?autoplay=0&rel=0" if vid else ""
            if u.path.startswith("/live/"):
                vid=u.path.split("/live/",1)[1].split("/",1)[0]
                return f"https://www.youtube.com/embed/{vid}?autoplay=0&rel=0" if vid else ""
            if u.path.startswith("/embed/"):
                return url
        if "youtu.be" in host:
            vid=u.path.strip("/").split("/")[0]
            return f"https://www.youtube.com/embed/{vid}?autoplay=0&rel=0" if vid else ""
    except Exception: pass
    return ""

@app.context_processor
def inject_globals():
    return {
        "session_user": current_user(),
        "is_admin": is_admin(),
        "is_creator": is_creator(),
        "team_name": get_setting("teamName", "AHVT"),
        "subtitle": get_setting("subtitle", "فريق الحكمة التطوعي"),
        "site_settings": {k: get_setting(k, "") for k in ["accentColor","navyColor","backgroundColor","fontFamily","heroTitle","heroText","announcement","customCss","teamLogo","universityLogo","favicon","telegramUrl","instagramUrl","siteDescription","showPublicAdmins","showPublicNews","joinButtonVisible","joinButtonText","joinButtonIcon","joinButtonPlacement","joinButtonMode","adminLinkVisible","heroBackground","heroVideo","publicBackground","cinematicMode","maintenanceMode","seoTitle","seoDescription","liveEnabled","liveTitle","liveDescription","liveUrl","livePlatform","liveMode","liveSessionId","liveStunServers","certificateIssuer","permissionsEnabled","payment_name","payment_provider","payment_account","payment_instructions"]},
        "public_nav": get_db().execute("SELECT * FROM nav_items WHERE visible=1 ORDER BY sort_order").fetchall(),
        "public_sections": get_db().execute("SELECT * FROM site_sections WHERE visible=1 ORDER BY sort_order").fetchall(),
        "current_endpoint": request.endpoint,
        "unread_notifications": (get_db().execute("SELECT COUNT(*) c FROM notifications WHERE user_id=? AND read_at IS NULL", (session.get("user_id"),)).fetchone()["c"] if session.get("user_id") else 0),
        "popup_notifications": (get_db().execute("SELECT * FROM notifications WHERE user_id=? AND read_at IS NULL ORDER BY created_at DESC LIMIT 3", (session.get("user_id"),)).fetchall() if session.get("user_id") else []),
        "site_alerts": get_db().execute("SELECT * FROM site_notifications WHERE active=1 AND (expires_at IS NULL OR expires_at='' OR expires_at>=?) ORDER BY created_at DESC LIMIT 3", (now_iso(),)).fetchall(),
        "telegram_url": get_setting("telegramUrl", "https://t.me/Hikmaht_bot"),
        "instagram_url": get_setting("instagramUrl", "https://www.instagram.com/hikma.ahvt?igsi=MXhodTN3cndpenI0NQ=="),
        "public_background": get_setting("publicBackground", ""),
        "latest_public_alerts": get_db().execute("SELECT slug,title FROM news WHERE status='published' ORDER BY published_at DESC,created_at DESC LIMIT 4").fetchall(),
        "live_embed_url": get_live_embed_url(get_setting("liveUrl", "")),
        "get_attachments": lambda parent_type, parent_id: attachments_for(get_db(), parent_type, parent_id),
    }

def log_action(action, target):
    db = get_db()
    u = current_user()
    db.execute("INSERT INTO audit_logs(id,action,target,by,date) VALUES(?,?,?,?,?)",
               (uid("log"), action, target, u["name"] if u else "—", now_iso()))
    db.commit()

def add_points(member_id, value, source):
    db = get_db()
    db.execute("INSERT INTO points(id,member_id,value,source,date) VALUES(?,?,?,?,?)",
               (uid("pt"), member_id, value, source, now_iso()))
    db.commit()

@app.route("/login", methods=["GET", "POST"])
def login():
    return redirect(url_for("admin_login"))

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        user = get_db().execute(
            "SELECT * FROM users WHERE lower(email)=? OR lower(username)=?",
            (email, email)
        ).fetchone()
        valid_plain = bool(user and user["active"] and user["role"] in ("CREATOR", "SUPER_ADMIN", "ADMIN") and user["password_plain"] is not None and password == user["password_plain"]) if user else False
        valid_hash = False
        if user and user["password_hash"]:
            try:
                valid_hash = check_password_hash(user["password_hash"], password)
            except Exception:
                valid_hash = False
        if not user or not user["active"] or user["role"] not in ("CREATOR", "SUPER_ADMIN", "ADMIN") or not (valid_plain or valid_hash):
            flash("بيانات الإدارة غير صحيحة", "error")
            return redirect(url_for("admin_login"))
        session["user_id"] = user["id"]
        db=get_db()
        db.execute("UPDATE users SET last_login=? WHERE id=?", (now_iso(), user["id"]))
        db.execute("INSERT INTO security_sessions(id,user_id,ip,user_agent,created_at,last_seen) VALUES(?,?,?,?,?,?)", (uid("ses"),user["id"],request.headers.get("X-Forwarded-For",request.remote_addr),request.headers.get("User-Agent",""),now_iso(),now_iso()))
        db.commit()
        return redirect(url_for("admin_dashboard"))
    has_admin = get_db().execute("SELECT COUNT(*) c FROM users WHERE role IN ('CREATOR','SUPER_ADMIN','ADMIN')").fetchone()["c"] > 0
    return render_template("admin_login.html", has_admin=has_admin)

@app.route("/signup", methods=["GET", "POST"])
def signup():
    # Public account creation is intentionally disabled in v2.
    return redirect(url_for("admin_login"))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("public_home"))

# ============================================================ COMPUTATION ENGINE ============================================================
def member_score(db, member_id):
    weights = get_weights()
    evs = db.execute("SELECT * FROM evaluations WHERE evaluated_user_id=?", (member_id,)).fetchall()
    if not evs:
        return None
    n = len(evs)
    sums = {k: 0 for k in CRITERIA_KEYS}
    for e in evs:
        for k in CRITERIA_KEYS:
            sums[k] += e[f"c_{k}"] or 0
    total_weight = sum(weights.values()) or 100
    score = 0
    for k in CRITERIA_KEYS:
        score += (sums[k] / n) * (weights.get(k, 0) / total_weight)
    return round(score)


def iraq_now_time():
    """Current Iraq local time, used when a supervisor records entry/exit."""
    if ZoneInfo:
        return datetime.now(ZoneInfo("Asia/Baghdad")).strftime("%H:%M")
    return datetime.utcnow().strftime("%H:%M")

def iraq_today():
    if ZoneInfo:
        return datetime.now(ZoneInfo("Asia/Baghdad")).date().isoformat()
    return datetime.utcnow().date().isoformat()

def duration_hours(start_time, end_time):
    """Return elapsed hours between HH:MM times; supports sessions crossing midnight."""
    if not start_time or not end_time:
        return 0.0
    try:
        sh, sm = [int(x) for x in start_time.split(":")[:2]]
        eh, em = [int(x) for x in end_time.split(":")[:2]]
        start = sh * 60 + sm
        end = eh * 60 + em
        if end < start:
            end += 24 * 60
        return round((end - start) / 60.0, 2)
    except (ValueError, TypeError):
        return 0.0

def member_volunteer_hours(db, member_id):
    attendance = db.execute(
        "SELECT COALESCE(SUM(hours),0) h FROM attendance WHERE member_id=? AND status IN ('Present','Late') AND (initiative_id IS NULL OR initiative_id = '')",
        (member_id,)).fetchone()["h"] or 0
    initiatives = db.execute(
        "SELECT COALESCE(SUM(hours),0) h FROM initiative_participants WHERE member_id=?",
        (member_id,)).fetchone()["h"] or 0
    return round(float(attendance) + float(initiatives), 2)

def member_attendance_pct(db, member_id):
    rows = db.execute("SELECT status FROM attendance WHERE member_id=?", (member_id,)).fetchall()
    if not rows:
        return None
    good = sum(1 for r in rows if r["status"] in ("Present", "Late"))
    return round(good / len(rows) * 100)


def member_points_total(db, member_id):
    row = db.execute("SELECT COALESCE(SUM(value),0) s FROM points WHERE member_id=?", (member_id,)).fetchone()
    return row["s"] or 0


def member_level(db, member_id):
    p = member_points_total(db, member_id)
    s = member_score(db, member_id) or 0
    if p >= 500 and s >= 90:
        return "سفير الحكمة"
    if p >= 300 and s >= 85:
        return "قائد فريق"
    if p >= 150 and s >= 75:
        return "متطوع متقدم"
    if p >= 50 and s >= 60:
        return "عضو نشط"
    return "متطوع"


def recommendation_text(s):
    if s is None:
        return "لا توجد بيانات كافية لتوليد توصية بعد."
    if s >= 85:
        return "أداء متميز واستمرارية عالية في المشاركة."
    if s >= 65:
        return "أداء جيد مع وجود فرص لتحسين بعض الجوانب."
    return "يوصى بمتابعة الأداء وتحسين الالتزام والمشاركة."


def initiative_participant_count(db, initiative_id):
    return db.execute("SELECT COUNT(*) c FROM initiative_participants WHERE initiative_id=?",
                       (initiative_id,)).fetchone()["c"]


def member_initiative_count(db, member_id):
    return db.execute("SELECT COUNT(*) c FROM initiative_participants WHERE member_id=?",
                       (member_id,)).fetchone()["c"]


def fmt_date(d):
    if not d:
        return "—"
    try:
        dt = datetime.fromisoformat(str(d)[:19])
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return str(d)


app.jinja_env.filters["fmt_date"] = fmt_date


# ============================================================ ADMIN CENTER / NEWS / PAGES ============================================================
@app.route("/admin")
def admin_dashboard():
    db=get_db()
    today=date.today().isoformat()
    stats={
        "members": db.execute("SELECT COUNT(*) c FROM members").fetchone()["c"],
        "active_members": db.execute("SELECT COUNT(*) c FROM members WHERE status IS NULL OR lower(status) IN ('active','نشط')").fetchone()["c"],
        "news": db.execute("SELECT COUNT(*) c FROM news WHERE status='published'").fetchone()["c"],
        "initiatives": db.execute("SELECT COUNT(*) c FROM initiatives").fetchone()["c"],
        "evaluations": db.execute("SELECT COUNT(*) c FROM evaluations").fetchone()["c"],
        "tasks": db.execute("SELECT COUNT(*) c FROM tasks").fetchone()["c"],
        "open_tasks": db.execute("SELECT COUNT(*) c FROM tasks WHERE lower(status) NOT IN ('done','completed','مكتملة','منجزة')").fetchone()["c"],
        "overdue": db.execute("SELECT COUNT(*) c FROM tasks WHERE deadline IS NOT NULL AND deadline < ? AND lower(status) NOT IN ('done','completed','مكتملة','منجزة')", (today,)).fetchone()["c"],
        "committees": db.execute("SELECT COUNT(*) c FROM committees").fetchone()["c"],
        "hours": db.execute("SELECT COALESCE(SUM(hours),0) h FROM initiatives").fetchone()["h"] or 0,
        "admins": db.execute("SELECT COUNT(*) c FROM users WHERE role IN ('ADMIN','SUPER_ADMIN') AND active=1").fetchone()["c"],
        "certificates": db.execute("SELECT COUNT(*) c FROM certificates").fetchone()["c"],
        "pending_approvals": db.execute("SELECT COUNT(*) c FROM approvals WHERE lower(status)='pending'").fetchone()["c"],
    }
    recent=db.execute("SELECT * FROM audit_logs ORDER BY date DESC LIMIT 8").fetchall()
    committees=db.execute("SELECT * FROM committees ORDER BY name").fetchall()
    upcoming=db.execute("SELECT * FROM initiatives WHERE date IS NOT NULL AND date >= ? ORDER BY date ASC LIMIT 5", (today,)).fetchall()
    overdue_tasks=db.execute("SELECT * FROM tasks WHERE deadline IS NOT NULL AND deadline < ? AND lower(status) NOT IN ('done','completed','مكتملة','منجزة') ORDER BY deadline ASC LIMIT 6", (today,)).fetchall()
    pending=db.execute("SELECT * FROM tasks WHERE lower(status) NOT IN ('done','completed','مكتملة','منجزة') ORDER BY deadline ASC LIMIT 8").fetchall()
    if is_creator():
        return render_template("owner_dashboard.html", stats=stats, recent=recent, committees=committees, upcoming=upcoming, overdue_tasks=overdue_tasks, pending=pending)
    return render_template("admin_dashboard.html", stats=stats, recent=recent, committees=committees, upcoming=upcoming, overdue_tasks=overdue_tasks, pending=pending)

@app.route("/admin/users")
def admin_users():
    users=get_db().execute("SELECT id,name,email,username,role,active,must_set_password,last_login,created_at FROM users ORDER BY role DESC, created_at DESC").fetchall()
    return render_template("admin_users.html", users=users)

@app.route("/admin/users/<uid_>/role", methods=["POST"])
def admin_user_role(uid_):
    if not is_creator():
        flash("تغيير صلاحيات الأدمن متاح لصانع التطبيق فقط", "error")
        return redirect(url_for("admin_users"))
    role=request.form.get("role","ADMIN")
    if role not in ("ADMIN","CREATOR"): role="ADMIN"
    db=get_db(); db.execute("UPDATE users SET role=? WHERE id=?", (role,uid_)); db.commit()
    log_action("Role changed", uid_)
    flash("تم تحديث الصلاحية", "ok")
    return redirect(url_for("admin_users"))

@app.route("/admin/users/new", methods=["GET", "POST"])
def admin_user_new():
    if not is_creator():
        flash("إنشاء الحسابات متاح لصانع التطبيق فقط", "error")
        return redirect(url_for("admin_users"))
    if request.method == "POST":
        name=request.form.get("name", "").strip()
        username=request.form.get("username", "").strip().lower()
        email=request.form.get("email", "").strip().lower() or None
        password=request.form.get("password", "")
        confirm=request.form.get("confirm_password", "")
        role=request.form.get("role", "ADMIN")
        position=request.form.get("position", "إداري").strip()
        committee=request.form.get("committee", "").strip()
        if role not in ("ADMIN", "SUPER_ADMIN"): role="ADMIN"
        if not name or not username or not password:
            flash("الاسم واسم المستخدم والرمز مطلوبة", "error")
            return redirect(request.url)
        if password != confirm:
            flash("الرمزان غير متطابقين", "error")
            return redirect(request.url)
        db=get_db()
        if db.execute("SELECT id FROM users WHERE lower(username)=?", (username,)).fetchone():
            flash("اسم المستخدم مستخدم مسبقًا", "error")
            return redirect(request.url)
        if email and db.execute("SELECT id FROM users WHERE lower(email)=?", (email,)).fetchone():
            flash("البريد مستخدم مسبقًا", "error")
            return redirect(request.url)
        db.execute("INSERT INTO users(id,name,email,username,password_hash,password_plain,role,active,must_set_password,last_login,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                   (uid("usr"),name,email,username,generate_password_hash(password),password,role,1,0,None,now_iso()))
        if request.form.get("add_directory"):
            if not db.execute("SELECT id FROM administrators WHERE name=?", (name,)).fetchone():
                db.execute("INSERT INTO administrators(id,name,position,committee,date,responsibilities) VALUES(?,?,?,?,?,?)",
                           (uid("adm"),name,position,committee,date.today().isoformat(),request.form.get("responsibilities", "")))
        db.commit(); log_action("Created admin account",name); flash("تم إنشاء الحساب", "ok")
        return redirect(url_for("admin_users"))
    return render_template("admin_user_form.html")

@app.route("/admin/users/<uid_>/password", methods=["POST"])
def admin_user_password(uid_):
    if not is_creator():
        flash("تغيير رموز الحسابات متاح لصانع التطبيق فقط", "error")
        return redirect(url_for("admin_users"))
    password=request.form.get("password", "")
    confirm=request.form.get("confirm_password", "")
    if len(password) < 6 or password != confirm:
        flash("الرمز يجب أن يكون 6 أحرف/أرقام على الأقل ومتطابقًا", "error")
        return redirect(url_for("admin_users"))
    db=get_db(); u=db.execute("SELECT * FROM users WHERE id=?",(uid_,)).fetchone()
    if not u:
        return redirect(url_for("admin_users"))
    db.execute("UPDATE users SET password_hash=?,password_plain=?,must_set_password=0 WHERE id=?",(generate_password_hash(password),password,uid_))
    db.commit(); log_action("Reset admin password",u["name"]); flash("تم تعيين الرمز", "ok")
    return redirect(url_for("admin_users"))

@app.route("/admin/users/<uid_>/toggle", methods=["POST"])
def admin_user_toggle(uid_):
    if not is_creator():
        flash("إدارة حالة الحسابات متاحة لصانع التطبيق فقط", "error")
        return redirect(url_for("admin_users"))
    db=get_db(); u=db.execute("SELECT * FROM users WHERE id=?",(uid_,)).fetchone()
    if not u:
        return redirect(url_for("admin_users"))
    if u["role"] == "CREATOR":
        flash("لا يمكن إيقاف حساب صانع التطبيق", "error")
        return redirect(url_for("admin_users"))
    db.execute("UPDATE users SET active=? WHERE id=?", (0 if u["active"] else 1, uid_)); db.commit()
    log_action("Toggled admin account",u["name"]); return redirect(url_for("admin_users"))

@app.route("/news")
def public_news():
    db=get_db(); q=request.args.get("q","").strip(); category=request.args.get("category","").strip()
    now=now_iso(); where=["status='published'","(scheduled_at IS NULL OR scheduled_at='' OR scheduled_at<=?)"]; args=[now]
    if q: where.append("(title LIKE ? OR excerpt LIKE ? OR content LIKE ?)"); like=f"%{q}%"; args += [like,like,like]
    if category: where.append("category=?"); args.append(category)
    sql="SELECT * FROM news WHERE "+" AND ".join(where)+" ORDER BY featured DESC,published_at DESC,created_at DESC"
    rows=db.execute(sql,args).fetchall(); categories=[r["category"] for r in db.execute("SELECT DISTINCT category FROM news WHERE status='published' ORDER BY category").fetchall() if r["category"]]
    featured=rows[0] if rows and not q and not category else None
    grid=rows[1:] if featured else rows
    return render_template("news.html", news=grid, featured_news=featured, categories=categories, q=q, active_category=category)

@app.route("/news/<slug>")
def public_news_detail(slug):
    n=get_db().execute("SELECT * FROM news WHERE slug=? AND status='published' AND (scheduled_at IS NULL OR scheduled_at='' OR scheduled_at<=?)", (slug,now_iso())).fetchone()
    if not n: return redirect(url_for("public_news"))
    return render_template("news_detail.html", n=n)

@app.route("/admin/news")
def news_admin():
    rows=get_db().execute("SELECT * FROM news ORDER BY created_at DESC").fetchall()
    return render_template("news_admin.html", news=rows)

@app.route("/admin/news/new", methods=["GET","POST"])
def news_new():
    if request.method=="POST":
        db=get_db(); title=request.form.get("title","").strip(); slug=request.form.get("slug","").strip().lower().replace(" ","-")
        if not title: flash("اكتب عنوان الخبر", "error"); return redirect(request.url)
        if not slug: slug=uuid.uuid4().hex[:10]
        author=(current_user()["name"] if current_user() else "AHVT")
        status=request.form.get("status","draft")
        featured=1 if request.form.get("featured") else 0; scheduled_at=request.form.get("scheduled_at") or None
        nid=uid("news")
        cover=save_image_as_jpg(request.files.get("cover_file"), "news")
        db.execute("INSERT INTO news(id,title,slug,excerpt,content,cover_image,category,author,status,published_at,created_at,featured,scheduled_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                   (nid,title,slug,request.form.get("excerpt",""),request.form.get("content",""),cover,request.form.get("category","عام"),author,status,now_iso() if status=="published" and not scheduled_at else None,now_iso(),featured,scheduled_at))
        save_multiple_attachments(request.files.getlist("media_files"), "news", nid, db, title)
        db.commit(); log_action("Created news",title); flash("تم حفظ الخبر", "ok"); return redirect(url_for("news_admin"))
    return render_template("news_form.html", n=None)

@app.route("/admin/news/<nid>/edit", methods=["GET","POST"])
def news_edit(nid):
    db=get_db(); n=db.execute("SELECT * FROM news WHERE id=?",(nid,)).fetchone()
    if not n: return redirect(url_for("news_admin"))
    if request.method=="POST":
        status=request.form.get("status","draft")
        featured=1 if request.form.get("featured") else 0; scheduled_at=request.form.get("scheduled_at") or None
        cover=save_image_as_jpg(request.files.get("cover_file"), "news") or n["cover_image"]
        db.execute("UPDATE news SET title=?,slug=?,excerpt=?,content=?,cover_image=?,category=?,status=?,published_at=?,featured=?,scheduled_at=? WHERE id=?",
                   (request.form.get("title",""),request.form.get("slug",""),request.form.get("excerpt",""),request.form.get("content",""),cover,request.form.get("category","عام"),status,now_iso() if status=="published" and not scheduled_at else None,featured,scheduled_at,nid))
        save_multiple_attachments(request.files.getlist("media_files"), "news", nid, db, request.form.get("title",""))
        db.commit(); log_action("Updated news",n["title"]); flash("تم تحديث الخبر", "ok"); return redirect(url_for("news_admin"))
    return render_template("news_form.html", n=n)

@app.route("/admin/news/<nid>/delete", methods=["POST"])
def news_delete(nid):
    db=get_db(); db.execute("DELETE FROM news WHERE id=?",(nid,)); db.commit(); log_action("Deleted news",nid); return redirect(url_for("news_admin"))

@app.route("/page/<slug>")
def public_page(slug):
    p=get_db().execute("SELECT * FROM pages WHERE slug=? AND status='published'",(slug,)).fetchone()
    if not p: return "<h2 style='font-family:Tajawal'>الصفحة غير موجودة</h2>",404
    return render_template("page_public.html", page=p)

@app.route("/admin/pages")
def page_admin():
    pages=get_db().execute("SELECT * FROM pages ORDER BY sort_order,title").fetchall()
    return render_template("pages_admin.html", pages=pages)

@app.route("/admin/pages/new", methods=["GET","POST"])
def page_new():
    if request.method=="POST":
        db=get_db(); title=request.form.get("title","").strip(); slug=request.form.get("slug","").strip().lower().replace(" ","-") or uuid.uuid4().hex[:10]
        db.execute("INSERT INTO pages(id,title,slug,content,status,show_in_nav,sort_order,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?)",
                   (uid("page"),title,slug,request.form.get("content",""),request.form.get("status","draft"),1 if request.form.get("show_in_nav") else 0,int(request.form.get("sort_order",0) or 0),now_iso(),now_iso()))
        db.commit(); log_action("Created page",title); return redirect(url_for("page_admin"))
    return render_template("page_form.html", page=None)

@app.route("/admin/pages/<pid>/edit", methods=["GET","POST"])
def page_edit(pid):
    db=get_db(); p=db.execute("SELECT * FROM pages WHERE id=?",(pid,)).fetchone()
    if not p: return redirect(url_for("page_admin"))
    if request.method=="POST":
        db.execute("UPDATE pages SET title=?,slug=?,content=?,status=?,show_in_nav=?,sort_order=?,updated_at=? WHERE id=?",
                   (request.form.get("title",""),request.form.get("slug",""),request.form.get("content",""),request.form.get("status","draft"),1 if request.form.get("show_in_nav") else 0,int(request.form.get("sort_order",0) or 0),now_iso(),pid))
        db.commit(); log_action("Updated page",p["title"]); return redirect(url_for("page_admin"))
    return render_template("page_form.html", page=p)

@app.route("/admin/pages/<pid>/delete", methods=["POST"])
def page_delete(pid):
    db=get_db(); db.execute("DELETE FROM pages WHERE id=?",(pid,)); db.commit(); return redirect(url_for("page_admin"))

# ============================================================ DASHBOARD ============================================================
@app.route("/")
@app.route("/dashboard")
def public_home():
    db = get_db()
    members = db.execute("SELECT * FROM members").fetchall()
    admins = db.execute("SELECT * FROM administrators").fetchall()
    initiatives = db.execute("SELECT * FROM initiatives").fetchall()

    scores = [member_score(db, m["id"]) for m in members]
    scores = [s for s in scores if s is not None]
    atts = [member_attendance_pct(db, m["id"]) for m in members]
    atts = [a for a in atts if a is not None]
    avg_score = round(sum(scores) / len(scores)) if scores else None
    avg_att = round(sum(atts) / len(atts)) if atts else None
    total_hours = sum((i["hours"] or 0) for i in initiatives)

    months = ["يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو",
              "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"]
    counts = [0] * 12
    for i in initiatives:
        if i["date"]:
            try:
                mi = datetime.fromisoformat(str(i["date"])[:10]).month - 1
                counts[mi] += 1
            except Exception:
                pass
    max_count = max(counts + [1])

    top_members = []
    for m in members:
        s = member_score(db, m["id"])
        if s is not None:
            top_members.append((m, s))
    top_members.sort(key=lambda x: x[1], reverse=True)
    top_members = top_members[:5]

    latest_news = db.execute("SELECT * FROM news WHERE status='published' ORDER BY published_at DESC, created_at DESC LIMIT 5").fetchall()
    committees = db.execute("SELECT * FROM committees ORDER BY name").fetchall()
    upcoming = db.execute("SELECT * FROM initiatives WHERE date IS NOT NULL AND date >= ? ORDER BY date ASC LIMIT 3", (date.today().isoformat(),)).fetchall()
    featured = db.execute("SELECT * FROM initiatives ORDER BY date DESC LIMIT 6").fetchall()
    public_admins = db.execute("SELECT * FROM administrators ORDER BY date ASC LIMIT 6").fetchall()
    featured_media = db.execute("SELECT * FROM media WHERE public=1 ORDER BY sort_order, created_at DESC LIMIT 4").fetchall()
    partners = db.execute("SELECT * FROM partners WHERE active=1 ORDER BY sort_order,name LIMIT 6").fetchall()
    impact_points = db.execute("SELECT id,name,location,latitude,longitude,date,status,map_status,map_note,map_date,map_visible FROM initiatives WHERE latitude IS NOT NULL AND longitude IS NOT NULL AND COALESCE(map_visible,1)=1 ORDER BY date DESC").fetchall()
    live_session_id=get_setting("liveSessionId", "")
    live_session=db.execute("SELECT ls.*,i.name initiative_name FROM live_sessions ls LEFT JOIN initiatives i ON i.id=ls.initiative_id WHERE ls.id=? AND ls.status='live'",(live_session_id,)).fetchone() if live_session_id else None
    return render_template("public_home.html",
        members_count=len(members), admins_count=len(admins), initiatives_count=len(initiatives),
        total_hours=total_hours, avg_score=avg_score, avg_att=avg_att, latest_news=latest_news,
        months=months, counts=counts, max_count=max_count, top_members=top_members,
        committees=committees, upcoming=upcoming, featured=featured, public_admins=public_admins, featured_media=featured_media, partners=partners, impact_points=impact_points, live_session=live_session, live_session_id=live_session_id, live_mode=get_setting("liveMode","internal"))



# ============================================================ PUBLIC EXPANSION / SEARCH / IMPACT ============================================================
@app.route("/about")
def about_public():
    return render_template("public_info.html", page_title="عن AHVT", page_kicker="ABOUT", page_text=get_setting("siteDescription", "منصة AHVT لإدارة الأثر، المبادرات، والأداء التطوعي."))

@app.route("/impact")
def impact_public():
    db=get_db()
    members=db.execute("SELECT COUNT(*) c FROM members").fetchone()["c"]
    initiatives=db.execute("SELECT COUNT(*) c FROM initiatives").fetchone()["c"]
    hours=db.execute("SELECT COALESCE(SUM(hours),0) h FROM initiatives").fetchone()["h"] or 0
    admins=db.execute("SELECT COUNT(*) c FROM administrators").fetchone()["c"]
    goals=db.execute("SELECT * FROM goals WHERE status='active' ORDER BY created_at DESC").fetchall()
    points=db.execute("SELECT id,name,location,latitude,longitude,date,status,map_status,map_note,map_date,map_visible FROM initiatives WHERE latitude IS NOT NULL AND longitude IS NOT NULL AND COALESCE(map_visible,1)=1 ORDER BY date DESC").fetchall()
    return render_template("impact.html", members=members, initiatives=initiatives, hours=round(hours), admins=admins, goals=goals, points=points)

@app.route("/impact-map")
def impact_map_public():
    points=get_db().execute("SELECT id,name,location,latitude,longitude,date,status,map_status,map_note,map_date,map_visible FROM initiatives WHERE latitude IS NOT NULL AND longitude IS NOT NULL AND COALESCE(map_visible,1)=1 ORDER BY date DESC").fetchall()
    return render_template("impact_map.html", points=points)

@app.route("/live")
def live_public():
    db=get_db()
    session_id=get_setting("liveSessionId", "")
    active=None
    if session_id:
        active=db.execute("SELECT ls.*, i.name initiative_name FROM live_sessions ls LEFT JOIN initiatives i ON i.id=ls.initiative_id WHERE ls.id=? AND ls.status='live'", (session_id,)).fetchone()
    active_sessions=db.execute("SELECT ls.*,i.name initiative_name FROM live_sessions ls LEFT JOIN initiatives i ON i.id=ls.initiative_id WHERE ls.status='live' ORDER BY ls.started_at DESC LIMIT 6").fetchall()
    return render_template("live.html", live_session=active, active_sessions=active_sessions, live_url=get_setting("liveUrl", ""), live_title=get_setting("liveTitle", "البث المباشر من الميدان"), live_description=get_setting("liveDescription", "تابع مبادرات فريق الحكمة التطوعي مباشرةً."), live_platform=get_setting("livePlatform", "YouTube"), live_enabled=get_setting("liveEnabled", "0"), live_mode=get_setting("liveMode", "internal"), live_embed_url=get_live_embed_url(get_setting("liveUrl", "")))

@app.route("/live/room/<sid>")
def live_room(sid):
    db=get_db()
    ls=db.execute("SELECT ls.*, i.name initiative_name FROM live_sessions ls LEFT JOIN initiatives i ON i.id=ls.initiative_id WHERE ls.id=? AND ls.status='live'", (sid,)).fetchone()
    if not ls: return redirect(url_for("live_public"))
    u=current_user()
    approved=False
    if u:
        approved=bool(db.execute("SELECT 1 FROM live_source_requests WHERE session_id=? AND requester_user_id=? AND status='approved' ORDER BY created_at DESC LIMIT 1",(sid,u["id"])).fetchone())
    return render_template("live_room.html", live_session=ls, is_broadcaster=is_admin(), source_mode=request.args.get("source") == "1", source_approved=approved)

@app.route("/live/request-source/<sid>", methods=["POST"])
def live_request_source(sid):
    db=get_db()
    ls=db.execute("SELECT id,title FROM live_sessions WHERE id=? AND status='live'",(sid,)).fetchone()
    if not ls: flash("البث غير موجود أو انتهى","error"); return redirect(url_for("live_public"))
    u=current_user()
    if not u:
        flash("يجب تسجيل الدخول لإرسال طلب مصدر بث.","error")
        return redirect(url_for("live_room",sid=sid))
    db.execute("INSERT INTO live_source_requests(id,session_id,requester_name,requester_user_id,device_label,status,created_at) VALUES(?,?,?,?,?,?,?)",
               (uid("src_req"),sid,u["name"],u["id"],request.form.get("device_label","هاتف/كاميرا"),"pending",now_iso()))
    db.commit()
    flash("تم إرسال طلب ربط الجهاز، بانتظار موافقة الإدارة.","ok")
    return redirect(url_for("live_room",sid=sid))
@app.route("/api/live/session/<sid>/join", methods=["POST"])
def live_join(sid):
    db=get_db(); ls=db.execute("SELECT * FROM live_sessions WHERE id=? AND status='live'",(sid,)).fetchone()
    if not ls: return jsonify({"ok":False,"error":"Live session not found"}),404
    data=request.get_json(silent=True) or {}; peer_id=(data.get("peer_id") or "").strip(); role=(data.get("role") or "viewer").strip()
    label=(data.get("label") or "").strip()
    source_index=int(data.get("source_index") or 0)
    if not peer_id: return jsonify({"ok":False,"error":"peer_id required"}),400
    if role in ("broadcaster","source") and not is_admin():
        u=current_user()
        approved=bool(u and db.execute("SELECT 1 FROM live_source_requests WHERE session_id=? AND requester_user_id=? AND status='approved' ORDER BY created_at DESC LIMIT 1",(sid,u["id"])).fetchone())
        if role!="source" or not approved:
            return jsonify({"ok":False,"error":"المصدر يحتاج موافقة الإدارة أولاً"}),403
    if role in ("broadcaster","source"):
        active_count=db.execute("SELECT COUNT(*) c FROM live_peers WHERE session_id=? AND role IN ('broadcaster','source')",(sid,)).fetchone()["c"]
        if active_count >= 6 and not db.execute("SELECT 1 FROM live_peers WHERE session_id=? AND peer_id=?",(sid,peer_id)).fetchone():
            return jsonify({"ok":False,"error":"وصل الحد الأقصى 6 مصادر/كاميرات لهذا البث"}),409
    now=now_iso(); db.execute("INSERT INTO live_peers(id,session_id,peer_id,role,created_at,last_seen,label,source_index) VALUES(?,?,?,?,?,?,?,?) ON CONFLICT(session_id,peer_id) DO UPDATE SET role=excluded.role,last_seen=excluded.last_seen,label=excluded.label,source_index=excluded.source_index",(uid("peer"),sid,peer_id,role,now,now,label,source_index)); db.commit()
    return jsonify({"ok":True,"session":sid,"peer_id":peer_id,"role":role})

@app.route("/api/live/session/<sid>/peers")
def live_peers_api(sid):
    peer_id=request.args.get("peer_id",""); db=get_db(); ls=db.execute("SELECT id FROM live_sessions WHERE id=? AND status='live'",(sid,)).fetchone()
    if not ls:return jsonify({"ok":False}),404
    db.execute("UPDATE live_peers SET last_seen=? WHERE session_id=? AND peer_id=?",(now_iso(),sid,peer_id))
    db.execute("DELETE FROM live_peers WHERE session_id=? AND last_seen<?",(sid,datetime.utcnow().replace(microsecond=0).isoformat()))
    viewer_count=db.execute("SELECT COUNT(*) c FROM live_peers WHERE session_id=? AND role='viewer'",(sid,)).fetchone()["c"]
    db.execute("UPDATE live_sessions SET viewer_peak=MAX(COALESCE(viewer_peak,0),?) WHERE id=?",(viewer_count,sid))
    db.commit()
    rows=db.execute("SELECT peer_id,role,label,source_index FROM live_peers WHERE session_id=? AND peer_id<>? ORDER BY created_at",(sid,peer_id)).fetchall()
    return jsonify({"ok":True,"peers":[dict(r) for r in rows]})

@app.route("/api/live/session/<sid>/signal", methods=["GET","POST"])
def live_signal(sid):
    db=get_db(); ls=db.execute("SELECT id FROM live_sessions WHERE id=? AND status='live'",(sid,)).fetchone()
    if not ls:return jsonify({"ok":False}),404
    if request.method=="POST":
        data=request.get_json(silent=True) or {}; sender=data.get("sender_id"); target=data.get("target_id"); kind=data.get("kind"); payload=data.get("payload")
        if not sender or not target or kind not in ("offer","answer","ice"): return jsonify({"ok":False,"error":"Invalid signal"}),400
        db.execute("INSERT INTO live_signals(id,session_id,sender_id,target_id,kind,payload,created_at,delivered_at) VALUES(?,?,?,?,?,?,?,NULL)",(uid("sig"),sid,sender,target,kind,json.dumps(payload),now_iso())); db.commit(); return jsonify({"ok":True})
    peer=request.args.get("peer_id",""); rows=db.execute("SELECT id,sender_id,kind,payload FROM live_signals WHERE session_id=? AND target_id=? AND delivered_at IS NULL ORDER BY created_at LIMIT 100",(sid,peer)).fetchall()
    ids=[r["id"] for r in rows]
    if ids: db.executemany("UPDATE live_signals SET delivered_at=? WHERE id=?",[(now_iso(),i) for i in ids]); db.commit()
    return jsonify({"ok":True,"signals":[{"id":r["id"],"sender_id":r["sender_id"],"kind":r["kind"],"payload":json.loads(r["payload"])} for r in rows]})

@app.route("/api/live/session/<sid>/leave", methods=["POST"])
def live_leave(sid):
    data=request.get_json(silent=True) or {}; peer_id=data.get("peer_id",""); db=get_db(); db.execute("DELETE FROM live_peers WHERE session_id=? AND peer_id=?",(sid,peer_id)); db.commit(); return jsonify({"ok":True})

@app.route("/events")
def events_public():
    db=get_db(); rows=db.execute("SELECT * FROM events ORDER BY date ASC, time ASC").fetchall()
    return render_template("events.html", events=rows)

@app.route("/media")
def media_public():
    rows=get_db().execute("SELECT m.*, i.name initiative_name FROM media m LEFT JOIN initiatives i ON i.id=m.initiative_id WHERE m.public=1 ORDER BY m.created_at DESC").fetchall()
    return render_template("media.html", media=rows)

@app.route("/partners")
def partners_public():
    rows=get_db().execute("SELECT * FROM partners WHERE active=1 ORDER BY sort_order, name").fetchall()
    return render_template("partners.html", partners=rows)

@app.route("/volunteer")
def volunteer_redirect():
    return redirect(get_setting("telegramUrl", "https://t.me/Hikmaht_bot"))

@app.route("/temporary-volunteers")
def temporary_volunteers_public():
    db=get_db()
    rows=db.execute("SELECT * FROM temporary_volunteers WHERE status='active' AND card_public=1 ORDER BY created_at DESC").fetchall()
    return render_template("temporary_volunteers.html", volunteers=rows)

@app.route("/temporary-volunteers/card/<vid>")
def temporary_volunteer_card(vid):
    db=get_db(); v=db.execute("SELECT * FROM temporary_volunteers WHERE id=?",(vid,)).fetchone()
    if not v: return "البطاقة غير موجودة",404
    if not v["card_public"] and not is_admin(): return "البطاقة خاصة",403
    return render_template("temporary_volunteer_card.html", volunteer=v)

@app.route("/admin/temporary-volunteers", methods=["GET","POST"])
def temporary_volunteers_admin():
    if not is_admin(): return redirect(url_for("admin_login"))
    db=get_db()
    if request.method=="POST":
        action=request.form.get("action","create")
        if action=="delete":
            vid=request.form.get("id"); db.execute("DELETE FROM temporary_volunteers WHERE id=?",(vid,)); db.commit(); log_action("Deleted temporary volunteer",vid); flash("تم حذف المتطوع الوقتي","ok"); return redirect(url_for("temporary_volunteers_admin"))
        photo=save_image_as_jpg(request.files.get("photo"),"temporary-volunteer")
        vid=uid("tv")
        db.execute("""INSERT INTO temporary_volunteers(id,name,phone,department,stage,specialty,task,committee,start_date,end_date,hours,benefits,status,photo,card_public,created_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(vid,request.form.get("name","").strip(),request.form.get("phone",""),request.form.get("department",""),request.form.get("stage",""),request.form.get("specialty",""),request.form.get("task",""),request.form.get("committee",""),request.form.get("start_date") or None,request.form.get("end_date") or None,float(request.form.get("hours",0) or 0),request.form.get("benefits",""),"active",photo,1 if request.form.get("card_public") else 0,now_iso()))
        db.commit(); log_action("Created temporary volunteer",vid); flash("تمت إضافة متطوع وقتي وإصدار بيانات بطاقته","ok"); return redirect(url_for("temporary_volunteers_admin"))
    rows=db.execute("SELECT * FROM temporary_volunteers ORDER BY created_at DESC").fetchall()
    return render_template("temporary_volunteers_admin.html", volunteers=rows)

@app.route("/search")
def search_public():
    q=request.args.get("q","").strip()
    results=[]
    if q:
        like=f"%{q}%"; db=get_db()
        for r in db.execute("SELECT id,title,slug,excerpt FROM news WHERE status='published' AND (title LIKE ? OR excerpt LIKE ? OR content LIKE ?) ORDER BY published_at DESC LIMIT 10",(like,like,like)).fetchall():
            results.append({"kind":"خبر","title":r["title"],"text":r["excerpt"] or "","url":url_for("public_news_detail",slug=r["slug"])})
        for r in db.execute("SELECT id,name,description FROM initiatives WHERE name LIKE ? OR description LIKE ? ORDER BY date DESC LIMIT 10",(like,like)).fetchall():
            results.append({"kind":"مبادرة","title":r["name"],"text":r["description"] or "","url":url_for("initiative_view",iid=r["id"])})
        for r in db.execute("SELECT id,name,position,committee FROM administrators WHERE name LIKE ? OR position LIKE ? OR committee LIKE ? LIMIT 10",(like,like,like)).fetchall():
            results.append({"kind":"إداري","title":r["name"],"text":r["position"] or r["committee"] or "","url":url_for("admins_list")+"#admin-"+r["id"]})
        for r in db.execute("SELECT id,name,description FROM committees WHERE name LIKE ? OR description LIKE ? LIMIT 10",(like,like)).fetchall():
            results.append({"kind":"لجنة","title":r["name"],"text":r["description"] or "","url":url_for("committees_list")})
    return render_template("search.html", q=q, results=results)

@app.route("/api/impact")
def api_impact():
    db=get_db()
    return jsonify({"ok":True,"team":get_setting("teamName","AHVT"),"members":db.execute("SELECT COUNT(*) c FROM members").fetchone()["c"],"initiatives":db.execute("SELECT COUNT(*) c FROM initiatives").fetchone()["c"],"hours":db.execute("SELECT COALESCE(SUM(hours),0) h FROM initiatives").fetchone()["h"] or 0,"news":db.execute("SELECT COUNT(*) c FROM news WHERE status='published'").fetchone()["c"]})

# ============================================================ ADMIN EXPANSION ============================================================
def notify_role(role, title, body, target_url=None):
    db=get_db(); users=db.execute("SELECT id FROM users WHERE role=? AND active=1",(role,)).fetchall()
    for u in users:
        db.execute("INSERT INTO notifications(id,user_id,target_role,type,title,body,url,read_at,created_at) VALUES(?,?,?,?,?,?,?,?,?)",(uid("ntf"),u["id"],role,"system",title,body,target_url,None,now_iso()))
    db.commit()

def notify_user(user_id,title,body,target_url=None,kind="system"):
    db=get_db(); db.execute("INSERT INTO notifications(id,user_id,target_role,type,title,body,url,read_at,created_at) VALUES(?,?,?,?,?,?,?,?,?)",(uid("ntf"),user_id,None,kind,title,body,target_url,None,now_iso())); db.commit()

@app.route("/notifications")
def notifications_page():
    u=current_user()
    if not u: return redirect(url_for("admin_login"))
    rows=get_db().execute("SELECT * FROM notifications WHERE user_id=? ORDER BY created_at DESC LIMIT 100",(u["id"],)).fetchall()
    return render_template("notifications.html", notifications=rows)

@app.route("/notifications/read/<nid>", methods=["POST"])
def notification_read(nid):
    u=current_user(); db=get_db(); db.execute("UPDATE notifications SET read_at=? WHERE id=? AND user_id=?",(now_iso(),nid,u["id"])); db.commit(); return jsonify({"ok":True})

@app.route("/notifications/read-all", methods=["POST"])
def notification_read_all():
    u=current_user(); get_db().execute("UPDATE notifications SET read_at=? WHERE user_id=? AND read_at IS NULL",(now_iso(),u["id"])); get_db().commit(); return jsonify({"ok":True})

@app.route("/admin/analytics")
def analytics_page():
    db=get_db(); months=[]
    for i in range(11,-1,-1):
        y=date.today().year; m=date.today().month-i
        while m<=0: y-=1; m+=12
        prefix=f"{y:04d}-{m:02d}"
        months.append({"label":prefix,"initiatives":db.execute("SELECT COUNT(*) c FROM initiatives WHERE date LIKE ?",(prefix+"%",)).fetchone()["c"],"news":db.execute("SELECT COUNT(*) c FROM news WHERE status='published' AND published_at LIKE ?",(prefix+"%",)).fetchone()["c"]})
    committees=[]
    for c in db.execute("SELECT * FROM committees ORDER BY name").fetchall():
        n=c["name"]; committees.append({"name":n,"members":db.execute("SELECT COUNT(*) c FROM members WHERE committee=?",(n,)).fetchone()["c"],"initiatives":db.execute("SELECT COUNT(*) c FROM initiatives WHERE committee=?",(n,)).fetchone()["c"]})
    return render_template("analytics.html",months=months,committees=committees)

@app.route("/admin/insights")
def insights_page():
    db=get_db(); insights=[]
    overdue=db.execute("SELECT COUNT(*) c FROM tasks WHERE deadline IS NOT NULL AND deadline < ? AND lower(status) NOT IN ('done','completed','مكتملة','منجزة')",(date.today().isoformat(),)).fetchone()["c"]
    if overdue: insights.append(("🚨","تنبيه تشغيلي",f"هناك {overdue} مهام متأخرة تحتاج متابعة."))
    members=db.execute("SELECT COUNT(*) c FROM members").fetchone()["c"]; initiatives=db.execute("SELECT COUNT(*) c FROM initiatives").fetchone()["c"]
    if members and initiatives: insights.append(("📈","فرصة",f"لدى الفريق {members} عضو مقابل {initiatives} مبادرة موثقة."))
    news=db.execute("SELECT COUNT(*) c FROM news WHERE status='published'").fetchone()["c"]
    if news==0: insights.append(("📰","فرصة إعلامية","لا توجد أخبار منشورة حاليًا؛ تفعيل مركز الأخبار سيحسن الحضور العام."))
    if not insights: insights.append(("✅","الوضع مستقر","لا توجد إشارات تشغيلية حرجة حاليًا."))
    return render_template("insights.html",insights=insights)

@app.route("/admin/risk")
def risk_page():
    db=get_db(); risks=[]
    overdue=db.execute("SELECT COUNT(*) c FROM tasks WHERE deadline IS NOT NULL AND deadline < ? AND lower(status) NOT IN ('done','completed','مكتملة','منجزة')",(date.today().isoformat(),)).fetchone()["c"]
    pending=db.execute("SELECT COUNT(*) c FROM approvals WHERE status='pending'").fetchone()["c"]
    if overdue: risks.append(("عالي","مهام متأخرة",f"{overdue} مهمة متأخرة."))
    if pending: risks.append(("متوسط","موافقات معلقة",f"{pending} طلب موافقة بانتظار المراجعة."))
    if not risks: risks.append(("منخفض","لا توجد مخاطر حرجة","الوضع التشغيلي مستقر حسب البيانات الحالية."))
    return render_template("risk.html",risks=risks)

@app.route("/admin/decisions")
def decision_center():
    db=get_db(); decisions=db.execute("SELECT * FROM approvals WHERE status='pending' ORDER BY created_at DESC").fetchall()
    return render_template("decisions.html",decisions=decisions)

@app.route("/admin/goals", methods=["GET","POST"])
def goals_page():
    db=get_db()
    if request.method=="POST":
        title=request.form.get("title","").strip(); target=float(request.form.get("target",0) or 0); period=request.form.get("period","")
        if title: db.execute("INSERT INTO goals(id,title,target,current,period,status,created_at) VALUES(?,?,?,?,?,?,?)",(uid("goal"),title,target,0,period,"active",now_iso())); db.commit(); log_action("Created goal",title)
        return redirect(url_for("goals_page"))
    goals=db.execute("SELECT * FROM goals ORDER BY created_at DESC").fetchall(); return render_template("goals.html",goals=goals)

@app.route("/admin/goals/<gid>/progress", methods=["POST"])
def goal_progress(gid):
    val=float(request.form.get("current",0) or 0); db=get_db(); db.execute("UPDATE goals SET current=? WHERE id=?",(val,gid)); db.commit(); log_action("Updated goal",gid); return redirect(url_for("goals_page"))

@app.route("/admin/approvals")
def approvals_page():
    rows=get_db().execute("SELECT * FROM approvals ORDER BY created_at DESC").fetchall(); return render_template("approvals.html", approvals=rows)

@app.route("/admin/approvals/<aid>/review", methods=["POST"])
def approval_review(aid):
    status=request.form.get("status","approved"); notes=request.form.get("notes",""); u=current_user(); db=get_db(); db.execute("UPDATE approvals SET status=?,reviewed_by=?,notes=?,reviewed_at=? WHERE id=?",(status,u["name"],notes,now_iso(),aid)); db.commit(); log_action("Reviewed approval",aid); return redirect(url_for("approvals_page"))

@app.route("/admin/events", methods=["GET","POST"])
def events_admin():
    db=get_db()
    if request.method=="POST":
        eid=uid("evt")
        cover=save_image_as_jpg(request.files.get("cover_file"), "event")
        db.execute("INSERT INTO events(id,title,date,time,location,description,status,cover_image,created_at,qr_path) VALUES(?,?,?,?,?,?,?,?,?,?)",(eid,request.form.get("title"),request.form.get("date"),request.form.get("time"),request.form.get("location"),request.form.get("description"),request.form.get("status","upcoming"),cover,now_iso(),_action_qr("event",eid,"action")))
        save_multiple_attachments(request.files.getlist("media_files"), "event", eid, db, request.form.get("title"))
        db.commit(); log_action("Created event",request.form.get("title")); return redirect(url_for("events_admin"))
    return render_template("events_admin.html",events=db.execute("SELECT * FROM events ORDER BY date DESC").fetchall())

@app.route("/admin/events/<eid>/delete", methods=["POST"])
def event_delete(eid):
    db=get_db(); db.execute("DELETE FROM events WHERE id=?",(eid,)); db.commit(); log_action("Deleted event",eid); return redirect(url_for("events_admin"))

@app.route("/admin/partners", methods=["GET","POST"])
def partners_admin():
    db=get_db()
    if request.method=="POST":
        pid=uid("pr")
        logo=save_image_as_jpg(request.files.get("logo_file"), "partner")
        db.execute("INSERT INTO partners(id,name,description,logo,url,sort_order,active) VALUES(?,?,?,?,?,?,?)",(pid,request.form.get("name"),request.form.get("description"),logo,request.form.get("url"),int(request.form.get("sort_order",0) or 0),1))
        save_multiple_attachments(request.files.getlist("media_files"), "partner", pid, db, request.form.get("name"))
        db.commit(); log_action("Created partner",request.form.get("name")); return redirect(url_for("partners_admin"))
    return render_template("partners_admin.html",partners=db.execute("SELECT * FROM partners ORDER BY sort_order,name").fetchall())

@app.route("/admin/media", methods=["GET","POST"])
def media_admin():
    db=get_db()
    if request.method=="POST":
        db.execute("INSERT INTO media(id,title,url,category,initiative_id,public,created_at) VALUES(?,?,?,?,?,?,?)",(uid("med"),request.form.get("title"),request.form.get("url"),request.form.get("category"),request.form.get("initiative_id") or None,1,now_iso())); db.commit(); log_action("Added media",request.form.get("title")); return redirect(url_for("media_admin"))
    return render_template("media_admin.html",media=db.execute("SELECT m.*,i.name initiative_name FROM media m LEFT JOIN initiatives i ON i.id=m.initiative_id ORDER BY m.created_at DESC").fetchall(),initiatives=db.execute("SELECT * FROM initiatives ORDER BY date DESC").fetchall())

@app.route("/admin/security")
def security_page():
    db=get_db(); rows=db.execute("SELECT s.*,u.name FROM security_sessions s LEFT JOIN users u ON u.id=s.user_id ORDER BY last_seen DESC LIMIT 100").fetchall(); return render_template("security.html",sessions=rows)

@app.route("/certificate/<mid>")
def certificate_legacy(mid):
    if not is_admin(): return redirect(url_for("admin_login"))
    return redirect(url_for("certificates_admin"))


# ============================================================ MEMBERS ============================================================
@app.route("/members")
def members_list():
    db = get_db()
    q = request.args.get("q", "").strip().lower()
    members = db.execute("SELECT * FROM members ORDER BY created_at DESC").fetchall()
    if q:
        members = [m for m in members if q in (m["name"] or "").lower()
                   or q in (m["committee"] or "").lower() or q in (m["position"] or "").lower()]
    data = []
    for m in members:
        data.append({
            "m": m, "score": member_score(db, m["id"]), "points": member_points_total(db, m["id"]),
            "att": member_attendance_pct(db, m["id"]), "level": member_level(db, m["id"]),
            "ini_count": member_initiative_count(db, m["id"])
        })
    return render_template("members_list.html", data=data, q=request.args.get("q", ""))


@app.route("/members/new", methods=["GET", "POST"])
def member_new():
    return member_form()


@app.route("/members/<mid>/edit", methods=["GET", "POST"])
def member_edit(mid):
    return member_form(mid)


def member_form(mid=None):
    db = get_db()
    committees = db.execute("SELECT * FROM committees").fetchall()
    member = db.execute("SELECT * FROM members WHERE id=?", (mid,)).fetchone() if mid else None
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("يرجى إدخال اسم العضو", "error")
            return redirect(request.url)
        old_photo = member["photo"] if member else ""
        photo = save_image_as_jpg(request.files.get("photo_file"), "member") or old_photo
        membership_no = request.form.get("membership_no", "").strip() or (member["membership_no"] if member and member["membership_no"] else f"HIKMA-M-{uuid.uuid4().hex[:8].upper()}")
        data = (name, request.form.get("email", ""), request.form.get("phone", ""),
                request.form.get("committee", ""), request.form.get("position", "عضو"),
                request.form.get("join_date", ""), request.form.get("status", "Active"),
                request.form.get("notes", ""), photo, request.form.get("department", ""), request.form.get("stage", ""),
                membership_no, request.form.get("card_issue_date", "") or (member["card_issue_date"] if member else date.today().isoformat()),
                request.form.get("card_expiry", "") or (member["card_expiry"] if member else ""), request.form.get("nfc_uid", "").strip(), 1 if request.form.get("card_public") else 0)
        if mid:
            db.execute("UPDATE members SET name=?,email=?,phone=?,committee=?,position=?,join_date=?,status=?,notes=?,photo=?,department=?,stage=?,membership_no=?,card_issue_date=?,card_expiry=?,nfc_uid=?,card_public=? WHERE id=?", data + (mid,))
            log_action("Updated", f"عضو: {name}")
        else:
            db.execute("INSERT INTO members(id,name,email,phone,committee,position,join_date,status,notes,photo,department,stage,membership_no,card_issue_date,card_expiry,nfc_uid,card_public,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (uid("mem"),) + data + (now_iso(),))
            log_action("Created", f"عضو: {name}")
        db.commit()
        flash("تم الحفظ بنجاح", "ok")
        return redirect(url_for("members_list"))
    return render_template("member_form.html", member=member, committees=committees)


@app.route("/members/<mid>/delete", methods=["POST"])
def member_delete(mid):
    db = get_db()
    m = db.execute("SELECT * FROM members WHERE id=?", (mid,)).fetchone()
    if m:
        db.execute("DELETE FROM members WHERE id=?", (mid,))
        db.execute("DELETE FROM initiative_participants WHERE member_id=?", (mid,))
        db.commit()
        log_action("Deleted", f"عضو: {m['name']}")
        flash("تم الحذف", "ok")
    return redirect(url_for("members_list"))


@app.route("/members/<mid>")
def member_view(mid):
    db = get_db()
    m = db.execute("SELECT * FROM members WHERE id=?", (mid,)).fetchone()
    if not m:
        return redirect(url_for("members_list"))
    s = member_score(db, mid)
    a = member_attendance_pct(db, mid)
    p = member_points_total(db, mid)
    weights = get_weights()
    evs = db.execute("SELECT * FROM evaluations WHERE evaluated_user_id=? ORDER BY date", (mid,)).fetchall()

    breakdown = []
    if s is not None:
        for k in CRITERIA_KEYS:
            cavg = sum((e[f"c_{k}"] or 0) for e in evs) / len(evs)
            breakdown.append({"label": CRITERIA_LABELS[k], "value": round(cavg), "weight": weights.get(k, 0)})

    timeline = []
    for e in evs:
        timeline.append({"date": e["date"], "title": "تقييم جديد", "sub": f"النوع: {e['type']}"})
    pts = db.execute("SELECT * FROM points WHERE member_id=?", (mid,)).fetchall()
    for pt in pts:
        timeline.append({"date": pt["date"], "title": "إضافة نقاط", "sub": f"+{pt['value']} — {pt['source'] or ''}"})
    timeline.append({"date": m["created_at"], "title": "انضمام العضو", "sub": m["committee"] or ""})
    timeline = [t for t in timeline if t["date"]]
    timeline.sort(key=lambda x: x["date"])

    ini_count = member_initiative_count(db, mid)
    volunteer_hours = member_volunteer_hours(db, mid)

    return render_template("member_view.html", m=m, score=s, att=a, points=p, volunteer_hours=volunteer_hours,
        level=member_level(db, mid), breakdown=breakdown, timeline=timeline,
        eval_count=len(evs), ini_count=ini_count, recommendation=recommendation_text(s))


@app.route("/admin/members/<mid>/photo", methods=["POST"])
def member_photo_upload(mid):
    f=request.files.get("photo_file")
    if not f or not f.filename: flash("اختر صورة", "error"); return redirect(url_for("member_view", mid=mid))
    path=save_image_as_jpg(f, "member")
    if not path: flash("تعذر حفظ الصورة", "error"); return redirect(url_for("member_view", mid=mid))
    db=get_db(); db.execute("UPDATE members SET photo=? WHERE id=?", (path,mid)); db.commit(); log_action("Updated member photo",mid)
    flash("تم تحديث الصورة الشخصية", "ok"); return redirect(url_for("member_view", mid=mid))

# ============================================================ ADMINISTRATORS ============================================================
@app.route("/administrators")
def admins_list():
    db = get_db()
    admins = db.execute("SELECT * FROM administrators ORDER BY name").fetchall()
    return render_template("administrators_list.html", admins=admins)


@app.route("/administrators/new", methods=["GET", "POST"])
def admin_new():
    return admin_form()


@app.route("/administrators/<aid>/edit", methods=["GET", "POST"])
def admin_edit(aid):
    return admin_form(aid)


def admin_form(aid=None):
    db = get_db()
    committees = db.execute("SELECT * FROM committees").fetchall()
    a = db.execute("SELECT * FROM administrators WHERE id=?", (aid,)).fetchone() if aid else None
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("يرجى إدخال الاسم", "error")
            return redirect(request.url)
        old_photo = a["photo"] if a else ""
        photo = save_image_as_jpg(request.files.get("photo_file"), "admin") or old_photo
        membership_no = request.form.get("membership_no", "").strip() or (a["membership_no"] if a and a["membership_no"] else f"HIKMA-A-{uuid.uuid4().hex[:8].upper()}")
        data = (name, request.form.get("position", ""), request.form.get("committee", ""),
                request.form.get("date", ""), request.form.get("responsibilities", ""), photo,
                request.form.get("department", ""), membership_no,
                request.form.get("card_issue_date", "") or (a["card_issue_date"] if a else date.today().isoformat()),
                request.form.get("card_expiry", "") or (a["card_expiry"] if a else ""), request.form.get("nfc_uid", "").strip(), 1 if request.form.get("card_public") else 0)
        if aid:
            db.execute("UPDATE administrators SET name=?,position=?,committee=?,date=?,responsibilities=?,photo=?,department=?,membership_no=?,card_issue_date=?,card_expiry=?,nfc_uid=?,card_public=? WHERE id=?", data + (aid,))
        else:
            db.execute("INSERT INTO administrators(id,name,position,committee,date,responsibilities,photo,department,membership_no,card_issue_date,card_expiry,nfc_uid,card_public) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)", (uid("adm"),) + data)
        db.commit()
        log_action("Updated" if aid else "Created", f"إداري: {name}")
        flash("تم الحفظ بنجاح", "ok")
        return redirect(url_for("admins_list"))
    return render_template("admin_form.html", a=a, committees=committees)


@app.route("/administrators/<aid>/delete", methods=["POST"])
def admin_delete(aid):
    db = get_db()
    db.execute("DELETE FROM administrators WHERE id=?", (aid,))
    db.commit()
    flash("تم الحذف", "ok")
    return redirect(url_for("admins_list"))


# ============================================================ COMMITTEES ============================================================
@app.route("/committees")
def committees_list():
    db = get_db()
    committees = db.execute("SELECT * FROM committees ORDER BY name").fetchall()
    data = []
    for c in committees:
        cnt = db.execute("SELECT COUNT(*) c FROM members WHERE committee=?", (c["name"],)).fetchone()["c"]
        data.append({"c": c, "member_count": cnt})
    return render_template("committees_list.html", data=data)


@app.route("/committees/new", methods=["GET", "POST"])
def committee_new():
    return committee_form()


@app.route("/committees/<cid>/edit", methods=["GET", "POST"])
def committee_edit(cid):
    return committee_form(cid)


def committee_form(cid=None):
    db = get_db()
    c = db.execute("SELECT * FROM committees WHERE id=?", (cid,)).fetchone() if cid else None
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("أدخل اسم اللجنة", "error")
            return redirect(request.url)
        data = (name, request.form.get("head", ""), request.form.get("description", ""))
        if cid:
            db.execute("UPDATE committees SET name=?,head=?,description=? WHERE id=?", data + (cid,))
        else:
            db.execute("INSERT INTO committees(id,name,head,description) VALUES(?,?,?,?)", (uid("com"),) + data)
        db.commit()
        log_action("Updated" if cid else "Created", f"لجنة: {name}")
        flash("تم الحفظ بنجاح", "ok")
        return redirect(url_for("committees_list"))
    return render_template("committee_form.html", c=c)


@app.route("/committees/<cid>/delete", methods=["POST"])
def committee_delete(cid):
    db = get_db()
    db.execute("DELETE FROM committees WHERE id=?", (cid,))
    db.commit()
    flash("تم الحذف", "ok")
    return redirect(url_for("committees_list"))


# ============================================================ INITIATIVES ============================================================
@app.route("/initiatives")
def initiatives_list():
    db = get_db()
    initiatives = db.execute("SELECT * FROM initiatives ORDER BY date DESC").fetchall()
    data = [{"i": i, "pcount": initiative_participant_count(db, i["id"])} for i in initiatives]
    return render_template("initiatives_list.html", data=data)


@app.route("/initiatives/new", methods=["GET", "POST"])
def initiative_new():
    return initiative_form()


@app.route("/initiatives/<iid>/edit", methods=["GET", "POST"])
def initiative_edit(iid):
    return initiative_form(iid)


def initiative_form(iid=None):
    db = get_db()
    committees = db.execute("SELECT * FROM committees").fetchall()
    members = db.execute("SELECT * FROM members ORDER BY name").fetchall()
    i = db.execute("SELECT * FROM initiatives WHERE id=?", (iid,)).fetchone() if iid else None
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("أدخل اسم المبادرة", "error")
            return redirect(request.url)
        lat_raw=request.form.get("latitude", "").strip()
        lng_raw=request.form.get("longitude", "").strip()
        try: lat=float(lat_raw) if lat_raw else None
        except Exception: lat=None
        try: lng=float(lng_raw) if lng_raw else None
        except Exception: lng=None
        data = (name, request.form.get("date", ""), request.form.get("location", ""),
                request.form.get("manager", ""), request.form.get("committee", ""),
                float(request.form.get("hours") or 0), request.form.get("status", "Planned"),
                request.form.get("description", ""), request.form.get("goals", ""), lat, lng, request.form.get("map_status","completed"), request.form.get("map_note",""), request.form.get("map_date","") or None, 1 if request.form.get("map_visible") else 0)
        if iid:
            db.execute("""UPDATE initiatives SET name=?,date=?,location=?,manager=?,committee=?,
                           hours=?,status=?,description=?,goals=?,latitude=?,longitude=?,map_status=?,map_note=?,map_date=?,map_visible=? WHERE id=?""", data + (iid,))
            log_action("Updated", f"مبادرة: {name}")
        else:
            db.execute("""INSERT INTO initiatives(id,name,date,location,manager,committee,hours,status,description,goals,latitude,longitude,map_status,map_note,map_date,map_visible)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (uid("ini"),) + data)
            log_action("Created Initiative", f"مبادرة: {name}")
        db.commit()
        parent_id = iid or db.execute("SELECT id FROM initiatives WHERE name=? ORDER BY created_at DESC LIMIT 1", (name,)).fetchone()["id"]
        save_multiple_attachments(request.files.getlist("media_files"), "initiative", parent_id, db, name)
        db.commit()
        flash("تم الحفظ بنجاح", "ok")
        return redirect(url_for("initiatives_list"))
    return render_template("initiative_form.html", i=i, committees=committees, members=members, attachments=attachments_for(db,"initiative",iid) if iid else [])


@app.route("/initiatives/<iid>/delete", methods=["POST"])
def initiative_delete(iid):
    db = get_db()
    db.execute("DELETE FROM initiatives WHERE id=?", (iid,))
    db.execute("DELETE FROM initiative_participants WHERE initiative_id=?", (iid,))
    db.commit()
    flash("تم الحذف", "ok")
    return redirect(url_for("initiatives_list"))


@app.route("/initiatives/<iid>")
def initiative_view(iid):
    db = get_db()
    i = db.execute("SELECT * FROM initiatives WHERE id=?", (iid,)).fetchone()
    if not i:
        return redirect(url_for("initiatives_list"))
    participants = db.execute("""
        SELECT m.*, ip.start_time, ip.end_time, ip.hours participant_hours
        FROM members m
        JOIN initiative_participants ip ON ip.member_id=m.id
        WHERE ip.initiative_id=? ORDER BY m.name
    """, (iid,)).fetchall()
    return render_template("initiative_view.html", i=i, participants=participants, public_view=not is_admin(), map_token=True)


@app.route("/initiatives/<iid>/participants", methods=["GET", "POST"])
def initiative_participants(iid):
    db = get_db()
    i = db.execute("SELECT * FROM initiatives WHERE id=?", (iid,)).fetchone()
    if not i:
        return redirect(url_for("initiatives_list"))
    if request.method == "POST":
        selected = set(request.form.getlist("participant"))
        existing = db.execute("SELECT member_id FROM initiative_participants WHERE initiative_id=?", (iid,)).fetchall()
        existing_ids = {r["member_id"] for r in existing}
        # Existing participants are never removed by the registration screen.
        # This prevents disabled checkboxes from accidentally deleting active records.
        now_time = iraq_now_time()
        for mid in selected:
            if mid not in existing_ids:
                db.execute(
                    "INSERT INTO initiative_participants(initiative_id, member_id, start_time, end_time, hours) VALUES(?,?,?,?,?)",
                    (iid, mid, now_time, None, 0.0))
        db.commit()
        flash("تم تسجيل المشاركين، ووقت الدخول سُجل تلقائيًا الآن", "ok")
        return redirect(url_for("initiative_participants", iid=iid))
    members = db.execute("SELECT * FROM members ORDER BY name").fetchall()
    selected_rows = db.execute(
        "SELECT member_id,start_time,end_time,hours FROM initiative_participants WHERE initiative_id=?", (iid,)).fetchall()
    selected = {r["member_id"]: r for r in selected_rows}
    return render_template("initiative_participants.html", i=i, members=members, selected=selected)


@app.post("/initiatives/<iid>/participants/<mid>/checkout")
def initiative_participant_checkout(iid, mid):
    db = get_db()
    row = db.execute("SELECT * FROM initiative_participants WHERE initiative_id=? AND member_id=?", (iid, mid)).fetchone()
    if not row:
        flash("المشارك غير مسجل", "error")
        return redirect(url_for("initiative_participants", iid=iid))
    if row["end_time"]:
        flash("تم تسجيل خروج هذا العضو مسبقًا ولا يمكن تسجيله مرة ثانية", "error")
        return redirect(url_for("initiative_participants", iid=iid))
    end_time = iraq_now_time()
    hours = duration_hours(row["start_time"], end_time)
    db.execute("UPDATE initiative_participants SET end_time=?, hours=? WHERE initiative_id=? AND member_id=?",
               (end_time, hours, iid, mid))
    db.commit()
    flash(f"تم تسجيل وقت خروج العضو وحساب {hours} ساعة", "ok")
    return redirect(url_for("initiative_participants", iid=iid))


# ============================================================ TASKS ============================================================
@app.route("/tasks")
def tasks_list():
    db = get_db()
    tasks = db.execute("SELECT * FROM tasks ORDER BY created_at DESC").fetchall()
    return render_template("tasks_list.html", tasks=tasks)


@app.route("/tasks/new", methods=["GET", "POST"])
def task_new():
    return task_form()


@app.route("/tasks/<tid>/edit", methods=["GET", "POST"])
def task_edit(tid):
    return task_form(tid)


def task_form(tid=None):
    db = get_db()
    members = db.execute("SELECT * FROM members ORDER BY name").fetchall()
    t = db.execute("SELECT * FROM tasks WHERE id=?", (tid,)).fetchone() if tid else None
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        if not title:
            flash("أدخل عنوان المهمة", "error")
            return redirect(request.url)
        assignee = request.form.get("assignee", "")
        status = request.form.get("status", "Todo")
        data = (title, assignee, request.form.get("deadline", ""),
                request.form.get("priority", "متوسطة"), status, request.form.get("description", ""), request.form.get("latitude", ""), request.form.get("longitude", ""), 1 if request.form.get("map_visible") else 0)
        was_completed = bool(tid) and t and t["status"] == "Completed"
        if tid:
            db.execute("""UPDATE tasks SET title=?,assignee=?,deadline=?,priority=?,status=?,description=?,latitude=?,longitude=?,map_visible=?
                           WHERE id=?""", data + (tid,))
        else:
            new_tid=uid("task")
            db.execute("""INSERT INTO tasks(id,title,assignee,deadline,priority,status,description,latitude,longitude,map_visible,created_at,qr_path)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""", (new_tid,) + data + (now_iso(), _action_qr("task",new_tid,"action")))
        db.commit()
        if status == "Completed" and not was_completed and assignee:
            member = db.execute("SELECT * FROM members WHERE name=?", (assignee,)).fetchone()
            if member:
                pts_cfg = get_points_config()
                add_points(member["id"], pts_cfg.get("task", 10), f"إنجاز مهمة: {title}")
        log_action("Updated" if tid else "Completed Task", f"مهمة: {title}")
        flash("تم الحفظ بنجاح", "ok")
        return redirect(url_for("tasks_list"))
    return render_template("task_form.html", t=t, members=members)


@app.route("/tasks/<tid>/delete", methods=["POST"])
def task_delete(tid):
    db = get_db()
    db.execute("DELETE FROM tasks WHERE id=?", (tid,))
    db.commit()
    flash("تم الحذف", "ok")
    return redirect(url_for("tasks_list"))


# ============================================================ ATTENDANCE ============================================================
@app.route("/attendance")
def attendance_list():
    db = get_db()
    rows = db.execute("SELECT * FROM attendance ORDER BY date DESC").fetchall()
    data = []
    for r in rows:
        m = db.execute("SELECT name FROM members WHERE id=?", (r["member_id"],)).fetchone()
        i = db.execute("SELECT name FROM initiatives WHERE id=?", (r["initiative_id"],)).fetchone() if r["initiative_id"] else None
        data.append({"a": r, "member_name": m["name"] if m else "—", "ini_name": i["name"] if i else "—"})
    return render_template("attendance_list.html", data=data)


@app.route("/attendance/new", methods=["GET", "POST"])
def attendance_new():
    db = get_db()
    members = db.execute("SELECT * FROM members ORDER BY name").fetchall()
    if not members:
        flash("أضف عضواً أولاً", "error")
        return redirect(url_for("attendance_list"))
    initiatives = db.execute("SELECT * FROM initiatives ORDER BY date DESC").fetchall()
    if request.method == "POST":
        member_id = request.form.get("member_id")
        status = request.form.get("status", "Present")
        att_date = iraq_today()
        initiative_id = request.form.get("initiative_id") or None
        start_time = iraq_now_time() if status in ("Present", "Late") else None
        end_time = None
        hours = 0.0
        db.execute("INSERT INTO attendance(id,member_id,date,status,initiative_id,start_time,end_time,hours) VALUES(?,?,?,?,?,?,?,?)",
                   (uid("att"), member_id, att_date, status, initiative_id, start_time, end_time, hours))
        db.commit()
        if status == "Present":
            pts_cfg = get_points_config()
            add_points(member_id, pts_cfg.get("attendance", 10), f"حضور بتاريخ {att_date}")
        log_action("Created", "حضور مسجل")
        flash("تم تسجيل الدخول تلقائيًا بالوقت الحالي", "ok")
        return redirect(url_for("attendance_list"))
    return render_template("attendance_form.html", members=members, initiatives=initiatives, today=iraq_today())


@app.post("/attendance/<aid>/checkout")
def attendance_checkout(aid):
    db = get_db()
    row = db.execute("SELECT * FROM attendance WHERE id=?", (aid,)).fetchone()
    if not row:
        flash("سجل الحضور غير موجود", "error")
        return redirect(url_for("attendance_list"))
    if not row["start_time"]:
        flash("لا يوجد وقت دخول لهذا السجل", "error")
        return redirect(url_for("attendance_list"))
    if row["end_time"]:
        flash("تم تسجيل خروج هذا السجل مسبقًا ولا يمكن تسجيله مرة ثانية", "error")
        return redirect(url_for("attendance_list"))
    end_time = iraq_now_time()
    hours = duration_hours(row["start_time"], end_time)
    db.execute("UPDATE attendance SET end_time=?, hours=? WHERE id=?", (end_time, hours, aid))
    db.commit()
    flash(f"تم تسجيل وقت الخروج وحساب {hours} ساعة", "ok")
    return redirect(url_for("attendance_list"))


# ============================================================ EVALUATIONS ============================================================
@app.route("/evaluations")
def evaluations_list():
    db = get_db()
    rows = db.execute("SELECT * FROM evaluations ORDER BY date DESC").fetchall()
    data = []
    for e in rows:
        m = db.execute("SELECT name FROM members WHERE id=?", (e["evaluated_user_id"],)).fetchone()
        vals = [e[f"c_{k}"] or 0 for k in CRITERIA_KEYS]
        avg_s = round(sum(vals) / len(vals))
        data.append({"e": e, "member_name": m["name"] if m else "—", "avg": avg_s})
    return render_template("evaluations_list.html", data=data)


@app.route("/evaluations/new", methods=["GET", "POST"])
def evaluation_new():
    db = get_db()
    members = db.execute("SELECT * FROM members ORDER BY name").fetchall()
    if not members:
        flash("أضف عضواً أولاً", "error")
        return redirect(url_for("evaluations_list"))
    preselect = request.args.get("member_id", "")
    if request.method == "POST":
        member_id = request.form.get("member_id")
        member = db.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
        criteria = {k: int(request.form.get(f"c_{k}", 50)) for k in CRITERIA_KEYS}
        u = current_user()
        db.execute("""INSERT INTO evaluations(id,evaluated_user_id,evaluator_id,evaluator_name,date,type,notes,
                       c_attendance,c_taskCompletion,c_initiativeParticipation,c_commitment,c_teamwork,c_creativity)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                   (uid("ev"), member_id, u["id"] if u else "", u["name"] if u else "",
                    request.form.get("date", date.today().isoformat()), request.form.get("type", "تقييم مبادرة"),
                    request.form.get("notes", ""), criteria["attendance"], criteria["taskCompletion"],
                    criteria["initiativeParticipation"], criteria["commitment"], criteria["teamwork"],
                    criteria["creativity"]))
        db.commit()
        avg_s = sum(criteria.values()) / 6
        if avg_s >= 85:
            pts_cfg = get_points_config()
            add_points(member_id, pts_cfg.get("excellent", 20), "تقييم ممتاز")
        log_action("Evaluated", f"تقييم {member['name'] if member else ''} ({request.form.get('type','')})")
        flash("تم حفظ التقييم وتحديث النقاط", "ok")
        return redirect(url_for("evaluations_list"))
    return render_template("evaluation_form.html", members=members, preselect=preselect,
                           criteria_keys=CRITERIA_KEYS, criteria_labels=CRITERIA_LABELS,
                           today=date.today().isoformat())


# ============================================================ ACHIEVEMENTS ============================================================
@app.route("/achievements")
def achievements():
    db = get_db()
    members = db.execute("SELECT * FROM members ORDER BY name").fetchall()
    data = []
    for m in members:
        data.append({
            "m": m, "level": member_level(db, m["id"]),
            "points": member_points_total(db, m["id"]), "score": member_score(db, m["id"])
        })
    return render_template("achievements.html", data=data)


# ============================================================ AUDIT LOG ============================================================
@app.route("/audit-log")
def audit_log():
    db = get_db()
    logs = db.execute("SELECT * FROM audit_logs ORDER BY date DESC").fetchall()
    return render_template("audit_log.html", logs=logs)


# ============================================================ MEMBER / ADMIN CARDS ============================================================
def _card_qr(kind, record_id, number):
    # The printed QR opens the scan gateway. Admins get one-tap attendance/action
    # registration; public visitors are redirected to the public card.
    target=url_for("scan_record",kind=kind,record_id=record_id,_external=True)
    img=qrcode.make(target)
    filename=f"card-{kind}-{secure_filename(str(number)) or uuid.uuid4().hex[:8]}.png"
    img.save(os.path.join(UPLOAD_DIR,filename))
    return f"uploads/{filename}"

def _action_qr(kind, record_id, label="scan"):
    target=url_for("scan_record",kind=kind,record_id=record_id,_external=True)
    img=qrcode.make(target)
    filename=f"qr-{kind}-{secure_filename(str(record_id))}-{label}.png"
    img.save(os.path.join(UPLOAD_DIR,filename))
    return f"uploads/{filename}"

def _card_template(db, kind):
    row=db.execute("SELECT * FROM card_templates WHERE kind=? ORDER BY updated_at DESC LIMIT 1",(kind,)).fetchone()
    if row: return row
    db.execute("INSERT INTO card_templates(id,name,kind,created_at,updated_at) VALUES(?,?,?,?,?)",(uid("ct"),f"قالب {kind}",kind,now_iso(),now_iso())); db.commit()
    return db.execute("SELECT * FROM card_templates WHERE kind=? ORDER BY updated_at DESC LIMIT 1",(kind,)).fetchone()

@app.route("/admin/cards")
def cards_admin():
    if not has_permission("cards"): return render_template("403.html",permission="cards"),403
    db=get_db()
    members=db.execute("SELECT * FROM members ORDER BY name").fetchall(); admins=db.execute("SELECT * FROM administrators ORDER BY name").fetchall()
    templates=db.execute("SELECT * FROM card_templates ORDER BY kind,updated_at DESC").fetchall()
    return render_template("cards_admin.html",members=members,admins=admins,templates=templates)

@app.route("/admin/cards/templates", methods=["POST"])
def card_template_save():
    if not has_permission("cards"): return render_template("403.html",permission="cards"),403
    db=get_db(); kind=request.form.get("kind","member")
    tid=request.form.get("template_id") or uid("ct")
    existing=db.execute("SELECT logo FROM card_templates WHERE id=?",(tid,)).fetchone()
    logo=save_image_as_jpg(request.files.get("logo_file"),"card-logo") or (existing["logo"] if existing else "")
    vals=(request.form.get("name") or f"قالب {kind}",kind,float(request.form.get("width_mm") or 85.6),float(request.form.get("height_mm") or 54),request.form.get("bg") or "#071522",request.form.get("accent") or "#20B486",request.form.get("text_color") or "#FFFFFF",request.form.get("font") or "Tajawal",logo,1 if request.form.get("public_default") else 0,json.dumps({"show_department":bool(request.form.get("show_department")),"show_stage":bool(request.form.get("show_stage")),"show_committee":bool(request.form.get("show_committee")),"show_position":bool(request.form.get("show_position")),"show_dates":bool(request.form.get("show_dates")),"show_nfc":bool(request.form.get("show_nfc"))},ensure_ascii=False),now_iso(),now_iso())
    db.execute("INSERT INTO card_templates(id,name,kind,width_mm,height_mm,bg,accent,text_color,font,logo,public_default,fields_json,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(id) DO UPDATE SET name=excluded.name,kind=excluded.kind,width_mm=excluded.width_mm,height_mm=excluded.height_mm,bg=excluded.bg,accent=excluded.accent,text_color=excluded.text_color,font=excluded.font,logo=excluded.logo,public_default=excluded.public_default,fields_json=excluded.fields_json,updated_at=excluded.updated_at",(tid,)+vals)
    db.commit(); log_action("Saved card template",f"{kind}:{vals[0]}"); flash("تم حفظ قالب البطاقة","ok"); return redirect(url_for("cards_admin"))

@app.route("/admin/cards/templates/<tid>/preview")
def card_template_preview(tid):
    if not has_permission("cards"): return render_template("403.html",permission="cards"),403
    db=get_db(); t=db.execute("SELECT * FROM card_templates WHERE id=?",(tid,)).fetchone()
    if not t:return "القالب غير موجود",404
    kind=t["kind"]; table="members" if kind=="member" else "administrators"
    person=db.execute(f"SELECT * FROM {table} ORDER BY name LIMIT 1").fetchone()
    if not person:return "أضف سجلًا أولاً لمعاينة البطاقة",404
    number=person["membership_no"] or f"AHVT-{kind.upper()}-PREVIEW"; qr=_card_qr(kind,person["id"],number); nfc_url=url_for("card_public",kind=kind,record_id=person["id"],_external=True)
    return render_template("membership_card.html",person=person,kind=kind,qr_path=qr,nfc_url=nfc_url,team_name=get_setting("teamName","فريق الحكمة التطوعي"),template=t,printable=True)

@app.route("/admin/cards/templates/<tid>/delete", methods=["POST"])
def card_template_delete(tid):
    if not has_permission("cards"): return render_template("403.html",permission="cards"),403
    db=get_db(); db.execute("DELETE FROM card_templates WHERE id=?",(tid,)); db.commit(); flash("تم حذف القالب","ok"); return redirect(url_for("cards_admin"))

@app.route("/admin/cards/<kind>/<record_id>")
def card_print(kind, record_id):
    if not has_permission("cards"): return render_template("403.html",permission="cards"),403
    db=get_db(); table="members" if kind=="member" else "administrators" if kind=="admin" else ""
    if not table:return "نوع البطاقة غير صحيح",404
    row=db.execute(f"SELECT * FROM {table} WHERE id=?",(record_id,)).fetchone()
    if not row:return "السجل غير موجود",404
    number=row["membership_no"] or (f"HIKMA-M-{record_id[-8:].upper()}" if kind=="member" else f"HIKMA-A-{record_id[-8:].upper()}")
    qr=_card_qr(kind,record_id,number); nfc_url=url_for("card_public",kind=kind,record_id=record_id,_external=True); tpl=_card_template(db,kind)
    return render_template("membership_card.html", person=row, kind=kind, qr_path=qr, nfc_url=nfc_url, team_name=get_setting("teamName","فريق الحكمة التطوعي"), template=tpl, printable=True)

@app.route("/card/<kind>/<record_id>")
def card_public(kind, record_id):
    db=get_db(); table="members" if kind=="member" else "administrators" if kind=="admin" else ""
    if not table:return "نوع البطاقة غير صحيح",404
    row=db.execute(f"SELECT * FROM {table} WHERE id=?",(record_id,)).fetchone()
    if not row:return "السجل غير موجود",404
    if not row["card_public"]: return render_template("403.html",permission="البطاقة خاصة"),403
    number=row["membership_no"] or (f"HIKMA-M-{record_id[-8:].upper()}" if kind=="member" else f"HIKMA-A-{record_id[-8:].upper()}")
    qr=_card_qr(kind,record_id,number); nfc_url=url_for("card_public",kind=kind,record_id=record_id,_external=True); tpl=_card_template(db,kind)
    return render_template("membership_card.html", person=row, kind=kind, qr_path=qr, nfc_url=nfc_url, team_name=get_setting("teamName","فريق الحكمة التطوعي"), template=tpl, printable=False)

@app.route("/admin/qr-scanner")
def qr_scanner():
    if not is_admin():
        return render_template("403.html", permission="scan"), 403
    mode=request.args.get("mode","attendance")
    record_id=request.args.get("record_id","")
    equipment_id=request.args.get("equipment_id","")
    return render_template("qr_scanner.html", mode=mode, record_id=record_id, equipment_id=equipment_id)

@app.post("/api/admin/qr/scan")
def admin_qr_scan_api():
    if not is_admin(): return jsonify({"ok":False,"error":"غير مصرح"}),403
    db=get_db(); data=request.get_json(silent=True) or {}
    raw=(data.get("raw") or "").strip(); mode=(data.get("mode") or "attendance").strip()
    record_id=(data.get("record_id") or "").strip(); equipment_id=(data.get("equipment_id") or "").strip()
    # Accept both absolute URLs and the relative /scan/... path printed in AHVT cards.
    from urllib.parse import urlparse
    path=urlparse(raw).path if raw else ""
    parts=[x for x in path.split("/") if x]
    if len(parts)>=3 and parts[0]=="scan":
        kind, rid=parts[1], parts[2]
    else:
        kind, rid="",raw.rsplit("/",1)[-1] if raw else ""
    if kind not in ("member","admin","equipment","task","event"):
        return jsonify({"ok":False,"error":"هذا QR ليس من بطاقات AHVT المدعومة"}),400
    if kind=="admin":
        return jsonify({"ok":False,"error":"بطاقة الإداري يمكن عرضها، لكن تسجيل الحضور الحالي مخصص لسجل الأعضاء"}),400
    if mode in ("attendance","event") and kind!="member":
        return jsonify({"ok":False,"error":"امسح QR بطاقة العضو"}),400
    if mode=="task" and kind!="member":
        return jsonify({"ok":False,"error":"امسح QR بطاقة العضو أولاً"}),400
    if mode=="equipment" and kind=="member":
        eid=equipment_id or record_id
        if not eid:
            return jsonify({"ok":False,"error":"اختر المعدة قبل المسح"}),400
        m=db.execute("SELECT * FROM members WHERE id=?",(rid,)).fetchone(); e=db.execute("SELECT * FROM podcast_equipment WHERE id=?",(eid,)).fetchone()
        if not m or not e:return jsonify({"ok":False,"error":"العضو أو المعدة غير موجود"}),404
        open_loan=db.execute("SELECT id FROM equipment_loans WHERE equipment_id=? AND member_id=? AND status='issued' ORDER BY issued_at DESC LIMIT 1",(eid,rid)).fetchone()
        if open_loan:
            db.execute("UPDATE equipment_loans SET status='returned',returned_at=? WHERE id=?",(now_iso(),open_loan["id"])); action="إرجاع"
        else:
            db.execute("INSERT INTO equipment_loans(id,equipment_id,member_id,issued_at,status,notes) VALUES(?,?,?,?,?,?)",(uid("loan"),eid,rid,now_iso(),"issued","تم التسجيل عبر كاميرا QR")); action="استلام"
        db.commit(); log_action(f"Equipment {action} via camera QR",f"{m['name']} · {e['name']}")
        return jsonify({"ok":True,"message":f"تم {action} {e['name']} — {m['name']}","name":m["name"],"time":now_iso()})
    if mode in ("attendance","event"):
        m=db.execute("SELECT * FROM members WHERE id=?",(rid,)).fetchone()
        if not m:return jsonify({"ok":False,"error":"العضو غير موجود"}),404
        today=iraq_today(); open_att=db.execute("SELECT * FROM attendance WHERE member_id=? AND date=? AND end_time IS NULL ORDER BY start_time DESC LIMIT 1",(rid,today)).fetchone()
        if open_att:
            start=open_att["start_time"] or iraq_now_time(); end=iraq_now_time()
            try:
                h=max(0,(datetime.strptime(end,"%H:%M")-datetime.strptime(start,"%H:%M")).seconds/3600)
            except Exception:h=0
            db.execute("UPDATE attendance SET end_time=?,hours=? WHERE id=?",(end,round(h,2),open_att["id"])); action="تسجيل الخروج"
        else:
            db.execute("INSERT INTO attendance(id,member_id,date,status,initiative_id,start_time,end_time,hours) VALUES(?,?,?,?,?,?,?,?)",(uid("att"),rid,today,"Present",None,iraq_now_time(),None,0.0)); action="تسجيل الدخول"
        db.commit(); log_action("Camera QR attendance",f"{m['name']} · {action}")
        return jsonify({"ok":True,"message":f"{action}: {m['name']}","name":m["name"],"time":iraq_now_time()})
    if mode=="task":
        tid=record_id; m=db.execute("SELECT * FROM members WHERE id=?",(rid,)).fetchone(); t=db.execute("SELECT * FROM tasks WHERE id=?",(tid,)).fetchone()
        if not m or not t:return jsonify({"ok":False,"error":"العضو أو المهمة غير موجود"}),404
        db.execute("INSERT OR IGNORE INTO task_completions(id,task_id,member_id,completed_at,notes) VALUES(?,?,?,?,?)",(uid("tc"),tid,rid,now_iso(),"تم عبر كاميرا QR")); db.commit(); log_action("Task camera QR",f"{m['name']} · {t['title']}")
        return jsonify({"ok":True,"message":f"تم تسجيل المهمة: {t['title']} — {m['name']}","name":m["name"],"time":iraq_now_time()})
    return jsonify({"ok":False,"error":"السياق غير مدعوم"}),400

@app.route("/scan/<kind>/<record_id>")
def scan_record(kind, record_id):
    db=get_db()
    if kind in ("member","admin"):
        table = "members" if kind=="member" else "administrators"
        row=db.execute(f"SELECT * FROM {table} WHERE id=?",(record_id,)).fetchone()
        if not row: return "السجل غير موجود",404
        if not row["card_public"] and not is_admin():
            return render_template("403.html",permission="البطاقة خاصة"),403
        return render_template("scan_action.html", kind="member", record=row, title="بطاقة QR",
                               action_url=url_for("scan_member_action",record_id=record_id),
                               equipment=db.execute("SELECT id,name,category FROM podcast_equipment ORDER BY name").fetchall(),
                               tasks=db.execute("SELECT id,title FROM tasks WHERE status NOT IN ('Done','Completed','مكتملة') ORDER BY deadline").fetchall(),
                               members=db.execute("SELECT id,name FROM members ORDER BY name").fetchall())
    if kind=="event":
        row=db.execute("SELECT * FROM events WHERE id=?",(record_id,)).fetchone()
        if not row:return "الفعالية غير موجودة",404
        return render_template("scan_action.html",kind=kind,record=row,title="تسجيل الفعالية",action_url=url_for("scan_task_action",record_id=record_id,kind="event"),equipment=[],tasks=[],members=db.execute("SELECT id,name FROM members ORDER BY name").fetchall())
    if kind=="task":
        row=db.execute("SELECT * FROM tasks WHERE id=?",(record_id,)).fetchone()
        if not row:return "المهمة غير موجودة",404
        return render_template("scan_action.html",kind=kind,record=row,title="تسجيل إنجاز المهمة",action_url=url_for("scan_task_action",record_id=record_id,kind="task"),equipment=[],tasks=[],members=db.execute("SELECT id,name FROM members ORDER BY name").fetchall())
    if kind=="equipment":
        row=db.execute("SELECT * FROM podcast_equipment WHERE id=?",(record_id,)).fetchone()
        if not row:return "المعدة غير موجودة",404
        return render_template("scan_action.html",kind=kind,record=row,title="استلام/تسليم معدة",action_url=url_for("scan_equipment_action",record_id=record_id),equipment=[],tasks=[],members=db.execute("SELECT id,name FROM members ORDER BY name").fetchall())
    return "نوع QR غير مدعوم",404

@app.post("/scan/member/<record_id>/action")
def scan_member_action(record_id):
    if not is_admin(): return render_template("403.html",permission="scan"),403
    db=get_db(); m=db.execute("SELECT * FROM members WHERE id=?",(record_id,)).fetchone()
    if not m:return "العضو غير موجود",404
    action=request.form.get("action","attendance")
    if action=="attendance":
        today=iraq_today(); existing=db.execute("SELECT id FROM attendance WHERE member_id=? AND date=? AND end_time IS NULL ORDER BY start_time DESC LIMIT 1",(record_id,today)).fetchone()
        if not existing:
            aid=uid("att"); db.execute("INSERT INTO attendance(id,member_id,date,status,initiative_id,start_time,end_time,hours) VALUES(?,?,?,?,?,?,?,?)",(aid,record_id,today,"Present",request.form.get("initiative_id") or None,iraq_now_time(),None,0.0)); db.commit(); add_points(record_id,get_points_config().get("attendance",10),f"حضور QR بتاريخ {today}"); log_action("QR attendance check-in",m["name"]); flash(f"تم تسجيل حضور {m['name']} بالـQR","ok")
        else: flash("العضو لديه حضور مفتوح اليوم بالفعل","error")
    elif action=="task":
        tid=request.form.get("task_id"); t=db.execute("SELECT * FROM tasks WHERE id=?",(tid,)).fetchone()
        if not t:return "المهمة غير موجودة",404
        db.execute("INSERT OR IGNORE INTO task_completions(id,task_id,member_id,completed_at,notes) VALUES(?,?,?,?,?)",(uid("tc"),tid,record_id,now_iso(),request.form.get("notes",""))); db.commit(); log_action("Task QR completion",f"{m['name']} · {t['title']}"); flash("تم تسجيل إنجاز المهمة عبر QR","ok")
    elif action=="equipment":
        eid=request.form.get("equipment_id"); e=db.execute("SELECT * FROM podcast_equipment WHERE id=?",(eid,)).fetchone()
        if not e:return "المعدة غير موجودة",404
        db.execute("INSERT INTO equipment_loans(id,equipment_id,member_id,issued_at,status,notes) VALUES(?,?,?,?,?,?)",(uid("loan"),eid,record_id,now_iso(),"issued","تم الاستلام عبر QR")); db.commit(); log_action("Equipment checkout via QR",f"{m['name']} · {e['name']}"); flash(f"تم تسجيل استلام {e['name']} للعضو {m['name']}","ok")
    return redirect(url_for("scan_record",kind="member",record_id=record_id))

@app.post("/scan/<kind>/<record_id>/action")
def scan_task_action(kind,record_id):
    if not is_admin(): return render_template("403.html",permission="scan"),403
    db=get_db(); member_id=request.form.get("member_id")
    if not member_id:return "اختر العضو أولاً",400
    m=db.execute("SELECT id,name FROM members WHERE id=?",(member_id,)).fetchone()
    if not m:return "العضو غير موجود",404
    if kind=="task":
        t=db.execute("SELECT * FROM tasks WHERE id=?",(record_id,)).fetchone()
        if not t:return "المهمة غير موجودة",404
        db.execute("INSERT OR IGNORE INTO task_completions(id,task_id,member_id,completed_at,notes) VALUES(?,?,?,?,?)",(uid("tc"),record_id,member_id,now_iso(),request.form.get("notes",""))); db.commit(); log_action("Task QR completion",f"{m['name']} · {t['title']}"); flash("تم تسجيل إنجاز المهمة عبر QR","ok")
    elif kind=="event":
        e=db.execute("SELECT * FROM events WHERE id=?",(record_id,)).fetchone()
        if not e:return "الفعالية غير موجودة",404
        today=iraq_today(); db.execute("INSERT INTO attendance(id,member_id,date,status,initiative_id,start_time,end_time,hours) VALUES(?,?,?,?,?,?,?,?)",(uid("att"),member_id,today,"Present",None,iraq_now_time(),None,0.0)); db.commit(); log_action("Event QR attendance",f"{m['name']} · {e['title']}"); flash("تم تسجيل حضور الفعالية عبر QR","ok")
    return redirect(url_for("scan_record",kind=kind,record_id=record_id))

@app.post("/scan/equipment/<record_id>/action")
def scan_equipment_action(record_id):
    if not is_admin(): return render_template("403.html",permission="scan"),403
    db=get_db(); e=db.execute("SELECT * FROM podcast_equipment WHERE id=?",(record_id,)).fetchone(); mid=request.form.get("member_id")
    if not e:return "المعدة غير موجودة",404
    m=db.execute("SELECT id,name FROM members WHERE id=?",(mid,)).fetchone() if mid else None
    if not m:return "اختر العضو أولاً",400
    open_loan=db.execute("SELECT id FROM equipment_loans WHERE equipment_id=? AND member_id=? AND status='issued' ORDER BY issued_at DESC LIMIT 1",(record_id,mid)).fetchone()
    if open_loan:
        db.execute("UPDATE equipment_loans SET status='returned',returned_at=? WHERE id=?",(now_iso(),open_loan["id"])); flash("تم تسجيل إرجاع المعدة عبر QR","ok"); log_action("Equipment return via QR",f"{m['name']} · {e['name']}")
    else:
        db.execute("INSERT INTO equipment_loans(id,equipment_id,member_id,issued_at,status,notes) VALUES(?,?,?,?,?,?)",(uid("loan"),record_id,mid,now_iso(),"issued","تم الاستلام عبر QR المعدة")); flash("تم تسجيل استلام المعدة عبر QR","ok"); log_action("Equipment checkout via QR",f"{m['name']} · {e['name']}")
    db.commit(); return redirect(url_for("scan_record",kind="equipment",record_id=record_id))

# ============================================================ HONOR LIST ============================================================
@app.route("/honor")
def honor_public():
    db=get_db()
    rows=db.execute("SELECT * FROM honor_list WHERE public=1 ORDER BY honor_date DESC, created_at DESC").fetchall()
    galleries={r["id"]: attachments_for(db,"honor",r["id"]) for r in rows}
    return render_template("honor_public.html", honors=rows, galleries=galleries)

@app.route("/admin/honor", methods=["GET","POST"])
def honor_admin():
    if not has_permission("honor"): return render_template("403.html", permission="honor"),403
    db=get_db()
    if request.method=="POST":
        name=request.form.get("name","").strip()
        if not name:
            flash("اكتب الاسم أولاً","error"); return redirect(url_for("honor_admin"))
        photo=save_image_as_jpg(request.files.get("photo_file"),"honor")
        hid=uid("honor")
        db.execute("INSERT INTO honor_list(id,name,photo,honor_type,reason,achievement,honor_date,occasion,department,certificate_id,badge,description,public,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                   (hid,name,photo,request.form.get("honor_type","تكريم"),request.form.get("reason","").strip(),request.form.get("achievement","").strip(),request.form.get("honor_date") or date.today().isoformat(),request.form.get("occasion","").strip(),request.form.get("department","").strip(),request.form.get("certificate_id") or None,request.form.get("badge","").strip(),request.form.get("description","").strip(),1 if request.form.get("public") else 0,now_iso(),now_iso()))
        save_multiple_attachments(request.files.getlist("gallery_files"), "honor", hid, db, "honor-gallery")
        db.commit(); log_action("Created honor",name); flash("تمت إضافة التكريم إلى قائمة الشرف","ok"); return redirect(url_for("honor_admin"))
    rows=db.execute("SELECT * FROM honor_list ORDER BY honor_date DESC,created_at DESC").fetchall()
    certs=db.execute("SELECT id,certificate_no,recipient_name FROM certificates ORDER BY created_at DESC LIMIT 200").fetchall()
    return render_template("honor_admin.html", honors=rows, certificates=certs, edit_honor=None, today=date.today().isoformat(), edit_gallery=[])

@app.route("/admin/honor/<hid>/edit", methods=["GET","POST"])
def honor_edit(hid):
    if not has_permission("honor"): return render_template("403.html", permission="honor"),403
    db=get_db(); h=db.execute("SELECT * FROM honor_list WHERE id=?",(hid,)).fetchone()
    if not h: flash("السجل غير موجود","error"); return redirect(url_for("honor_admin"))
    if request.method=="POST":
        photo=h["photo"] or ""
        new_photo=save_image_as_jpg(request.files.get("photo_file"),"honor")
        if new_photo: photo=new_photo
        db.execute("UPDATE honor_list SET name=?,photo=?,honor_type=?,reason=?,achievement=?,honor_date=?,occasion=?,department=?,certificate_id=?,badge=?,description=?,public=?,updated_at=? WHERE id=?",
                   (request.form.get("name","").strip(),photo,request.form.get("honor_type","تكريم"),request.form.get("reason","").strip(),request.form.get("achievement","").strip(),request.form.get("honor_date") or date.today().isoformat(),request.form.get("occasion","").strip(),request.form.get("department","").strip(),request.form.get("certificate_id") or None,request.form.get("badge","").strip(),request.form.get("description","").strip(),1 if request.form.get("public") else 0,now_iso(),hid))
        save_multiple_attachments(request.files.getlist("gallery_files"), "honor", hid, db, "honor-gallery")
        db.commit(); log_action("Edited honor",h["name"]); flash("تم تعديل التكريم","ok"); return redirect(url_for("honor_admin"))
    certs=db.execute("SELECT id,certificate_no,recipient_name FROM certificates ORDER BY created_at DESC LIMIT 200").fetchall()
    rows=db.execute("SELECT * FROM honor_list ORDER BY honor_date DESC,created_at DESC").fetchall()
    return render_template("honor_admin.html", honors=rows, certificates=certs, edit_honor=h, today=date.today().isoformat(), edit_gallery=attachments_for(db,"honor",h["id"]))

@app.route("/admin/honor/<hid>/delete", methods=["POST"])
def honor_delete(hid):
    if not has_permission("delete"): return render_template("403.html", permission="delete"),403
    db=get_db(); h=db.execute("SELECT * FROM honor_list WHERE id=?",(hid,)).fetchone()
    if h:
        db.execute("DELETE FROM honor_list WHERE id=?",(hid,)); db.commit(); log_action("Deleted honor",h["name"]); flash("تم حذف السجل من قائمة الشرف","ok")
    return redirect(url_for("honor_admin"))

# ============================================================ CERTIFICATES ============================================================
def _certificate_qr(cid, certificate_no):
    verify_url = url_for("certificate_verify", certificate_no=certificate_no, _external=True)
    img = qrcode.make(verify_url); filename = f"cert-{certificate_no}.png"
    img.save(os.path.join(UPLOAD_DIR, filename)); return f"uploads/{filename}"

def _certificate_logo_upload(f, prefix):
    return save_image_as_jpg(f, prefix) if f and f.filename else ""

@app.route("/admin/certificates", methods=["GET", "POST"])
def certificates_admin():
    if not has_permission("certificates"): return render_template("403.html", permission="certificates"), 403
    db=get_db()
    if request.method=="POST":
        recipient=request.form.get("recipient_name","").strip()
        if not recipient: flash("اكتب اسم المستفيد","error"); return redirect(url_for("certificates_admin"))
        initiative_id=request.form.get("initiative_id") or None; ini=db.execute("SELECT * FROM initiatives WHERE id=?",(initiative_id,)).fetchone() if initiative_id else None
        issue_date=request.form.get("issue_date") or date.today().isoformat(); hours=float(request.form.get("hours") or 0)
        if hours<=0 and ini and ini["hours"]: hours=float(ini["hours"] or 0)
        no="HIKMA-"+date.today().strftime("%Y")+"-"+uuid.uuid4().hex[:8].upper(); cid=uid("cert"); token=uuid.uuid4().hex; u=current_user()
        template=request.form.get("template","classic") if request.form.get("template") in ("classic","minimal","impact") else "classic"
        issuer=request.form.get("issuer_name","").strip() or get_setting("certificateIssuer",get_setting("teamName","فريق الحكمة التطوعي"))
        writer=request.form.get("writer_name","").strip() or (u["name"] if u else "")
        sql="INSERT INTO certificates(id,certificate_no,recipient_name,certificate_type,initiative_id,initiative_name,issue_date,hours,note,issued_by,created_at,status,verify_token,template,recipient_member_id,issuer_name,custom_title,custom_intro,custom_body,custom_footer,logo1,logo2,logo3,writer_name,paper_size,custom_eyebrow,custom_date_label,custom_hours_label,writer_label,design_bg,design_accent,design_text,design_font) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
        vals=(cid,no,recipient,request.form.get("certificate_type","شهادة مشاركة"),initiative_id,ini["name"] if ini else request.form.get("initiative_name","").strip(),issue_date,hours,request.form.get("note","").strip(),u["name"] if u else issuer,now_iso(),"valid",token,template,request.form.get("recipient_member_id") or None,issuer,request.form.get("custom_title","").strip(),request.form.get("custom_intro","").strip(),request.form.get("custom_body","").strip(),request.form.get("custom_footer","").strip(),_certificate_logo_upload(request.files.get("logo1"),"cert-logo1"),_certificate_logo_upload(request.files.get("logo2"),"cert-logo2"),_certificate_logo_upload(request.files.get("logo3"),"cert-logo3"),writer,"A4 landscape",request.form.get("custom_eyebrow","").strip(),request.form.get("custom_date_label","تاريخ الإصدار").strip(),request.form.get("custom_hours_label","ساعة مشاركة").strip(),request.form.get("writer_label","إعداد وكتابة").strip(),request.form.get("design_bg","#071625").strip(),request.form.get("design_accent","#20B486").strip(),request.form.get("design_text","#FFFFFF").strip(),request.form.get("design_font","Tajawal").strip())
        try:
            db.execute(sql, vals)
            db.commit()
        except sqlite3.IntegrityError:
            db.rollback()
            # Extremely defensive retry: regenerate both identifiers in case a legacy
            # database already contains the generated certificate number/token.
            no="HIKMA-"+date.today().strftime("%Y")+"-"+uuid.uuid4().hex[:10].upper()
            token=uuid.uuid4().hex
            vals=list(vals); vals[1]=no; vals[12]=token; vals[0]=uid("cert") ; cid=vals[0]
            try:
                db.execute(sql, tuple(vals)); db.commit()
            except Exception:
                db.rollback(); app.logger.exception("Certificate insert failed after retry")
                flash("تعذر حفظ الشهادة. تأكد من الحقول ثم حاول مرة أخرى.", "error")
                return redirect(url_for("certificates_admin"))
        except Exception:
            db.rollback(); app.logger.exception("Certificate insert failed")
            flash("تعذر حفظ الشهادة. تم تسجيل الخطأ للمراجعة.", "error")
            return redirect(url_for("certificates_admin"))
        created = db.execute("SELECT * FROM certificates WHERE id=?", (cid,)).fetchone()
        if not created:
            flash("تعذر حفظ الشهادة في سجل الشهادات. حاول مرة أخرى.", "error")
            return redirect(url_for("certificates_admin"))
        qr_path = _certificate_qr(cid, no)
        db.execute("UPDATE certificates SET qr_path=? WHERE id=?", (qr_path, cid))
        db.commit()
        created = db.execute("SELECT * FROM certificates WHERE id=?", (cid,)).fetchone()
        if not created:
            flash("تم حفظ الشهادة لكن تعذر تجهيز QR. افتحها من سجل الشهادات.", "error")
            return redirect(url_for("certificates_admin"))
        log_action("Created certificate", f"{recipient} · {no}")
        return redirect(url_for("certificate_view", cid=cid))
    rows=db.execute("SELECT * FROM certificates ORDER BY created_at DESC LIMIT 100").fetchall()
    return render_template("certificates_admin.html",certificates=rows,initiatives=db.execute("SELECT id,name FROM initiatives ORDER BY date DESC").fetchall(),members=db.execute("SELECT id,name FROM members ORDER BY name").fetchall(),today=date.today().isoformat(),edit_certificate=None)

@app.route("/admin/certificates/<cid>/edit", methods=["GET","POST"])
def certificate_edit(cid):
    if not has_permission("certificates"): return render_template("403.html",permission="certificates"),403
    db=get_db(); c=db.execute("SELECT * FROM certificates WHERE id=?",(cid,)).fetchone()
    if not c:return "الشهادة غير موجودة",404
    if request.method=="POST":
        logos=[c["logo1"],c["logo2"],c["logo3"]]
        for idx in range(3):
            slot=f"logo{idx+1}"
            if request.form.get(f"remove_{slot}"): logos[idx]=""
            up=_certificate_logo_upload(request.files.get(slot),f"cert-{slot}")
            if up: logos[idx]=up
        db.execute("UPDATE certificates SET recipient_name=?,certificate_type=?,initiative_name=?,issue_date=?,hours=?,note=?,template=?,issuer_name=?,custom_title=?,custom_intro=?,custom_body=?,custom_footer=?,logo1=?,logo2=?,logo3=?,writer_name=?,paper_size=?,custom_eyebrow=?,custom_date_label=?,custom_hours_label=?,writer_label=?,design_bg=?,design_accent=?,design_text=?,design_font=? WHERE id=?",(request.form.get("recipient_name","").strip(),request.form.get("certificate_type","شهادة مشاركة"),request.form.get("initiative_name","").strip(),request.form.get("issue_date") or date.today().isoformat(),float(request.form.get("hours") or 0),request.form.get("note","").strip(),request.form.get("template","classic"),request.form.get("issuer_name","").strip() or get_setting("certificateIssuer",get_setting("teamName","فريق الحكمة التطوعي")),request.form.get("custom_title","").strip(),request.form.get("custom_intro","").strip(),request.form.get("custom_body","").strip(),request.form.get("custom_footer","").strip(),logos[0],logos[1],logos[2],request.form.get("writer_name","").strip() or current_user()["name"],"A4 landscape",request.form.get("custom_eyebrow","").strip(),request.form.get("custom_date_label","تاريخ الإصدار").strip(),request.form.get("custom_hours_label","ساعة مشاركة").strip(),request.form.get("writer_label","إعداد وكتابة").strip(),request.form.get("design_bg","#071625").strip(),request.form.get("design_accent","#20B486").strip(),request.form.get("design_text","#FFFFFF").strip(),request.form.get("design_font","Tajawal").strip(),cid))
        db.commit(); log_action("Edited certificate",c["certificate_no"]); flash("تم تعديل الشهادة","ok"); return redirect(url_for("certificate_view",cid=cid))
    return render_template("certificates_admin.html",certificates=[],initiatives=db.execute("SELECT id,name FROM initiatives ORDER BY date DESC").fetchall(),members=db.execute("SELECT id,name FROM members ORDER BY name").fetchall(),today=date.today().isoformat(),edit_certificate=c)

@app.route("/admin/certificates/<cid>/delete", methods=["POST"])
def certificate_delete(cid):
    if not has_permission("certificates"):
        return render_template("403.html", permission="certificates"), 403
    db=get_db(); c=db.execute("SELECT * FROM certificates WHERE id=?", (cid,)).fetchone()
    if not c:
        flash("الشهادة غير موجودة", "error")
        return redirect(url_for("certificates_admin"))
    # Remove the generated QR image when possible.
    qr = c["qr_path"] or ""
    if qr.startswith("uploads/"):
        try:
            os.remove(os.path.join(UPLOAD_DIR, qr.split("uploads/",1)[1]))
        except OSError:
            pass
    db.execute("DELETE FROM certificates WHERE id=?", (cid,)); db.commit()
    log_action("Deleted certificate", c["certificate_no"])
    flash("تم حذف الشهادة", "ok")
    return redirect(url_for("certificates_admin"))

@app.route("/admin/certificates/<cid>/revoke", methods=["POST"])
def certificate_revoke(cid):
    if not has_permission("certificates"): return render_template("403.html",permission="certificates"),403
    db=get_db(); c=db.execute("SELECT * FROM certificates WHERE id=?",(cid,)).fetchone()
    if not c:return "الشهادة غير موجودة",404
    db.execute("UPDATE certificates SET status='revoked' WHERE id=?",(cid,)); db.commit(); log_action("Revoked certificate",c["certificate_no"]); flash("تم إلغاء الشهادة","ok"); return redirect(url_for("certificates_admin"))

@app.route("/admin/certificates/<cid>/reissue", methods=["POST"])
def certificate_reissue(cid):
    if not has_permission("certificates"): return render_template("403.html",permission="certificates"),403
    db=get_db(); c=db.execute("SELECT * FROM certificates WHERE id=?",(cid,)).fetchone()
    if not c:return "الشهادة غير موجودة",404
    no="HIKMA-"+date.today().strftime("%Y")+"-"+uuid.uuid4().hex[:8].upper(); token=uuid.uuid4().hex; new_id=uid("cert")
    sql="INSERT INTO certificates(id,certificate_no,recipient_name,certificate_type,initiative_id,initiative_name,issue_date,hours,note,issued_by,created_at,status,verify_token,template,recipient_member_id,issuer_name,custom_title,custom_intro,custom_body,custom_footer,logo1,logo2,logo3,writer_name,paper_size,custom_eyebrow,custom_date_label,custom_hours_label,writer_label,design_bg,design_accent,design_text,design_font) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
    vals=(new_id,no,c["recipient_name"],c["certificate_type"],c["initiative_id"],c["initiative_name"],date.today().isoformat(),c["hours"],c["note"],current_user()["name"],now_iso(),"valid",token,c["template"],c["recipient_member_id"],c["issuer_name"],c["custom_title"],c["custom_intro"],c["custom_body"],c["custom_footer"],c["logo1"],c["logo2"],c["logo3"],c["writer_name"],"A4 landscape",c["custom_eyebrow"],c["custom_date_label"],c["custom_hours_label"],c["writer_label"],c["design_bg"],c["design_accent"],c["design_text"],c["design_font"])
    db.execute(sql,vals); db.commit(); db.execute("UPDATE certificates SET qr_path=? WHERE id=?",(_certificate_qr(new_id,no),new_id)); db.commit(); log_action("Reissued certificate",f"{c['certificate_no']} -> {no}"); flash("تم إصدار نسخة جديدة برقم جديد","ok"); return redirect(url_for("certificate_view",cid=new_id))

@app.route("/certificate/view/<cid>")
def certificate_view(cid):
    if not is_admin(): return redirect(url_for("admin_login"))
    db=get_db(); c=db.execute("SELECT * FROM certificates WHERE id=? OR certificate_no=?", (cid, cid)).fetchone()
    if not c:
        return "الشهادة غير موجودة في سجل الشهادات", 404
    return render_template("certificate.html", certificate=c, team_logo=get_setting("teamLogo",""), university_logo=get_setting("universityLogo",""), verify_url=url_for("certificate_verify",certificate_no=c["certificate_no"],_external=True))

@app.route("/verify/certificate/<certificate_no>")
def certificate_verify(certificate_no):
    c=get_db().execute("SELECT * FROM certificates WHERE certificate_no=?",(certificate_no,)).fetchone()
    return render_template("certificate_verify.html", certificate=c)

# ============================================================ BRAND / LOGO ============================================================
@app.route("/admin/logo/<kind>", methods=["POST"])
def upload_logo(kind):
    if not is_creator():
        flash("تعديل الشعارات متاح لصانع التطبيق فقط", "error")
        return redirect(url_for("settings_page"))
    if kind not in ("team", "university", "favicon"):
        return redirect(url_for("settings_page"))
    f=request.files.get("logo_file")
    if not f or not f.filename:
        flash("اختر صورة أولاً", "error")
        return redirect(url_for("settings_page"))
    ext=os.path.splitext(secure_filename(f.filename))[1].lower()
    allowed={".png", ".jpg", ".jpeg", ".webp", ".ico"}
    if ext not in allowed:
        flash("صيغة الصورة غير مدعومة", "error")
        return redirect(url_for("settings_page"))
    path=save_image_as_jpg(f, f"{kind}-logo")
    if not path:
        flash("تعذر حفظ الصورة", "error"); return redirect(url_for("settings_page"))
    set_setting({"team":"teamLogo","university":"universityLogo","favicon":"favicon"}[kind], path)
    log_action("Updated logo",kind); flash("تم تحديث الشعار", "ok")
    return redirect(url_for("settings_page"))

# ============================================================ EXPERIENCE CONTROL CENTER ============================================================
@app.route("/admin/navigation", methods=["GET","POST"])
def navigation_page():
    if request.method=="POST":
        db=get_db()
        ids=request.form.getlist("nav_id")
        for idx,nid in enumerate(ids):
            visible=1 if request.form.get(f"visible_{nid}") else 0
            label=request.form.get(f"label_{nid}","").strip()
            icon=request.form.get(f"icon_{nid}","").strip()
            db.execute("UPDATE nav_items SET label=?,icon=?,visible=?,sort_order=? WHERE id=?",(label,icon,visible,(idx+1)*10,nid))
        # Optional custom navigation item
        if request.form.get("new_label") and request.form.get("new_url"):
            db.execute("INSERT INTO nav_items(id,label,endpoint,url,icon,visible,sort_order,is_system,created_at) VALUES(?,?,?,?,?,?,?,?,?)",(uid("nav"),request.form.get("new_label"),"",request.form.get("new_url"),request.form.get("new_icon","↗"),1,999,0,now_iso()))
        db.commit(); log_action("Updated navigation","Public navigation")
        flash("تم حفظ ترتيب وأسماء وإظهار القوائم", "ok")
        return redirect(url_for("navigation_page"))
    return render_template("navigation_admin.html", nav_items=get_db().execute("SELECT * FROM nav_items ORDER BY sort_order").fetchall())

@app.route("/admin/navigation/<nid>/delete", methods=["POST"])
def navigation_delete(nid):
    db=get_db(); row=db.execute("SELECT * FROM nav_items WHERE id=?",(nid,)).fetchone()
    if row and not row["is_system"]: db.execute("DELETE FROM nav_items WHERE id=?",(nid,)); db.commit(); log_action("Deleted navigation",row["label"])
    return redirect(url_for("navigation_page"))

@app.route("/admin/sections", methods=["GET","POST"])
def sections_page():
    if request.method=="POST":
        db=get_db(); ids=request.form.getlist("section_id")
        for idx,sid in enumerate(ids):
            visible=1 if request.form.get(f"visible_{sid}") else 0
            title=request.form.get(f"title_{sid}","").strip()
            db.execute("UPDATE site_sections SET title=?,visible=?,sort_order=?,background=? WHERE id=?",(title,visible,(idx+1)*10,request.form.get(f"background_{sid}","light"),sid))
        db.commit(); log_action("Updated site sections","Public sections"); flash("تم حفظ ترتيب وإظهار أقسام الصفحة", "ok"); return redirect(url_for("sections_page"))
    return render_template("sections_admin.html", sections=get_db().execute("SELECT * FROM site_sections ORDER BY sort_order").fetchall())

@app.route("/admin/permissions", methods=["GET","POST"])
def permissions_page():
    if not is_creator(): return render_template("403.html",permission="permissions"),403
    roles=["ADMIN","SUPER_ADMIN","CREATOR"]
    permissions=[r["permission"] for r in get_db().execute("SELECT DISTINCT permission FROM role_permissions ORDER BY permission").fetchall()]
    if request.method=="POST":
        db=get_db()
        for role in roles:
            for perm in permissions:
                allowed=1 if request.form.get(f"p_{role}_{perm}") else 0
                db.execute("INSERT INTO role_permissions(id,role,permission,allowed) VALUES(?,?,?,?) ON CONFLICT(role,permission) DO UPDATE SET allowed=excluded.allowed",(uid("perm"),role,perm,allowed))
        db.commit(); log_action("Updated permissions","Role matrix"); flash("تم حفظ مصفوفة الصلاحيات", "ok"); return redirect(url_for("permissions_page"))
    matrix={(r["role"],r["permission"]):r["allowed"] for r in get_db().execute("SELECT role,permission,allowed FROM role_permissions").fetchall()}
    return render_template("permissions_admin.html",roles=roles,permissions=permissions,matrix=matrix)

@app.route("/admin/control-center")
def control_center():
    return redirect(url_for("settings_page"))

@app.route("/admin/media/upload", methods=["POST"])
def media_upload():
    if not has_permission("media"): return render_template("403.html",permission="media"),403
    db=get_db(); files=request.files.getlist("media_files") or request.files.getlist("media_file"); saved=0
    for f in files:
        if not f or not f.filename: continue
        ext=os.path.splitext(secure_filename(f.filename))[1].lower()
        if ext in {".jpg",".jpeg",".png",".webp",".gif"}: path=save_image_as_jpg(f,"media"); mtype="image"
        elif ext in {".mp4",".webm",".mov",".m4v",".avi"}: path=save_uploaded_video(f,"media"); mtype="video"
        else: continue
        if not path: continue
        db.execute("INSERT INTO media(id,title,url,category,initiative_id,public,created_at,media_type,sort_order) VALUES(?,?,?,?,?,?,?,?,?)",(uid("med"),request.form.get("title") or secure_filename(f.filename),path,request.form.get("category","عام"),request.form.get("initiative_id") or None,1 if request.form.get("public") else 0,now_iso(),mtype,saved)); saved+=1
    db.commit(); log_action("Uploaded media",f"{saved} file(s)"); flash(f"تم رفع {saved} ملف/ملفات", "ok"); return redirect(url_for("media_admin"))

@app.route("/admin/asset/<kind>", methods=["POST"])
def upload_public_asset(kind):
    if not is_creator(): return render_template("403.html",permission="system"),403
    f=request.files.get("asset_file")
    allowed={"hero_background":{".jpg",".jpeg",".png",".webp"},"hero_video":{".mp4",".webm",".mov",".m4v",".avi"},"public_background":{".jpg",".jpeg",".png",".webp"}}
    if kind not in allowed or not f or not f.filename: flash("اختر ملفًا مناسبًا", "error"); return redirect(url_for("settings_page"))
    ext=os.path.splitext(secure_filename(f.filename))[1].lower()
    if ext not in allowed[kind]: flash("صيغة الملف غير مدعومة", "error"); return redirect(url_for("settings_page"))
    path=save_uploaded_video(f,kind) if kind=="hero_video" else save_image_as_jpg(f,kind)
    set_setting(kind,path); log_action("Uploaded public asset",kind); flash("تم رفع الملف وتفعيله", "ok"); return redirect(url_for("settings_page"))

@app.route("/admin/live/<sid>/recording", methods=["POST"])
def live_recording_upload(sid):
    if not has_permission("live"): return jsonify({"ok":False,"error":"Live permission required"}),403
    db=get_db(); ls=db.execute("SELECT id FROM live_sessions WHERE id=?",(sid,)).fetchone()
    f=request.files.get("recording")
    if not ls or not f or not f.filename: return jsonify({"ok":False,"error":"Recording missing"}),400
    path=save_uploaded_video(f,"live-recording")
    if not path: return jsonify({"ok":False,"error":"صيغة التسجيل غير مدعومة"}),400
    db.execute("UPDATE live_sessions SET recording_path=? WHERE id=?",(path,sid)); db.commit(); log_action("Archived live recording",sid); return jsonify({"ok":True,"path":path})

# ============================================================ LIVE CONTROL ============================================================
@app.route("/admin/live", methods=["GET", "POST"])
def live_admin():
    if not has_permission("live"):
        return render_template("403.html", permission="live"), 403
    db=get_db()
    if request.method=="POST":
        action=request.form.get("action","start_internal")
        if action=="stop":
            sid=request.form.get("session_id") or get_setting("liveSessionId","")
            if sid:
                db.execute("UPDATE live_sessions SET status='ended',ended_at=? WHERE id=?",(now_iso(),sid))
                db.execute("DELETE FROM live_peers WHERE session_id=?",(sid,))
                db.execute("DELETE FROM live_signals WHERE session_id=?",(sid,))
            db.commit(); log_action("Stopped live broadcast",sid); flash("تم إيقاف البث", "ok"); return redirect(url_for("live_admin"))
        if action=="request_source":
            sid=request.form.get("session_id")
            if not sid: flash("اختر بثًا أولاً","error")
            else:
                db.execute("INSERT INTO live_source_requests(id,session_id,requester_name,requester_user_id,device_label,status,created_at) VALUES(?,?,?,?,?,?,?)",
                           (uid("src_req"),sid,current_user()["name"] if current_user() else "زائر",current_user()["id"] if current_user() else None,request.form.get("device_label","هاتف/كاميرا"),"pending",now_iso()))
                db.commit(); flash("تم إرسال طلب إضافة مصدر بث","ok")
            return redirect(url_for("live_admin"))
        if action=="review_source":
            rid=request.form.get("request_id"); decision=request.form.get("decision","approved")
            db.execute("UPDATE live_source_requests SET status=?,reviewed_at=?,reviewed_by=? WHERE id=?",(decision,now_iso(),current_user()["name"] if current_user() else "",rid)); db.commit()
            flash("تم تحديث طلب مصدر البث","ok"); return redirect(url_for("live_admin"))
        title=request.form.get("live_title","البث المباشر من الميدان").strip() or "البث المباشر من الميدان"
        desc=request.form.get("live_description","").strip()
        initiative_id=request.form.get("initiative_id") or None
        mode=request.form.get("mode","internal")
        broadcast_type=request.form.get("broadcast_type","لجنة")
        event_location=request.form.get("event_location","").strip()
        scheduled_date=request.form.get("scheduled_date") or None
        scheduled_time=request.form.get("scheduled_time") or None
        links=[x.strip() for x in request.form.getlist("external_links") if x.strip()]
        if mode=="internal":
            active_count=db.execute("SELECT COUNT(*) c FROM live_sessions WHERE status='live'").fetchone()["c"]
            if active_count >= 6:
                flash("الحد الأقصى 6 بثوث مباشرة في الوقت نفسه","error"); return redirect(url_for("live_admin"))
            sid=uid("live")
            db.execute("""INSERT INTO live_sessions(
                id,title,description,initiative_id,mode,external_url,external_platform,status,created_by,
                started_at,created_at,channel_name,broadcast_type,event_location,scheduled_date,scheduled_time,
                external_links,viewer_peak
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0)""",
                (sid,title,desc,initiative_id,"internal","","Internal","live",current_user()["id"],now_iso(),now_iso(),
                 broadcast_type, broadcast_type, event_location, scheduled_date, scheduled_time, json.dumps(links,ensure_ascii=False)))
            set_setting("liveEnabled","1"); set_setting("liveMode","internal"); set_setting("liveSessionId",sid)
            set_setting("liveTitle",title); set_setting("liveDescription",desc); set_setting("liveUrl",""); set_setting("livePlatform","Internal")
            db.commit(); log_action("Started internal live broadcast",title)
            flash("تم فتح غرفة البث. لم يتم تشغيل أي مصدر حتى تضيف الكاميرا/الهاتف بنفسك.","ok")
            return redirect(url_for("live_room",sid=sid))
        ext=request.form.get("live_url","").strip()
        platform=request.form.get("live_platform","YouTube")
        if ext: links.append(ext)
        sid=uid("live")
        db.execute("""INSERT INTO live_sessions(
            id,title,description,initiative_id,mode,external_url,external_platform,status,created_by,
            started_at,created_at,channel_name,broadcast_type,event_location,scheduled_date,scheduled_time,
            external_links,viewer_peak
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0)""",
            (sid,title,desc,initiative_id,"external",ext,platform,"live",current_user()["id"],now_iso(),now_iso(),
                 broadcast_type, broadcast_type, event_location, scheduled_date, scheduled_time, json.dumps(links,ensure_ascii=False)))
        db.commit()
        set_setting("liveEnabled","1"); set_setting("liveMode","external"); set_setting("liveSessionId",sid)
        set_setting("liveTitle",title); set_setting("liveDescription",desc); set_setting("liveUrl",ext); set_setting("livePlatform",platform)
        log_action("Started external live broadcast",title); flash("تم تفعيل البث الخارجي وحفظ روابطه","ok"); return redirect(url_for("live_admin"))
    active_sessions=db.execute("""SELECT ls.*,i.name initiative_name,
        (SELECT COUNT(*) FROM live_peers lp WHERE lp.session_id=ls.id AND lp.role='viewer') viewer_count,
        (SELECT COUNT(*) FROM live_peers lp WHERE lp.session_id=ls.id AND lp.role IN ('broadcaster','source')) source_count
        FROM live_sessions ls LEFT JOIN initiatives i ON i.id=ls.initiative_id
        WHERE ls.status='live' ORDER BY ls.started_at DESC LIMIT 6""").fetchall()
    sessions=db.execute("SELECT ls.*,i.name initiative_name FROM live_sessions ls LEFT JOIN initiatives i ON i.id=ls.initiative_id ORDER BY ls.created_at DESC LIMIT 50").fetchall()
    source_requests=db.execute("""SELECT r.*,ls.title session_title FROM live_source_requests r
        LEFT JOIN live_sessions ls ON ls.id=r.session_id ORDER BY r.created_at DESC LIMIT 50""").fetchall()
    return render_template("live_admin.html",active_sessions=active_sessions,active_session=(active_sessions[0] if active_sessions else None),
                           sessions=sessions,source_requests=source_requests,
                           initiatives=db.execute("SELECT id,name FROM initiatives ORDER BY date DESC").fetchall())

# ============================================================ PODCAST ============================================================
def _podcast_cover(f): return save_image_as_jpg(f,"podcast-cover") if f and f.filename else ""

@app.route("/podcast")
def podcast_public():
    db=get_db(); shows=db.execute("SELECT * FROM podcast_shows WHERE status='active' ORDER BY updated_at DESC").fetchall(); eps=db.execute("SELECT e.*,g.name guest_name FROM podcast_episodes e LEFT JOIN podcast_guests g ON g.id=e.guest_id WHERE e.status='published' ORDER BY e.published_at DESC,e.created_at DESC").fetchall(); clips=db.execute("SELECT * FROM podcast_clips ORDER BY created_at DESC LIMIT 12").fetchall()
    return render_template("podcast.html",shows=shows,episodes=eps,clips=clips)

@app.route("/podcast/<eid>")
def podcast_episode(eid):
    db=get_db(); e=db.execute("SELECT e.*,g.name guest_name,g.bio guest_bio,g.photo guest_photo,g.specialty guest_specialty FROM podcast_episodes e LEFT JOIN podcast_guests g ON g.id=e.guest_id WHERE e.id=? AND e.status='published'",(eid,)).fetchone()
    if not e:return "الحلقة غير موجودة",404
    clips=db.execute("SELECT * FROM podcast_clips WHERE episode_id=? ORDER BY created_at",(eid,)).fetchall(); return render_template("podcast_episode.html",e=e,clips=clips)

@app.route("/admin/podcast", methods=["GET","POST"])
def podcast_admin():
    if not has_permission("podcast_admin"): return render_template("403.html",permission="podcast_admin"),403
    db=get_db()
    if request.method=="POST":
        action=request.form.get("action")
        if action=="show":
            cover=_podcast_cover(request.files.get("cover")); db.execute("INSERT INTO podcast_shows(id,title,subtitle,description,cover_image,host,status,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?)",(uid("show"),request.form.get("title"),request.form.get("subtitle"),request.form.get("description"),cover,request.form.get("host"),"active",now_iso(),now_iso()))
        elif action=="guest":
            photo=save_image_as_jpg(request.files.get("photo"),"podcast-guest"); db.execute("INSERT INTO podcast_guests(id,name,bio,photo,specialty,created_at) VALUES(?,?,?,?,?,?)",(uid("guest"),request.form.get("name"),request.form.get("bio"),photo,request.form.get("specialty"),now_iso()))
        elif action=="episode":
            cover=_podcast_cover(request.files.get("cover")); audio=save_uploaded_audio(request.files.get("audio"),"podcast-audio"); video=save_uploaded_video(request.files.get("video"),"podcast-video")
            status="published" if request.form.get("publish") else "draft"; pub=now_iso() if status=="published" else None
            db.execute("INSERT INTO podcast_episodes(id,show_id,season,episode_no,title,description,guest_id,host,director,producer,recorded_at,published_at,duration,cover_image,audio_file,video_file,transcript,status,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(uid("ep"),request.form.get("show_id") or None,int(request.form.get("season") or 1),int(request.form.get("episode_no") or 1),request.form.get("title"),request.form.get("description"),request.form.get("guest_id") or None,request.form.get("host"),request.form.get("director"),request.form.get("producer"),request.form.get("recorded_at"),pub,request.form.get("duration"),cover,audio,video,request.form.get("transcript"),status,now_iso(),now_iso()))
        elif action=="equipment":
            db.execute("INSERT INTO podcast_equipment(id,category,name,quantity,status,notes) VALUES(?,?,?,?,?,?)",(uid("eq"),request.form.get("category"),request.form.get("name"),int(request.form.get("quantity") or 1),request.form.get("status","available"),request.form.get("notes")))
        db.commit(); flash("تم الحفظ","ok"); return redirect(url_for("podcast_admin"))
    return render_template("podcast_admin.html",shows=db.execute("SELECT * FROM podcast_shows ORDER BY updated_at DESC").fetchall(),guests=db.execute("SELECT * FROM podcast_guests ORDER BY name").fetchall(),episodes=db.execute("SELECT e.*,g.name guest_name,s.title show_title FROM podcast_episodes e LEFT JOIN podcast_guests g ON g.id=e.guest_id LEFT JOIN podcast_shows s ON s.id=e.show_id ORDER BY e.created_at DESC").fetchall(),equipment=db.execute("SELECT * FROM podcast_equipment ORDER BY category,name").fetchall())

@app.route("/admin/podcast/episodes/<eid>/delete", methods=["POST"])
def podcast_episode_delete(eid):
    if not has_permission("podcast_admin"): return render_template("403.html",permission="podcast_admin"),403
    db=get_db(); db.execute("DELETE FROM podcast_episodes WHERE id=?",(eid,)); db.execute("DELETE FROM podcast_clips WHERE episode_id=?",(eid,)); db.commit(); flash("تم حذف الحلقة","ok"); return redirect(url_for("podcast_admin"))

@app.route("/admin/podcast/clips", methods=["POST"])
def podcast_clip_add():
    if not has_permission("podcast_admin"): return render_template("403.html",permission="podcast_admin"),403
    db=get_db(); video=save_uploaded_video(request.files.get("clip"),"podcast-clip"); cover=save_image_as_jpg(request.files.get("cover"),"podcast-clip-cover");
    if video: db.execute("INSERT INTO podcast_clips(id,episode_id,title,file_path,cover_image,created_at) VALUES(?,?,?,?,?,?)",(uid("clip"),request.form.get("episode_id"),request.form.get("title"),video,cover,now_iso())); db.commit()
    flash("تم حفظ المقطع","ok"); return redirect(url_for("podcast_admin"))

# ============================================================ CARDS / PUBLIC VISIBILITY ============================================================
@app.route("/admin/cards/<kind>/<record_id>/visibility", methods=["POST"])
def card_visibility(kind,record_id):
    if not has_permission("cards"): return render_template("403.html",permission="cards"),403
    table="members" if kind=="member" else "administrators" if kind=="admin" else ""
    if not table:return "نوع البطاقة غير صحيح",404
    db=get_db(); db.execute(f"UPDATE {table} SET card_public=? WHERE id=?",(1 if request.form.get("public") else 0,record_id)); db.commit(); flash("تم تحديث ظهور البطاقة","ok"); return redirect(url_for("cards_admin"))

# ============================================================ PRIVATE CHAT ============================================================
def _require_chat_user():
    u=current_user()
    if not u or not u["active"]: return None
    return u

def _conversation(db, me, other):
    a,b=sorted([me,other])
    row=db.execute("SELECT * FROM chat_conversations WHERE user_a=? AND user_b=?",(a,b)).fetchone()
    if row:return row
    cid=uid("chat"); db.execute("INSERT INTO chat_conversations(id,user_a,user_b,created_at,updated_at) VALUES(?,?,?,?,?)",(cid,a,b,now_iso(),now_iso())); db.commit()
    return db.execute("SELECT * FROM chat_conversations WHERE id=?",(cid,)).fetchone()

def _chat_other(conv, me): return conv["user_b"] if conv["user_a"]==me else conv["user_a"]

@app.route("/chat")
def chat_page():
    u=_require_chat_user()
    if not u:return redirect(url_for("login"))
    db=get_db(); q=request.args.get("q","").strip()
    people=db.execute("SELECT id,name,email,role FROM users WHERE active=1 AND id<>? ORDER BY name",(u["id"],)).fetchall()
    if q: people=[x for x in people if q.lower() in ((x["name"] or "")+" "+(x["email"] or "")).lower()]
    convs=db.execute("SELECT * FROM chat_conversations WHERE user_a=? OR user_b=? ORDER BY updated_at DESC",(u["id"],u["id"])).fetchall()
    items=[]
    for c in convs:
        oid=_chat_other(c,u["id"]); person=db.execute("SELECT id,name,role FROM users WHERE id=?",(oid,)).fetchone(); last=db.execute("SELECT * FROM chat_messages WHERE conversation_id=? ORDER BY created_at DESC LIMIT 1",(c["id"],)).fetchone()
        unread=db.execute("SELECT COUNT(*) c FROM chat_messages WHERE conversation_id=? AND sender_id<>? AND deleted_at IS NULL AND created_at>?",(c["id"],u["id"],c["updated_at"] or "")).fetchone()["c"] if False else 0
        items.append({"conversation":c,"person":person,"last":last,"unread":unread})
    active_id=request.args.get("with")
    active=None; messages=[]; active_person=None
    if active_id:
        active_person=db.execute("SELECT id,name,role FROM users WHERE id=? AND active=1",(active_id,)).fetchone()
        if active_person:
            active=_conversation(db,u["id"],active_id); messages=db.execute("SELECT m.*,u.name sender_name FROM chat_messages m JOIN users u ON u.id=m.sender_id WHERE m.conversation_id=? ORDER BY m.created_at",(active["id"],)).fetchall()
    return render_template("chat.html",people=people,conversations=items,active=active,active_person=active_person,messages=messages)

@app.route("/chat/open/<uid_>")
def chat_open(uid_):
    u=_require_chat_user()
    if not u:return redirect(url_for("login"))
    if uid_==u["id"]: return redirect(url_for("chat_page"))
    other=get_db().execute("SELECT id FROM users WHERE id=? AND active=1",(uid_,)).fetchone()
    if not other:return "المستخدم غير موجود",404
    _conversation(get_db(),u["id"],uid_)
    return redirect(url_for("chat_page",**{"with":uid_}))

@app.post("/chat/<cid>/send")
def chat_send(cid):
    u=_require_chat_user()
    if not u:return jsonify({"ok":False,"error":"login"}),401
    db=get_db(); c=db.execute("SELECT * FROM chat_conversations WHERE id=? AND (user_a=? OR user_b=?)",(cid,u["id"],u["id"])).fetchone()
    if not c:return "المحادثة غير موجودة",404
    body=(request.form.get("body") or "").strip(); f=request.files.get("media")
    path=""; media_type=""
    if f and f.filename:
        ext=os.path.splitext(secure_filename(f.filename))[1].lower()
        if ext in {".jpg",".jpeg",".png",".webp",".gif"}: path=save_image_as_jpg(f,"chat"); media_type="image"
        elif ext in {".mp4",".webm",".mov",".m4v",".avi"}: path=save_uploaded_video(f,"chat"); media_type="video"
        elif ext in {".mp3",".wav",".m4a",".ogg"}: path=save_uploaded_audio(f,"chat"); media_type="audio"
    if not body and not path:return redirect(url_for("chat_page",**{"with":_chat_other(c,u["id"])}))
    mid=uid("msg"); db.execute("INSERT INTO chat_messages(id,conversation_id,sender_id,body,media_path,media_type,created_at) VALUES(?,?,?,?,?,?,?)",(mid,cid,u["id"],body,path,media_type,now_iso())); db.execute("UPDATE chat_conversations SET updated_at=? WHERE id=?",(now_iso(),cid)); other=_chat_other(c,u["id"]); db.execute("INSERT INTO notifications(id,user_id,target_role,type,title,body,url,created_at) VALUES(?,?,?,?,?,?,?,?)",(uid("notif"),other,"","chat","رسالة جديدة",f"لديك رسالة جديدة من {u['name']}",url_for("chat_page",**{"with":u["id"]}),now_iso())); db.commit(); return redirect(url_for("chat_page",**{"with":other}))

@app.post("/chat/message/<mid>/edit")
def chat_edit(mid):
    u=_require_chat_user(); db=get_db()
    row=db.execute("SELECT m.*,c.user_a,c.user_b FROM chat_messages m JOIN chat_conversations c ON c.id=m.conversation_id WHERE m.id=? AND m.sender_id=?",(mid,u["id"] if u else "")).fetchone()
    if not row:return "الرسالة غير موجودة",404
    body=(request.form.get("body") or "").strip()
    if body: db.execute("UPDATE chat_messages SET body=?,edited_at=? WHERE id=?",(body,now_iso(),mid)); db.commit()
    return redirect(url_for("chat_page",**{"with":_chat_other(row,u["id"])}))

@app.post("/chat/message/<mid>/delete")
def chat_delete(mid):
    u=_require_chat_user(); db=get_db(); row=db.execute("SELECT m.*,c.user_a,c.user_b FROM chat_messages m JOIN chat_conversations c ON c.id=m.conversation_id WHERE m.id=? AND m.sender_id=?",(mid,u["id"] if u else "")).fetchone()
    if not row:return "الرسالة غير موجودة",404
    db.execute("UPDATE chat_messages SET body='',media_path='',media_type='',deleted_at=? WHERE id=?",(now_iso(),mid)); db.commit(); return redirect(url_for("chat_page",**{"with":_chat_other(row,u["id"])}))

@app.post("/chat/message/<mid>/react")
def chat_react(mid):
    u=_require_chat_user(); db=get_db(); msg=db.execute("SELECT m.*,c.user_a,c.user_b FROM chat_messages m JOIN chat_conversations c ON c.id=m.conversation_id WHERE m.id=? AND (c.user_a=? OR c.user_b=?)",(mid,u["id"] if u else "",u["id"] if u else "")).fetchone()
    if not msg:return "الرسالة غير موجودة",404
    reaction=request.form.get("reaction","❤️")[:20]; existing=db.execute("SELECT id FROM chat_reactions WHERE message_id=? AND user_id=?",(mid,u["id"])).fetchone()
    if existing: db.execute("UPDATE chat_reactions SET reaction=?,created_at=? WHERE id=?",(reaction,now_iso(),existing["id"]))
    else: db.execute("INSERT INTO chat_reactions(id,message_id,user_id,reaction,created_at) VALUES(?,?,?,?,?)",(uid("react"),mid,u["id"],reaction,now_iso()))
    db.commit(); return redirect(url_for("chat_page",**{"with":_chat_other(msg,u["id"])}))

# ============================================================ SETTINGS ============================================================
@app.route("/settings", methods=["GET", "POST"])
def settings_page():
    weights = get_weights()
    points_cfg = get_points_config()
    if request.method == "POST":
        form_type = request.form.get("form_type")
        if form_type == "identity":
            set_setting("teamName", request.form.get("team_name", "AHVT"))
            set_setting("subtitle", request.form.get("subtitle", ""))
            flash("تم الحفظ بنجاح", "ok")
        elif form_type == "weights":
            new_w = {k: int(request.form.get(f"w_{k}", 0) or 0) for k in CRITERIA_KEYS}
            if sum(new_w.values()) != 100:
                flash("مجموع الأوزان يجب أن يساوي 100%", "error")
                return redirect(url_for("settings_page"))
            set_setting("weights", json.dumps(new_w))
            flash("تم حفظ الأوزان", "ok")
        elif form_type == "points":
            new_p = {k: int(request.form.get(f"p_{k}", 0) or 0) for k in DEFAULT_POINTS.keys()}
            set_setting("points", json.dumps(new_p))
            flash("تم حفظ القيم", "ok")
        elif form_type == "appearance":
            for key, field in [("accentColor","accent_color"),("navyColor","navy_color"),("backgroundColor","background_color"),("fontFamily","font_family"),("heroTitle","hero_title"),("heroText","hero_text"),("announcement","announcement"),("customCss","custom_css"),("heroBackground","hero_background"),("heroVideo","hero_video"),("publicBackground","public_background"),("seoTitle","seo_title"),("seoDescription","seo_description"),("certificateIssuer","certificate_issuer")]:
                set_setting(key, request.form.get(field, ""))
            set_setting("cinematicMode", "1" if request.form.get("cinematic_mode") else "0")
            flash("تم تحديث الهوية والمظهر", "ok")
        elif form_type == "public_controls":
            for key, field in [("joinButtonVisible","join_visible"),("joinButtonText","join_text"),("joinButtonIcon","join_icon"),("joinButtonPlacement","join_placement"),("joinButtonMode","join_mode"),("adminLinkVisible","admin_link_visible"),("telegramUrl","telegram_url"),("showPublicAdmins","show_admins"),("showPublicNews","show_news")]:
                set_setting(key, request.form.get(field, "0" if field in ("join_visible","admin_link_visible","show_admins","show_news") else ""))
            flash("تم تحديث زر الانضمام والتحكم العام", "ok")
        elif form_type == "payment":
            if not is_admin():
                flash("معلومات الدفع متاحة للإدارة فقط", "error")
            else:
                for key, field in [("payment_name","payment_name"),("payment_provider","payment_provider"),("payment_account","payment_account"),("payment_instructions","payment_instructions")]:
                    set_setting(key, request.form.get(field, ""))
                flash("تم حفظ معلومات الدفع", "ok")
        elif form_type == "permissions_toggle":
            if not is_creator():
                flash("تفعيل نظام الصلاحيات متاح للمالك فقط", "error")
            else:
                set_setting("permissionsEnabled", "1" if request.form.get("permissions_enabled") else "0")
                flash("تم تحديث نظام الصلاحيات", "ok")
        return redirect(url_for("settings_page"))
    return render_template("settings.html", weights=weights, points_cfg=points_cfg,
                           criteria_keys=CRITERIA_KEYS, criteria_labels=CRITERIA_LABELS,
                           points_labels={"attendance": "حضور", "task": "إنجاز مهمة", "leader": "قيادة مبادرة",
                                          "participation": "مشاركة فعالة", "excellent": "تقييم ممتاز"})



# ============================================================ COMPREHENSIVE REPORT EXPORT ============================================================
REPORT_FONT_REG = os.path.join(BASE_DIR, "static", "fonts", "NotoSansArabic-Regular.ttf")
REPORT_FONT_BOLD = os.path.join(BASE_DIR, "static", "fonts", "NotoSansArabic-Bold.ttf")
try:
    if os.path.exists(REPORT_FONT_REG):
        pdfmetrics.registerFont(TTFont("HikmaArabic", REPORT_FONT_REG))
    if os.path.exists(REPORT_FONT_BOLD):
        pdfmetrics.registerFont(TTFont("HikmaArabicBold", REPORT_FONT_BOLD))
except Exception:
    pass


def _report_ar(value):
    """Safely prepare Arabic/text for ReportLab Paragraphs. Escaping prevents user data like < > & from breaking PDF generation."""
    from xml.sax.saxutils import escape
    text = "—" if value is None or str(value).strip() == "" else str(value)
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        text = get_display(arabic_reshaper.reshape(text))
    except Exception:
        pass
    return escape(text).replace("\n", "<br/>")


def _report_font(bold=False):
    return "HikmaArabicBold" if bold and "HikmaArabicBold" in pdfmetrics.getRegisteredFontNames() else ("HikmaArabic" if "HikmaArabic" in pdfmetrics.getRegisteredFontNames() else "Helvetica")


def _report_rows(db, start=None, end=None):
    def filtered(sql, args=()):
        return db.execute(sql, args).fetchall()

    members = filtered("SELECT * FROM members ORDER BY name")
    admins = filtered("SELECT * FROM administrators ORDER BY name")
    committees = filtered("SELECT * FROM committees ORDER BY name")
    initiatives = filtered("SELECT * FROM initiatives ORDER BY date DESC, name")
    participants = filtered("""SELECT ip.*, i.name initiative_name, m.name member_name
        FROM initiative_participants ip
        LEFT JOIN initiatives i ON i.id=ip.initiative_id
        LEFT JOIN members m ON m.id=ip.member_id
        ORDER BY i.date DESC, i.name, m.name""")
    attendance = filtered("""SELECT a.*, m.name member_name FROM attendance a
        LEFT JOIN members m ON m.id=a.member_id ORDER BY a.date DESC, m.name""")
    tasks = filtered("SELECT * FROM tasks ORDER BY deadline ASC, title")
    events = filtered("SELECT * FROM events ORDER BY date DESC, time DESC")
    news = filtered("SELECT * FROM news ORDER BY published_at DESC, created_at DESC")
    evaluations = filtered("""SELECT e.*, m.name member_name FROM evaluations e
        LEFT JOIN members m ON m.id=e.evaluated_user_id ORDER BY e.date DESC""")
    points = filtered("""SELECT p.*, m.name member_name FROM points p
        LEFT JOIN members m ON m.id=p.member_id ORDER BY p.date DESC""")
    applications = filtered("SELECT * FROM applications ORDER BY created_at DESC")
    certificates = filtered("SELECT * FROM certificates ORDER BY issue_date DESC, created_at DESC")
    podcast = filtered("SELECT * FROM podcast_episodes ORDER BY published_at DESC, created_at DESC") if _table_exists(db, "podcast_episodes") else []
    live = filtered("SELECT * FROM live_sessions ORDER BY created_at DESC")

    # Optional period filtering for activity-like tables while leaving master directories complete.
    if start or end:
        def in_period(row, key):
            v = str(row[key] or "") if key in row.keys() else ""
            return (not start or v >= start) and (not end or v <= end)
        initiatives = [r for r in initiatives if in_period(r, "date")]
        participants = [r for r in participants if any(i["id"] == r["initiative_id"] for i in initiatives)]
        attendance = [r for r in attendance if in_period(r, "date")]
        events = [r for r in events if in_period(r, "date")]
        news = [r for r in news if in_period(r, "published_at") or in_period(r, "created_at")]
        evaluations = [r for r in evaluations if in_period(r, "date")]
        points = [r for r in points if in_period(r, "date")]
        certificates = [r for r in certificates if in_period(r, "issue_date")]
        tasks = [r for r in tasks if in_period(r, "deadline") or in_period(r, "created_at")]

    attendance_hours = sum(float(r["hours"] or 0) for r in attendance if not r["initiative_id"])
    initiative_hours = sum(float(r["hours"] or 0) for r in participants)
    initiative_declared_hours = sum(float(r["hours"] or 0) for r in initiatives)
    total_hours = attendance_hours + initiative_hours
    member_hours_map = {m["id"]: 0.0 for m in members}
    for r in attendance:
        if r["member_id"] in member_hours_map and not r["initiative_id"]:
            member_hours_map[r["member_id"]] += float(r["hours"] or 0)
    for r in participants:
        if r["member_id"] in member_hours_map:
            member_hours_map[r["member_id"]] += float(r["hours"] or 0)
    member_points_map = {m["id"]: 0.0 for m in members}
    for r in points:
        if r["member_id"] in member_points_map:
            member_points_map[r["member_id"]] += float(r["value"] or 0)
    published_news = sum(1 for r in news if (r["status"] or "").lower() == "published")
    active_members = sum(1 for r in members if not r["status"] or str(r["status"]).lower() in ("active", "نشط"))
    active_admins = len(admins)
    return locals()


def _table_exists(db, name):
    return bool(db.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (name,)).fetchone())


def _report_pdf(path, data, start=None, end=None):
    page = landscape(A4)
    doc = SimpleDocTemplate(path, pagesize=page, rightMargin=7*mm, leftMargin=7*mm, topMargin=7*mm, bottomMargin=7*mm,
                            title="AHVT — التقرير الشامل", author="AHVT")
    font = _report_font(False); bold = _report_font(True)
    navy = colors.HexColor("#071625"); navy2 = colors.HexColor("#0D2740"); accent = colors.HexColor("#20B486")
    text = colors.HexColor("#172033"); muted = colors.HexColor("#64748B"); line = colors.HexColor("#D7DEE8")
    styles = {
        "title": ParagraphStyle("rt", fontName=bold, fontSize=22, leading=26, alignment=TA_RIGHT, textColor=colors.white),
        "sub": ParagraphStyle("rs", fontName=font, fontSize=9, leading=13, alignment=TA_RIGHT, textColor=colors.HexColor("#C8D8E8")),
        "h": ParagraphStyle("rh", fontName=bold, fontSize=12, leading=15, alignment=TA_RIGHT, textColor=navy),
        "p": ParagraphStyle("rp", fontName=font, fontSize=7.1, leading=9.2, alignment=TA_RIGHT, textColor=text),
        "small": ParagraphStyle("rsm", fontName=font, fontSize=6.2, leading=7.6, alignment=TA_RIGHT, textColor=text),
        "center": ParagraphStyle("rc", fontName=font, fontSize=7, leading=9, alignment=TA_CENTER, textColor=text),
        "stat": ParagraphStyle("rst", fontName=bold, fontSize=13, leading=16, alignment=TA_CENTER, textColor=navy),
    }
    story=[]
    title = get_setting("teamName", "فريق الحكمة التطوعي")
    story.append(Table([[Paragraph(_report_ar("AHVT"), styles["title"]), Paragraph(_report_ar(title), styles["title"])]], colWidths=[115*mm, 150*mm], rowHeights=[24*mm], style=[
        ("BACKGROUND",(0,0),(-1,-1),navy),("LINEBELOW",(0,0),(-1,-1),4,accent),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("LEFTPADDING",(0,0),(-1,-1),8*mm),("RIGHTPADDING",(0,0),(-1,-1),8*mm)
    ]))
    period = f"{start or 'من البداية'} — {end or 'حتى الآن'}"
    story.append(Spacer(1,3*mm)); story.append(Paragraph(_report_ar(f"التقرير التنفيذي الشامل · الفترة: {period} · تاريخ الإصدار: {date.today().isoformat()}"), styles["sub"]))
    story.append(Spacer(1,4*mm))
    stats = [
        ("الأعضاء", len(data["members"])), ("الإداريون", data["active_admins"]), ("المبادرات", len(data["initiatives"])),
        ("الفعاليات", len(data["events"])), ("الأخبار", data["published_news"]), ("ساعات التطوع", round(data["total_hours"],2)),
        ("المهام", len(data["tasks"])), ("الشهادات", len(data["certificates"]))
    ]
    stat_cells=[]
    for label,val in stats:
        stat_cells.append([Paragraph(_report_ar(label), styles["small"]), Paragraph(_report_ar(val), styles["stat"])])
    stat_table=Table([stat_cells[:4],stat_cells[4:]], colWidths=[66*mm]*4, rowHeights=[15*mm,15*mm])
    stat_table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#EEF3F8")),("BOX",(0,0),(-1,-1),0.4,line),("INNERGRID",(0,0),(-1,-1),0.35,line),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("ALIGN",(0,0),(-1,-1),"CENTER")]))
    story.append(stat_table); story.append(Spacer(1,4*mm))

    def section(title, headers, rows, widths=None, font_size=6.2, max_rows=None):
        elems=[Paragraph(_report_ar(title), styles["h"]), Spacer(1,1.5*mm)]
        rows = list(rows)
        if max_rows: rows=rows[:max_rows]
        data_rows=[[Paragraph(_report_ar(h), styles["small"]) for h in headers]]
        for row in rows:
            data_rows.append([Paragraph(_report_ar(v), styles["small"]) for v in row])
        if not rows:
            data_rows.append([Paragraph(_report_ar("لا توجد بيانات"), styles["small"])] + [""]*(len(headers)-1))
        t=Table(data_rows, colWidths=widths, repeatRows=1, hAlign="RIGHT")
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),navy2),("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),0.25,line),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("ALIGN",(0,0),(-1,-1),"RIGHT"),("LEFTPADDING",(0,0),(-1,-1),2.2),("RIGHTPADDING",(0,0),(-1,-1),2.2),
            ("TOPPADDING",(0,0),(-1,-1),1.7),("BOTTOMPADDING",(0,0),(-1,-1),1.7),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#F7F9FC")])
        ]))
        elems.append(t); elems.append(Spacer(1,3*mm)); return elems

    # Compact directories and activity data.
    story += section("الأعضاء", ["الاسم","القسم","المرحلة","اللجنة","المنصب","الحالة","الانضمام","الساعات","النقاط"], [
        [r["name"],r["department"],r["stage"],r["committee"],r["position"],r["status"] or "نشط",r["join_date"],round(data["member_hours_map"].get(r["id"],0),2),round(data["member_points_map"].get(r["id"],0),2)] for r in data["members"]
    ], [40*mm,28*mm,18*mm,28*mm,24*mm,18*mm,22*mm,18*mm,18*mm])
    story += section("اللجان", ["اللجنة","الرئيس","الوصف"], [[r["name"],r["head"],r["description"]] for r in data["committees"]], [55*mm,50*mm,110*mm])
    story += section("الإداريون", ["الاسم","الإدارة/اللجنة","المنصب","التاريخ","المسؤوليات"], [[r["name"],r["committee"],r["position"],r["date"],r["responsibilities"]] for r in data["admins"]], [42*mm,35*mm,30*mm,25*mm,82*mm])
    story += section("المبادرات والمشاركون والساعات", ["المبادرة","التاريخ","الموقع","اللجنة","المدير","الحالة","ساعات المبادرة","المشاركون"], [[r["name"],r["date"],r["location"],r["committee"],r["manager"],r["status"],round(float(r["hours"] or 0),2), initiative_participant_count(get_db(),r["id"])] for r in data["initiatives"]], [48*mm,22*mm,38*mm,28*mm,30*mm,22*mm,24*mm,22*mm])
    story += section("الحضور وساعات التطوع", ["العضو","التاريخ","الحالة","المبادرة","من","إلى","الساعات"], [[r["member_name"],r["date"],r["status"],r["initiative_id"],r["start_time"],r["end_time"],round(float(r["hours"] or 0),2)] for r in data["attendance"]], [55*mm,24*mm,25*mm,55*mm,25*mm,25*mm,20*mm])
    story += section("الفعاليات والنشاطات", ["العنوان","التاريخ","الوقت","الموقع","الحالة","الوصف"], [[r["title"],r["date"],r["time"],r["location"],r["status"],r["description"]] for r in data["events"]], [52*mm,24*mm,20*mm,40*mm,25*mm,64*mm])
    story += section("الأخبار والنشر", ["العنوان","التصنيف","الكاتب","الحالة","النشر"], [[r["title"],r["category"],r["author"],r["status"],r["published_at"]] for r in data["news"]], [92*mm,38*mm,42*mm,28*mm,30*mm])
    story += section("المهام", ["المهمة","المكلف","الموعد","الأولوية","الحالة"], [[r["title"],r["assignee"],r["deadline"],r["priority"],r["status"]] for r in data["tasks"]], [80*mm,55*mm,30*mm,30*mm,35*mm])
    story += section("التقييمات", ["العضو","التاريخ","النوع","الحضور","المهام","المبادرات","الالتزام","الفريق","الإبداع"], [[r["member_name"],r["date"],r["type"],r["c_attendance"],r["c_taskCompletion"],r["c_initiativeParticipation"],r["c_commitment"],r["c_teamwork"],r["c_creativity"]] for r in data["evaluations"]], [45*mm,22*mm,28*mm,22*mm,22*mm,24*mm,22*mm,22*mm,22*mm])
    story += section("الشهادات", ["رقم الشهادة","المستفيد","النوع","المبادرة","التاريخ","الساعات","الكاتب","الحالة"], [[r["certificate_no"],r["recipient_name"],r["certificate_type"],r["initiative_name"],r["issue_date"],r["hours"],r["writer_name"],r["status"]] for r in data["certificates"]], [34*mm,42*mm,34*mm,42*mm,24*mm,20*mm,34*mm,22*mm])
    story += section("الأهداف والطلبات", ["الهدف","المستهدف","الحالي","الفترة","الحالة"], [[r["title"],r["target"],r["current"],r["period"],r["status"]] for r in db.execute("SELECT * FROM goals ORDER BY created_at DESC").fetchall()], [75*mm,30*mm,30*mm,35*mm,35*mm])
    story.append(Spacer(1,2*mm)); story.append(Paragraph(_report_ar(f"إجمالي ساعات الحضور: {data['attendance_hours']:.2f} · ساعات المشاركات بالمبادرات: {data['initiative_hours']:.2f} · إجمالي الساعات المحسوبة: {data['total_hours']:.2f} · المبادرات بساعات معلنة: {data['initiative_declared_hours']:.2f}"), styles["small"]))

    def footer(canvas, doc):
        canvas.saveState(); canvas.setFont(font,6); canvas.setFillColor(muted)
        canvas.drawRightString(page[0]-7*mm,4*mm,_report_ar(f"فريق الحكمة التطوعي · AHVT · صفحة {doc.page}")); canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)


def _report_xlsx(path, data):
    wb=Workbook(); wb.remove(wb.active)
    navy="071625"; accent="20B486"; white="FFFFFF"; thin=Side(style="thin", color="D7DEE8")
    def sheet(name, headers, rows):
        ws=wb.create_sheet(name[:31]); ws.sheet_view.rightToLeft=True
        ws.append(headers)
        for r in rows: ws.append(list(r))
        for cell in ws[1]:
            cell.font=XLFont(bold=True,color=white); cell.fill=PatternFill("solid",fgColor=navy); cell.alignment=Alignment(horizontal="center")
        for row in ws.iter_rows():
            for c in row:
                c.border=Border(bottom=thin); c.alignment=Alignment(vertical="top",wrap_text=True)
        for col in ws.columns:
            width=min(max(max(len(str(c.value or "")) for c in col)+2,10),36); ws.column_dimensions[get_column_letter(col[0].column)].width=width
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
        return ws
    sheet("ملخص", ["المؤشر","القيمة"], [("الأعضاء",len(data["members"])),("الإداريون",data["active_admins"]),("المبادرات",len(data["initiatives"])),("الفعاليات",len(data["events"])),("الأخبار المنشورة",data["published_news"]),("ساعات التطوع",data["total_hours"]),("المهام",len(data["tasks"])),("الشهادات",len(data["certificates"]))])
    sheet("الأعضاء", ["الاسم","البريد","الهاتف","القسم","المرحلة","اللجنة","المنصب","تاريخ الانضمام","الحالة","الساعات","النقاط"], [[r["name"],r["email"],r["phone"],r["department"],r["stage"],r["committee"],r["position"],r["join_date"],r["status"],data["member_hours_map"].get(r["id"],0),data["member_points_map"].get(r["id"],0)] for r in data["members"]])
    sheet("اللجان", ["اللجنة","الرئيس","الوصف"], [[r["name"],r["head"],r["description"]] for r in data["committees"]])
    sheet("الإداريون", ["الاسم","المنصب","اللجنة","التاريخ","المسؤوليات"], [[r["name"],r["position"],r["committee"],r["date"],r["responsibilities"]] for r in data["admins"]])
    sheet("المبادرات", ["الاسم","التاريخ","الموقع","المدير","اللجنة","الساعات","الحالة","الوصف","الأهداف"], [[r["name"],r["date"],r["location"],r["manager"],r["committee"],r["hours"],r["status"],r["description"],r["goals"]] for r in data["initiatives"]])
    sheet("المشاركون", ["المبادرة","العضو","من","إلى","الساعات"], [[r["initiative_name"],r["member_name"],r["start_time"],r["end_time"],r["hours"]] for r in data["participants"]])
    sheet("الحضور", ["العضو","التاريخ","الحالة","المبادرة","من","إلى","الساعات"], [[r["member_name"],r["date"],r["status"],r["initiative_id"],r["start_time"],r["end_time"],r["hours"]] for r in data["attendance"]])
    sheet("الفعاليات", ["العنوان","التاريخ","الوقت","الموقع","الحالة","الوصف"], [[r["title"],r["date"],r["time"],r["location"],r["status"],r["description"]] for r in data["events"]])
    sheet("الأخبار", ["العنوان","التصنيف","الكاتب","الحالة","النشر","المحتوى المختصر"], [[r["title"],r["category"],r["author"],r["status"],r["published_at"],r["excerpt"]] for r in data["news"]])
    sheet("المهام", ["المهمة","المكلف","الموعد","الأولوية","الحالة","الوصف"], [[r["title"],r["assignee"],r["deadline"],r["priority"],r["status"],r["description"]] for r in data["tasks"]])
    sheet("التقييمات", ["العضو","التاريخ","النوع","الحضور","المهام","المبادرات","الالتزام","الفريق","الإبداع","الملاحظات"], [[r["member_name"],r["date"],r["type"],r["c_attendance"],r["c_taskCompletion"],r["c_initiativeParticipation"],r["c_commitment"],r["c_teamwork"],r["c_creativity"],r["notes"]] for r in data["evaluations"]])
    sheet("النقاط", ["العضو","القيمة","المصدر","التاريخ"], [[r["member_name"],r["value"],r["source"],r["date"]] for r in data["points"]])
    sheet("الشهادات", ["الرقم","المستفيد","النوع","المبادرة","التاريخ","الساعات","الكاتب","الحالة"], [[r["certificate_no"],r["recipient_name"],r["certificate_type"],r["initiative_name"],r["issue_date"],r["hours"],r["writer_name"],r["status"]] for r in data["certificates"]])
    wb.save(path)


@app.route("/reports/export/complete.pdf")
def export_complete_pdf():
    if not has_permission("reports"): return render_template("403.html", permission="reports"),403
    start=request.args.get("from"," ").strip() or None; end=request.args.get("to"," ").strip() or None
    try:
        data=_report_rows(get_db(),start,end)
        buf=io.BytesIO(); _report_pdf(buf,data,start,end); buf.seek(0)
    except Exception:
        # Never expose a 500 to the owner: create a compact fallback PDF with the
        # headline statistics and the exact error in the server log only.
        app.logger.exception("Complete PDF export failed; using fallback PDF")
        data=_report_rows(get_db(),start,end)
        buf=io.BytesIO(); doc=SimpleDocTemplate(buf,pagesize=landscape(A4),rightMargin=10*mm,leftMargin=10*mm,topMargin=10*mm,bottomMargin=10*mm)
        f=_report_font(False); b=_report_font(True)
        st=[ParagraphStyle("fb1",fontName=b,fontSize=22,leading=26,alignment=TA_RIGHT,textColor=colors.HexColor("#071625")),ParagraphStyle("fb2",fontName=f,fontSize=10,leading=14,alignment=TA_RIGHT,textColor=colors.HexColor("#334155"))]
        story=[Table([[Paragraph(_report_ar("AHVT · التقرير الشامل"),st[0])]],colWidths=[277*mm],rowHeights=[20*mm],style=[("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#071625")),("LINEBELOW",(0,0),(-1,-1),4,colors.HexColor("#20B486")),("RIGHTPADDING",(0,0),(-1,-1),8*mm)]),Spacer(1,5*mm)]
        for label,val in [("الأعضاء",len(data["members"])),("الإداريون",data["active_admins"]),("المبادرات",len(data["initiatives"])),("الفعاليات",len(data["events"])),("المهام",len(data["tasks"])),("الشهادات",len(data["certificates"])),("ساعات التطوع",round(data["total_hours"],2))]:
            story.append(Paragraph(_report_ar(f"{label}: {val}"),st[1])); story.append(Spacer(1,1.5*mm))
        doc.build(story); buf.seek(0)
    return send_file(buf,mimetype="application/pdf",as_attachment=True,download_name=f"AHVT-Complete-Report-{date.today().isoformat()}.pdf")


@app.route("/reports/export/complete.xlsx")
def export_complete_xlsx():
    if not has_permission("reports"): return render_template("403.html", permission="reports"),403
    start=request.args.get("from"," ").strip() or None; end=request.args.get("to"," ").strip() or None
    data=_report_rows(get_db(),start,end)
    buf=io.BytesIO(); _report_xlsx(buf,data); buf.seek(0)
    return send_file(buf,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",as_attachment=True,download_name=f"AHVT-Complete-Data-{date.today().isoformat()}.xlsx")


# ============================================================ REPORTS ============================================================
@app.route("/reports")
def reports_page():
    db = get_db()
    members = db.execute("SELECT * FROM members ORDER BY name").fetchall()
    initiatives = db.execute("SELECT * FROM initiatives ORDER BY date DESC").fetchall()
    return render_template("reports.html", members=members, initiatives=initiatives)


@app.route("/reports/member/<mid>")
def report_member(mid):
    db = get_db()
    m = db.execute("SELECT * FROM members WHERE id=?", (mid,)).fetchone()
    if not m:
        return redirect(url_for("reports_page"))
    s = member_score(db, mid)
    a = member_attendance_pct(db, mid)
    p = member_points_total(db, mid)
    volunteer_hours = member_volunteer_hours(db, mid)
    weights = get_weights()
    initiatives = db.execute("""SELECT i.* FROM initiatives i
        JOIN initiative_participants ip ON ip.initiative_id=i.id WHERE ip.member_id=?""", (mid,)).fetchall()
    return render_template("report_member.html", m=m, score=s, att=a, points=p, volunteer_hours=volunteer_hours, weights=weights,
        criteria_labels=CRITERIA_LABELS, initiatives=initiatives,
        recommendation=recommendation_text(s), today=date.today().isoformat())


@app.route("/reports/initiative/<iid>")
def report_initiative(iid):
    db = get_db()
    i = db.execute("SELECT * FROM initiatives WHERE id=?", (iid,)).fetchone()
    if not i:
        return redirect(url_for("reports_page"))
    participants = db.execute("""SELECT m.* FROM members m
        JOIN initiative_participants ip ON ip.member_id=m.id WHERE ip.initiative_id=?""", (iid,)).fetchall()
    return render_template("report_initiative.html", i=i, participants=participants, today=date.today().isoformat())


@app.route("/reports/monthly")
def report_monthly():
    db = get_db()
    today = date.today()
    initiatives = db.execute("SELECT * FROM initiatives").fetchall()
    ins = [i for i in initiatives if i["date"] and str(i["date"])[:7] == today.strftime("%Y-%m")]
    members_count = db.execute("SELECT COUNT(*) c FROM members").fetchone()["c"]
    hours = sum((i["hours"] or 0) for i in ins)
    return render_template("report_period.html", title="التقرير الشهري للفريق", initiatives=ins,
        members_count=members_count, hours=hours, today=today.isoformat())


@app.route("/reports/annual")
def report_annual():
    db = get_db()
    year = date.today().year
    initiatives = db.execute("SELECT * FROM initiatives").fetchall()
    ins = [i for i in initiatives if i["date"] and str(i["date"])[:4] == str(year)]
    members_count = db.execute("SELECT COUNT(*) c FROM members").fetchone()["c"]
    hours = sum((i["hours"] or 0) for i in ins)
    return render_template("report_period.html", title=f"التقرير السنوي للفريق — {year}", initiatives=ins,
        members_count=members_count, hours=hours, today=date.today().isoformat())


@app.route("/admin/site-notifications", methods=["GET","POST"])
def site_notifications_admin():
    if not has_permission("site_notifications"): return render_template("403.html",permission="site_notifications"),403
    db=get_db()
    if request.method=="POST":
        action=request.form.get("action")
        if action=="create":
            db.execute("INSERT INTO site_notifications(id,title,body,active,created_at,expires_at) VALUES(?,?,?,?,?,?)",(uid("snt"),request.form.get("title"),request.form.get("body"),1,now_iso(),request.form.get("expires_at") or None))
        elif action=="toggle":
            db.execute("UPDATE site_notifications SET active=CASE WHEN active=1 THEN 0 ELSE 1 END WHERE id=?",(request.form.get("id"),))
        elif action=="delete":
            db.execute("DELETE FROM site_notifications WHERE id=?",(request.form.get("id"),))
        db.commit(); flash("تم تحديث إشعارات الموقع","ok"); return redirect(url_for("site_notifications_admin"))
    return render_template("site_notifications_admin.html",alerts=db.execute("SELECT * FROM site_notifications ORDER BY created_at DESC").fetchall())

# ============================================================ EXTENDED PUBLIC/ADMIN MODULES ============================================================
@app.route("/academy")
def academy_public():
    db=get_db()
    courses=db.execute("SELECT c.*, COUNT(l.id) lesson_count FROM academy_courses c LEFT JOIN academy_lessons l ON l.course_id=c.id WHERE c.public=1 AND c.status='published' GROUP BY c.id ORDER BY c.created_at DESC").fetchall()
    episodes=db.execute("SELECT id,title,cover_image,season,episode_no,duration,published_at FROM podcast_episodes WHERE status='published' ORDER BY published_at DESC LIMIT 6").fetchall()
    recordings=db.execute("SELECT id,title,description,recording_path,started_at,ended_at,viewer_peak FROM live_sessions WHERE recording_path IS NOT NULL ORDER BY started_at DESC LIMIT 6").fetchall()
    return render_template("academy_public.html", courses=courses, episodes=episodes, recordings=recordings)

@app.route("/academy/<cid>")
def academy_course_public(cid):
    db=get_db(); course=db.execute("SELECT * FROM academy_courses WHERE id=? AND public=1 AND status='published'",(cid,)).fetchone()
    if not course:return "الدورة غير موجودة",404
    lessons=db.execute("SELECT * FROM academy_lessons WHERE course_id=? ORDER BY sort_order,created_at",(cid,)).fetchall()
    learner_key=session.get("academy_learner_key")
    enrollment=db.execute("SELECT * FROM academy_enrollments WHERE course_id=? AND learner_key=?",(cid,learner_key)).fetchone() if learner_key else None
    if enrollment and int(course["is_paid"] or 0):
        approved=db.execute("SELECT 1 FROM academy_payments WHERE enrollment_id=? AND status='approved' LIMIT 1",(enrollment["id"],)).fetchone()
        if not approved:
            enrollment=None
    done=set()
    if enrollment:
        done={r["lesson_id"] for r in db.execute("SELECT lesson_id FROM academy_progress WHERE enrollment_id=? AND completed=1",(enrollment["id"],)).fetchall()}
    return render_template("academy_course.html",course=course,lessons=lessons,done=done,enrollment=enrollment)

@app.route("/academy/enroll/<cid>", methods=["POST"])
def academy_enroll(cid):
    db=get_db(); course=db.execute("SELECT * FROM academy_courses WHERE id=? AND public=1 AND status='published'",(cid,)).fetchone()
    if not course:return "الدورة غير موجودة",404
    learner_key=session.get("academy_learner_key") or uid("learner")
    session["academy_learner_key"]=learner_key
    name=(request.form.get("name") or "متعلم AHVT").strip(); email=(request.form.get("email") or "").strip()
    row=db.execute("SELECT id FROM academy_enrollments WHERE course_id=? AND learner_key=?",(cid,learner_key)).fetchone()
    if not row:
        eid=uid("enroll"); db.execute("INSERT INTO academy_enrollments(id,course_id,learner_key,learner_name,learner_email,enrolled_at,last_seen) VALUES(?,?,?,?,?,?,?)",(eid,cid,learner_key,name,email,now_iso(),now_iso()))
    else:
        eid=row["id"]; db.execute("UPDATE academy_enrollments SET learner_name=?,learner_email=?,last_seen=? WHERE id=?",(name,email,now_iso(),eid))
    db.commit()
    if int(course["is_paid"] or 0) and float(course["price_iqd"] or 0)>0:
        flash("هذه الدورة مدفوعة. لا يتم تفعيلها إلا بعد رفع إثبات الدفع ومراجعته من الإدارة.","ok")
        return redirect(url_for("academy_payment",cid=cid))
    return redirect(url_for("academy_course_public",cid=cid))

@app.route("/academy/payment/<cid>", methods=["GET","POST"])
def academy_payment(cid):
    db=get_db(); course=db.execute("SELECT * FROM academy_courses WHERE id=? AND public=1 AND status='published'",(cid,)).fetchone()
    if not course:return "الدورة غير موجودة",404
    learner_key=session.get("academy_learner_key")
    if not learner_key:
        return redirect(url_for("academy_course_public",cid=cid))
    en=db.execute("SELECT * FROM academy_enrollments WHERE course_id=? AND learner_key=?",(cid,learner_key)).fetchone()
    if not en:
        return redirect(url_for("academy_course_public",cid=cid))
    existing=db.execute("SELECT * FROM academy_payments WHERE course_id=? AND learner_key=? ORDER BY created_at DESC LIMIT 1",(cid,learner_key)).fetchone()
    if request.method=="POST":
        proof=request.files.get("proof")
        if not proof or not proof.filename:
            flash("ارفع صورة الوصل أو ملف PDF لإثبات الدفع.","error"); return redirect(url_for("academy_payment",cid=cid))
        ext=os.path.splitext(secure_filename(proof.filename))[1].lower()
        if ext not in {".jpg",".jpeg",".png",".webp",".pdf"}:
            flash("المسموح: JPG أو PNG أو WEBP أو PDF.","error"); return redirect(url_for("academy_payment",cid=cid))
        namef=f"payment-{uuid.uuid4().hex[:14]}{ext}"; proof.save(os.path.join(UPLOAD_DIR,namef))
        path=f"uploads/{namef}"
        db.execute("INSERT INTO academy_payments(id,course_id,enrollment_id,learner_key,learner_name,learner_email,amount,currency,proof_path,learner_note,status,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",(uid("pay"),cid,en["id"],learner_key,en["learner_name"],en["learner_email"],float(course["price_iqd"] or 0),"IQD",path,request.form.get("note","").strip(),"pending",now_iso()))
        db.commit(); flash("تم رفع إثبات الدفع. سيتم تفعيل الدورة بعد مراجعة المالك أو الأدمن.","ok")
        return redirect(url_for("academy_course_public",cid=cid))
    payinfo={k:get_setting(k,"") for k in ["payment_name","payment_provider","payment_account","payment_instructions"]}
    return render_template("academy_payment.html",course=course,enrollment=en,existing=existing,payinfo=payinfo)

@app.route("/academy/payments")
def academy_my_payments():
    learner_key=session.get("academy_learner_key")
    if not learner_key:return redirect(url_for("academy_public"))
    db=get_db(); payments=db.execute("SELECT p.*,c.title FROM academy_payments p JOIN academy_courses c ON c.id=p.course_id WHERE p.learner_key=? ORDER BY p.created_at DESC",(learner_key,)).fetchall()
    return render_template("academy_payments.html",payments=payments)

@app.route("/admin/academy/payments", methods=["GET","POST"])
def academy_payments_admin():
    if not has_permission("academy_manage"): return render_template("403.html",permission="academy_manage"),403
    db=get_db()
    if request.method=="POST":
        pid=request.form.get("payment_id"); status=request.form.get("status"); note=request.form.get("admin_note","").strip()
        if status not in {"approved","rejected","pending"}: return "حالة غير صحيحة",400
        pmt=db.execute("SELECT * FROM academy_payments WHERE id=?",(pid,)).fetchone()
        if not pmt:return "طلب الدفع غير موجود",404
        db.execute("UPDATE academy_payments SET status=?,admin_note=?,reviewed_at=?,reviewed_by=? WHERE id=?",(status,note,now_iso(),current_user()["id"],pid))
        if status=="approved":
            db.execute("UPDATE academy_enrollments SET last_seen=? WHERE id=?",(now_iso(),pmt["enrollment_id"]))
        db.commit(); flash("تم تحديث طلب الدفع", "ok"); return redirect(url_for("academy_payments_admin"))
    payments=db.execute("SELECT p.*,c.title FROM academy_payments p JOIN academy_courses c ON c.id=p.course_id ORDER BY CASE p.status WHEN 'pending' THEN 0 ELSE 1 END,p.created_at DESC").fetchall()
    return render_template("academy_payments_admin.html",payments=payments)

@app.route("/academy/learn/<cid>/<lid>")
def academy_learn(cid,lid):
    db=get_db(); course=db.execute("SELECT * FROM academy_courses WHERE id=? AND public=1 AND status='published'",(cid,)).fetchone(); lesson=db.execute("SELECT * FROM academy_lessons WHERE id=? AND course_id=?",(lid,cid)).fetchone()
    if not course or not lesson:return "المحتوى غير موجود",404
    if int(course["is_paid"] or 0):
        learner_key=session.get("academy_learner_key")
        if not learner_key:return redirect(url_for("academy_course_public",cid=cid))
        approved=db.execute("SELECT 1 FROM academy_payments p JOIN academy_enrollments e ON e.id=p.enrollment_id WHERE p.course_id=? AND p.learner_key=? AND p.status='approved' LIMIT 1",(cid,learner_key)).fetchone()
        if not approved:
            flash("هذه الدورة مدفوعة ولم يتم تفعيلها بعد. ارفع إثبات الدفع وانتظر مراجعة الإدارة.","error")
            return redirect(url_for("academy_payment",cid=cid))
    quizzes=db.execute("SELECT * FROM academy_quizzes WHERE lesson_id=? ORDER BY sort_order",(lid,)).fetchall()
    resources=db.execute("SELECT * FROM academy_resources WHERE lesson_id=? ORDER BY created_at DESC",(lid,)).fetchall()
    return render_template("academy_lesson.html",course=course,lesson=lesson,quizzes=quizzes,resources=resources)

@app.route("/academy/lesson/<cid>/<lid>/complete", methods=["POST"])
def academy_lesson_complete(cid,lid):
    learner_key=session.get("academy_learner_key")
    if not learner_key:return redirect(url_for("academy_course_public",cid=cid))
    db=get_db(); en=db.execute("SELECT * FROM academy_enrollments WHERE course_id=? AND learner_key=?",(cid,learner_key)).fetchone()
    if not en:return redirect(url_for("academy_course_public",cid=cid))
    db.execute("INSERT INTO academy_progress(id,enrollment_id,lesson_id,completed,completed_at) VALUES(?,?,?,?,?) ON CONFLICT(enrollment_id,lesson_id) DO UPDATE SET completed=1,completed_at=excluded.completed_at",(uid("prog"),en["id"],lid,1,now_iso()))
    total=db.execute("SELECT COUNT(*) c FROM academy_lessons WHERE course_id=?",(cid,)).fetchone()["c"] or 1; done=db.execute("SELECT COUNT(*) c FROM academy_progress WHERE enrollment_id=? AND completed=1",(en["id"],)).fetchone()["c"]; pct=round(done*100/total,1)
    completed_at=now_iso() if pct>=100 else None
    db.execute("UPDATE academy_enrollments SET progress=?,last_seen=?,completed_at=? WHERE id=?",(pct,now_iso(),completed_at,en["id"])); db.commit()
    flash("تم تسجيل إكمال الدرس" if pct<100 else "مبروك! أكملت الدورة", "ok")
    return redirect(url_for("academy_course_public",cid=cid))

@app.route("/academy/quiz/<cid>/<lid>", methods=["POST"])
def academy_quiz_submit(cid,lid):
    learner_key=session.get("academy_learner_key")
    if not learner_key:return redirect(url_for("academy_learn",cid=cid,lid=lid))
    db=get_db(); en=db.execute("SELECT * FROM academy_enrollments WHERE course_id=? AND learner_key=?",(cid,learner_key)).fetchone()
    qs=db.execute("SELECT * FROM academy_quizzes WHERE lesson_id=? ORDER BY sort_order",(lid,)).fetchall(); score=0
    for q in qs:
        try: score += float(q["points"] or 1) if int(request.form.get("q_"+q["id"],-1))==int(q["answer_index"]) else 0
        except Exception: pass
    if en: db.execute("UPDATE academy_progress SET score=? WHERE enrollment_id=? AND lesson_id=?",(score,en["id"],lid))
    db.commit(); flash(f"نتيجتك: {score:g} من {sum(float(q['points'] or 1) for q in qs):g}","ok")
    return redirect(url_for("academy_learn",cid=cid,lid=lid))

@app.route("/academy/media")
def academy_media():
    db=get_db(); recordings=db.execute("SELECT * FROM live_sessions WHERE recording_path IS NOT NULL ORDER BY started_at DESC LIMIT 100").fetchall(); episodes=db.execute("SELECT * FROM podcast_episodes WHERE status='published' ORDER BY published_at DESC LIMIT 100").fetchall(); return render_template("academy_media.html",recordings=recordings,episodes=episodes)

@app.route("/admin/academy", methods=["GET","POST"])
def academy_admin():
    if not has_permission("academy_manage"): return render_template("403.html",permission="academy_manage"),403
    db=get_db()
    if request.method=="POST":
        action=request.form.get("action")
        if action=="course":
            cover=save_image_as_jpg(request.files.get("cover"),"academy")
            cid=uid("course"); db.execute("INSERT INTO academy_courses(id,title,description,instructor,level,duration,cover_image,public,status,created_at,updated_at,is_paid,price_iqd,payment_required) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(cid,request.form.get("title"),request.form.get("description"),request.form.get("instructor"),request.form.get("level"),request.form.get("duration"),cover,1 if request.form.get("public") else 0,"published",now_iso(),now_iso(),1 if request.form.get("is_paid") else 0,float(request.form.get("price_iqd") or 0),1 if request.form.get("is_paid") else 0))
        elif action=="lesson":
            lesson_id=uid("lesson")
            video_path=save_uploaded_video(request.files.get("video_file"),"academy-video") if request.files.get("video_file") else ""
            db.execute("INSERT INTO academy_lessons(id,course_id,title,content,video,duration,sort_order,created_at) VALUES(?,?,?,?,?,?,?,?)",(lesson_id,request.form.get("course_id"),request.form.get("title"),request.form.get("content"),video_path,request.form.get("duration"),int(request.form.get("sort_order") or 0),now_iso()))
            for f in request.files.getlist("resource_files"):
                if not f or not f.filename: continue
                ext=os.path.splitext(secure_filename(f.filename))[1].lower()
                if ext in {".jpg",".jpeg",".png",".webp",".gif"}:
                    path=save_image_as_jpg(f,"academy-resource"); rtype="image"
                elif ext in {".mp4",".webm",".mov",".m4v",".avi"}:
                    path=save_uploaded_video(f,"academy-resource"); rtype="video"
                else: continue
                if path:
                    db.execute("INSERT INTO academy_resources(id,course_id,lesson_id,title,file_path,url,resource_type,created_at) VALUES(?,?,?,?,?,?,?,?)",(uid("ares"),request.form.get("course_id"),lesson_id,secure_filename(f.filename),path,"",rtype,now_iso()))
        elif action=="quiz":
            opts=[x.strip() for x in request.form.get("options","").split("|") if x.strip()]; db.execute("INSERT INTO academy_quizzes(id,lesson_id,question,options_json,answer_index,points,sort_order) VALUES(?,?,?,?,?,?,?)",(uid("quiz"),request.form.get("lesson_id"),request.form.get("question"),json.dumps(opts,ensure_ascii=False),int(request.form.get("answer_index") or 0),float(request.form.get("points") or 1),int(request.form.get("sort_order") or 0)))
        db.commit(); flash("تم حفظ الأكاديمية", "ok"); return redirect(url_for("academy_admin"))
    courses=db.execute("SELECT * FROM academy_courses ORDER BY created_at DESC").fetchall(); lessons=db.execute("SELECT * FROM academy_lessons ORDER BY course_id,sort_order").fetchall(); return render_template("academy_admin.html",courses=courses,lessons=lessons)

@app.route("/calendar")
def calendar_public():
    db=get_db(); events=db.execute("SELECT * FROM events ORDER BY date,time LIMIT 200").fetchall(); lives=db.execute("SELECT * FROM live_sessions WHERE scheduled_date IS NOT NULL ORDER BY scheduled_date,scheduled_time LIMIT 100").fetchall(); eps=db.execute("SELECT * FROM podcast_episodes WHERE status='published' ORDER BY recorded_at DESC LIMIT 100").fetchall()
    return render_template("calendar_public.html",events=events,lives=lives,episodes=eps)

@app.route("/media-center")
def media_center_public():
    db=get_db(); media=db.execute("SELECT * FROM media WHERE public=1 ORDER BY created_at DESC").fetchall(); return render_template("media_center_public.html",media=media)

@app.route("/equipment")
def equipment_public():
    db=get_db(); equipment=db.execute("SELECT * FROM podcast_equipment ORDER BY category,name").fetchall(); loans=db.execute("SELECT equipment_id,status,issued_at FROM equipment_loans WHERE status='issued'").fetchall(); active={r['equipment_id'] for r in loans}
    return render_template("equipment_public.html",equipment=equipment,active_ids=active)

@app.route("/admin/equipment", methods=["GET","POST"])
def equipment_admin():
    if not has_permission("equipment"): return render_template("403.html",permission="equipment"),403
    db=get_db()
    if request.method=="POST":
        action=request.form.get("action")
        if action=="add":
            eid=uid("eq"); db.execute("INSERT INTO podcast_equipment(id,category,name,quantity,status,notes) VALUES(?,?,?,?,?,?)",(eid,request.form.get("category"),request.form.get("name"),int(request.form.get("quantity") or 1),request.form.get("status","available"),request.form.get("notes"))); db.commit(); db.execute("UPDATE podcast_equipment SET qr_path=? WHERE id=?",(_action_qr("equipment",eid,"equipment"),eid))
        elif action=="loan":
            eq=request.form.get("equipment_id"); mid=request.form.get("member_id")
            db.execute("INSERT INTO equipment_loans(id,equipment_id,member_id,issued_at,status,notes) VALUES(?,?,?,?,?,?)",(uid("loan"),eq,mid,now_iso(),"issued",request.form.get("notes")))
        elif action=="return":
            db.execute("UPDATE equipment_loans SET status='returned',returned_at=? WHERE id=?",(now_iso(),request.form.get("loan_id")))
        db.commit(); flash("تم تحديث المعدات","ok"); return redirect(url_for("equipment_admin"))
    equipment_rows=db.execute("SELECT * FROM podcast_equipment ORDER BY category,name").fetchall()
    for e in equipment_rows:
        if not e["qr_path"]:
            db.execute("UPDATE podcast_equipment SET qr_path=? WHERE id=?",(_action_qr("equipment",e["id"],"equipment"),e["id"]))
    db.commit()
    equipment_rows=db.execute("SELECT * FROM podcast_equipment ORDER BY category,name").fetchall()
    return render_template("equipment_admin.html",equipment=equipment_rows,loans=db.execute("SELECT l.*,e.name equipment_name,m.name member_name FROM equipment_loans l LEFT JOIN podcast_equipment e ON e.id=l.equipment_id LEFT JOIN members m ON m.id=l.member_id ORDER BY l.issued_at DESC LIMIT 100").fetchall(),members=db.execute("SELECT id,name FROM members ORDER BY name").fetchall())

@app.route("/admin/equipment/checkout/<mid>", methods=["POST"])
def equipment_checkout_qr(mid):
    if not has_permission("equipment"): return render_template("403.html",permission="equipment"),403
    db=get_db(); member=db.execute("SELECT id,name FROM members WHERE id=?",(mid,)).fetchone(); eq=request.form.get("equipment_id")
    if not member:return "العضو غير موجود",404
    db.execute("INSERT INTO equipment_loans(id,equipment_id,member_id,issued_at,status,notes) VALUES(?,?,?,?,?,?)",(uid("loan"),eq,mid,now_iso(),"issued","تم الاستلام عبر QR البطاقة")); db.commit(); log_action("Equipment checkout via member QR",member["name"]); flash(f"تم تسليم المعدة إلى {member['name']}","ok"); return redirect(url_for("equipment_admin"))

@app.route("/finance")
def finance_public():
    db=get_db(); rows=db.execute("SELECT * FROM finance_entries WHERE public=1 ORDER BY entry_date DESC,created_at DESC").fetchall(); total_in=sum(float(r['amount'] or 0) for r in rows if r['entry_type']=='income'); total_out=sum(float(r['amount'] or 0) for r in rows if r['entry_type']=='expense')
    return render_template("finance_public.html",rows=rows,total_in=total_in,total_out=total_out,balance=total_in-total_out)

@app.route("/admin/finance", methods=["GET","POST"])
def finance_admin():
    if not has_permission("finance"): return render_template("403.html",permission="finance"),403
    db=get_db()
    if request.method=="POST":
        db.execute("INSERT INTO finance_entries(id,entry_type,title,amount,category,entry_date,description,public,created_at) VALUES(?,?,?,?,?,?,?,?,?)",(uid("fin"),request.form.get("entry_type"),request.form.get("title"),float(request.form.get("amount") or 0),request.form.get("category"),request.form.get("entry_date") or iraq_today(),request.form.get("description"),1 if request.form.get("public") else 0,now_iso())); db.commit(); flash("تم حفظ الحركة المالية","ok"); return redirect(url_for("finance_admin"))
    return render_template("finance_admin.html",rows=db.execute("SELECT * FROM finance_entries ORDER BY entry_date DESC,created_at DESC").fetchall())

@app.route("/assistant")
def assistant_public():
    db=get_db(); q=request.args.get("q","").strip(); members=db.execute("SELECT COUNT(*) c FROM members").fetchone()['c']; admins=db.execute("SELECT COUNT(*) c FROM administrators").fetchone()['c']; initiatives=db.execute("SELECT COUNT(*) c FROM initiatives").fetchone()['c']; events=db.execute("SELECT COUNT(*) c FROM events").fetchone()['c']; hours=db.execute("SELECT COALESCE(SUM(hours),0) h FROM attendance").fetchone()['h']; answer="أهلاً بك في مساعد AHVT. اسأل عن أعضاء الفريق، المبادرات، الفعاليات، الساعات أو الإحصائيات العامة."
    if q:
        ql=q.lower()
        if "عضو" in q: answer=f"عدد أعضاء الفريق المسجلين حاليًا: {members}."
        elif "إداري" in q or "اداري" in q: answer=f"عدد الإداريين المسجلين: {admins}."
        elif "مباد" in q: answer=f"عدد المبادرات المسجلة: {initiatives}."
        elif "فعالي" in q: answer=f"عدد الفعاليات المسجلة: {events}."
        elif "ساع" in q: answer=f"إجمالي ساعات الحضور المسجلة: {float(hours or 0):.2f} ساعة."
        else: answer="أستطيع حاليًا إعطاء إحصائيات AHVT العامة، والبحث المتقدم يبقى داخل الإدارة لحماية البيانات."
    return render_template("assistant_public.html",answer=answer,q=q,stats={"members":members,"admins":admins,"initiatives":initiatives,"events":events,"hours":hours})

@app.route("/backup/export/full")
def backup_full_export():
    if not has_permission("backup"): return render_template("403.html",permission="backup"),403
    import tempfile, zipfile as _zipfile, sqlite3 as _sqlite3
    db=get_db(); tmpdir=tempfile.mkdtemp(prefix="ahvt_backup_"); dbcopy=os.path.join(tmpdir,"ahvt.sqlite")
    dest=_sqlite3.connect(dbcopy); db.backup(dest); dest.close()
    manifest={"version":26,"created_at":now_iso(),"team":"AHVT","includes":["database","uploads"]}
    manifest_path=os.path.join(tmpdir,"manifest.json"); open(manifest_path,"w",encoding="utf-8").write(json.dumps(manifest,ensure_ascii=False,indent=2))
    out=io.BytesIO()
    with _zipfile.ZipFile(out,"w",_zipfile.ZIP_DEFLATED) as z:
        z.write(dbcopy,"data/ahvt.sqlite"); z.write(manifest_path,"manifest.json")
        if os.path.isdir(UPLOAD_DIR):
            for base,_,files in os.walk(UPLOAD_DIR):
                for fn in files:
                    full=os.path.join(base,fn); z.write(full,os.path.join("uploads",os.path.relpath(full,UPLOAD_DIR)))
    shutil.rmtree(tmpdir,ignore_errors=True); out.seek(0); return send_file(out,mimetype="application/zip",as_attachment=True,download_name=f"AHVT-Full-Backup-{date.today().isoformat()}.zip")

# ============================================================ BACKUP ============================================================
@app.route("/backup/export")
def backup_export():
    """Export a consistent JSON snapshot without changing live data."""
    db = get_db()
    tables = [
        "users", "members", "administrators", "committees", "initiatives",
        "initiative_participants", "tasks", "attendance", "evaluations", "points",
        "audit_logs", "news", "pages", "events", "partners", "media", "goals",
        "approvals", "security_sessions", "site_sections", "nav_items", "role_permissions",
        "notifications", "applications", "live_sessions", "live_peers", "live_signals",
        "certificates", "honor_list", "attachments", "card_templates", "task_completions", "chat_conversations", "chat_messages", "chat_reactions", "podcast_shows", "podcast_guests", "podcast_episodes", "podcast_clips", "podcast_equipment", "temporary_volunteers", "academy_courses", "academy_lessons", "academy_quizzes", "academy_resources", "academy_enrollments", "academy_progress", "academy_payments", "finance_entries", "equipment_loans", "live_source_requests", "site_notifications"
    ]
    data = {"version": 26, "created_at": now_iso(), "tables": {}}
    with db:
        for t in tables:
            rows = db.execute(f"SELECT * FROM {t}").fetchall()
            data["tables"][t] = [dict(r) for r in rows]
        data["tables"]["settings"] = [dict(r) for r in db.execute("SELECT * FROM settings").fetchall()]
    payload = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
    buf = io.BytesIO(payload)
    buf.seek(0)
    fname = f"AHVT-Full-Backup-{date.today().isoformat()}.json"
    return send_file(buf, mimetype="application/json", as_attachment=True, download_name=fname)


@app.route("/backup/import", methods=["POST"])
def backup_import():
    file = request.files.get("backup_file")
    if not file or not file.filename.lower().endswith(".json"):
        flash("الملف غير صالح — اختر نسخة AHVT بصيغة JSON", "error")
        return redirect(url_for("reports_page"))

    try:
        data = json.load(file.stream)
        tables_data = data.get("tables", data)
        if not isinstance(tables_data, dict):
            raise ValueError("invalid backup structure")

        db = get_db()
        table_order = [
            "live_signals", "live_peers", "initiative_participants", "attendance",
            "evaluations", "points", "notifications", "security_sessions", "audit_logs",
            "approvals", "tasks", "task_completions", "chat_conversations", "chat_messages", "chat_reactions", "media", "events", "goals", "certificates", "applications",
            "news", "pages", "site_sections", "nav_items", "role_permissions", "partners", "attachments",
            "card_templates", "podcast_shows", "podcast_guests", "podcast_episodes", "podcast_clips", "podcast_equipment", "academy_courses", "academy_lessons", "finance_entries", "equipment_loans", "site_notifications",
            "initiatives", "committees", "administrators", "members", "users", "settings"
        ]

        # Never delete data until the backup has been parsed and validated.
        # The whole restore is one transaction: any failure rolls everything back.
        db.execute("BEGIN IMMEDIATE")
        try:
            for table in table_order:
                if table not in tables_data:
                    continue
                db.execute(f"DELETE FROM {table}")

            restored = 0
            for table in table_order:
                rows = tables_data.get(table)
                if rows is None:
                    continue
                if table == "settings":
                    for row in rows:
                        if not isinstance(row, dict) or "key" not in row:
                            raise ValueError("invalid settings row")
                        db.execute("INSERT INTO settings(key,value) VALUES(?,?)", (row.get("key"), row.get("value")))
                        restored += 1
                    continue

                if not isinstance(rows, list):
                    raise ValueError(f"invalid table: {table}")

                # Restore only columns that actually exist in the current DB.
                # This makes V11 compatible with older backups and preserves
                # defaults/migrations for columns introduced later.
                cols = {r["name"] for r in db.execute(f"PRAGMA table_info({table})").fetchall()}
                for row in rows:
                    if not isinstance(row, dict):
                        raise ValueError(f"invalid row in {table}")
                    keys = [k for k in row.keys() if k in cols]
                    if not keys:
                        continue
                    placeholders = ",".join("?" for _ in keys)
                    db.execute(
                        f"INSERT INTO {table}({','.join(keys)}) VALUES({placeholders})",
                        [row.get(k) for k in keys]
                    )
                    restored += 1

            db.commit()
        except Exception:
            db.rollback()
            raise

        flash(f"تم استيراد النسخة الاحتياطية بنجاح ({restored} سجل)", "ok")
    except Exception as exc:
        app.logger.exception("Backup import failed: %s", exc)
        flash("فشل استيراد النسخة الاحتياطية. لم يتم حذف أو تغيير بياناتك.", "error")
    return redirect(url_for("public_home"))


if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=os.environ.get("FLASK_DEBUG", "0") == "1")
else:
    init_db()
